# -*- coding: utf-8 -*-
"""IR 팩트시트에서 뽑은 값을 data_sheet_{기간}.csv 에 병합한다.

사용:
  python ir_fill.py --period 2606                 # 병합 결과를 새 CSV로 저장
  python ir_fill.py --period 2606 --dry           # 저장 없이 출력만
  python ir_fill.py --period 2606 --inplace       # 원본 CSV 덮어쓰기

팩트시트 경로는 FACTSHEETS 에 회사×연도로 등록한다.
IR 값은 기존 CSV 값이 비어 있을 때만 채우며, 이미 값이 있으면 건드리지 않는다
(--overwrite 로 반대 동작).
"""
import os, csv, sys, argparse
sys.stdout.reconfigure(encoding='utf-8')
from ir_parse import extract_company, SPECS

def _ir_dir():
    """팩트시트가 있는 폴더를 찾는다.

    IR_DIR 환경변수가 최우선. 없으면 흔한 위치를 순서대로 본다.
    브라우저로 받으면 대개 사용자 다운로드 폴더에 떨어지므로 그쪽을 먼저 본다.
    """
    env = os.environ.get('IR_DIR')
    if env:
        return env
    for cand in (os.path.join(os.getcwd(), 'ir'),
                 os.path.join(os.path.expanduser('~'), 'Downloads'),
                 os.path.join(os.path.expanduser('~'), '다운로드'),
                 '/mnt/user-data/uploads/Downloads'):
        if os.path.isdir(cand):
            return cand
    return os.path.join(os.path.expanduser('~'), 'Downloads')


DL = _ir_dir()

# 팩트시트는 `data/{회사명}/IR/{기간}/{회사명}_{기간}_팩트시트.xlsx` 에 둔다
# (ir_store.py 가 다운로드 폴더에서 이 구조로 정리해준다).
# 분기 팩트시트에는 그 연도의 앞선 분기가 함께 들어 있으므로, 해당 기간 파일이 없으면
# 같은 연도의 더 나중 분기 파일을 쓴다.
IR_ROOT = os.path.join('data', '{comp}', 'IR')
KINDS = ('팩트시트', '팩트북')


def find_factsheet(company, period):
    """(경로, 실제사용기간) — 없으면 (None, None)"""
    root = IR_ROOT.format(comp=company)
    if not os.path.isdir(root):
        return None, None
    year, mm = period[:2], int(period[2:])
    # 해당 기간 → 같은 연도의 이후 분기 순으로 찾는다
    candidates = [period] + [f'{year}{m:02d}' for m in (3, 6, 9, 12) if m > mm]
    for pd in candidates:
        d = os.path.join(root, pd)
        if not os.path.isdir(d):
            continue
        for kind in KINDS:
            for f in sorted(os.listdir(d)):
                if kind in f and f.lower().endswith(('.xlsx', '.xls')):
                    return os.path.join(d, f), pd
    return None, None


SHORT = {'미래에셋생명': '미래', '삼성생명': '삼성', '한화생명': '한화',
         '동양생명': '동양', '삼성화재': '삼성화재', '현대해상': '현대해상',
         '메리츠화재': '메리츠화재', 'KB손해보험': 'KB손보', 'KB생명': 'KB라이프',
         '교보생명': '교보', '신한라이프생명': '신한라이프', 'DB손해보험': 'DB손보'}



def resolve_short(ws, row, col=3, depth=4):
    """회사약칭 칸이 '=C108' 처럼 다른 행을 참조하는 경우까지 풀어서 읽는다.

    작업 파일에서 같은 회사의 뒷 기간 행은 회사명을 위 행 참조 수식으로 넣어두는데,
    openpyxl 은 수식 문자열 그대로를 돌려주므로 그냥 비교하면 그 행 전체를 놓친다.
    """
    import re
    v = ws.cell(row, col).value
    for _ in range(depth):
        if not isinstance(v, str) or not v.startswith('='):
            break
        m = re.fullmatch(r'=\$?([A-Z]{1,3})\$?(\d+)', v.strip())
        if not m:
            return None
        from openpyxl.utils import column_index_from_string
        v = ws.cell(int(m.group(2)), column_index_from_string(m.group(1))).value
    return str(v).strip() if v not in (None, '') else None


def find_disclosure(company, period):
    """정기경영공시 PDF 경로 (없으면 None)"""
    d = os.path.join(IR_ROOT.format(comp=company), period)
    if not os.path.isdir(d):
        return None
    for f in sorted(os.listdir(d)):
        if '경영공시' in f and f.lower().endswith('.pdf'):
            return os.path.join(d, f)
    return None


def collect_disclosure(period):
    """정기경영공시 PDF에서 뽑은 {회사약칭: {col: 값}}

    금감원 정기경영공시는 모든 보험사가 같은 목차·같은 표 제목을 쓰므로
    팩트시트와 달리 회사별 매핑이 필요 없다.
    """
    from ir_pdf import parse_disclosure
    year = 2000 + int(period[:2])
    out, notes = {}, []
    for comp, short in SHORT.items():
        pdf = find_disclosure(comp, period)
        if not pdf:
            continue
        try:
            vals, _src = parse_disclosure(pdf, year)
        except Exception as e:                       # 목차가 다른 PDF 도 있다
            notes.append(f'{comp} 경영공시 파싱 실패: {e}')
            continue
        if vals:
            out[short] = vals
            notes.append(f'{comp}: 경영공시 PDF에서 {len(vals)}개 컬럼')
    return out, notes


def collect(period):
    """{회사약칭: {col: 값}}"""
    year = 2000 + int(period[:2])
    out, notes = {}, []
    for comp in SPECS:
        full, used = find_factsheet(comp, period)
        if not full:
            notes.append(f'{comp}: 팩트시트 없음 (data/{comp}/IR/{period}/ 를 확인하세요)')
            continue
        if used != period:
            notes.append(f'{comp}: {period} 자료가 없어 {used} 팩트시트로 대신 읽음')
        vals, miss = extract_company(comp, full, period)

        # Col80(전기말 CSM)은 직전 연도 팩트시트의 기말 CSM(Col75)이 정확하다.
        # 당해 팩트시트에는 전년도 '연간' 열이 없고 동기 분기만 실리는 경우가 있어
        # (삼성화재 26.2Q 팩트시트가 그렇다) 직전 연도 파일에서 다시 읽는다.
        prior_path, _pu = find_factsheet(comp, f'{(year - 1) % 100:02d}12')
        if prior_path:
            pv, _ = extract_company(comp, prior_path,
                                    f'{(year - 1) % 100:02d}12')
            if 75 in pv:
                vals[80] = pv[75]
                miss = [m for m in miss if m[0] != 80]   # 직전 연도 파일로 해결됨

        out[SHORT[comp]] = vals
        for c, l, why in miss:
            notes.append(f'{comp} Col{c}: 추출실패({l}, {why})')

    # 정기경영공시 PDF 는 팩트시트가 없는 회사(교보 등)를 메운다.
    # 팩트시트 값이 더 정밀하므로 이미 뽑힌 컬럼은 덮어쓰지 않는다.
    dvals, dnotes = collect_disclosure(period)
    for short, cols in dvals.items():
        tgt = out.setdefault(short, {})
        for c, v in cols.items():
            tgt.setdefault(c, v)
    notes += dnotes
    return out, notes


def merge(period, values, overwrite=False):
    src = os.path.join('data', f'data_sheet_{period}.csv')
    rows = list(csv.DictReader(open(src, encoding='utf-8-sig')))
    fields = list(rows[0].keys())
    added = filled = 0
    for col in sorted({c for v in values.values() for c in v}):
        name = f'Col{col}'
        if name not in fields:
            fields.append(name)
            for r in rows:
                r[name] = ''
            added += 1
    for r in rows:
        short = r['key'][4:]
        for col, val in values.get(short, {}).items():
            name = f'Col{col}'
            if r.get(name) in (None, '') or overwrite:
                r[name] = repr(val)
                filled += 1
    # 컬럼 번호 순 정렬 (key, Col2, Col4, ...)
    def order(f):
        return (0, 0) if f == 'key' else (1, int(f[3:]))
    fields.sort(key=order)
    return rows, fields, added, filled


# 채울 엑셀 파일. WORK_XLSX 환경변수나 --excel-path 로 바꾼다.
# (작업 파일이 해마다 바뀌므로 하드코딩하지 않는다)
EXCEL = os.environ.get('WORK_XLSX', 'DATA_작업_빈칸.xlsx')
IR_FILL_COLOR = 'CCE5FF'   # 연한 파랑 — IR 자료에서 온 값임을 눈으로 구분


def write_excel(period, values, overwrite=False, path=EXCEL):
    """DATA 시트에 직접 값을 써넣는다.

    parse_annual.py 가 1차로 채운 뒤 남은 빈칸을 IR 값으로 메우는 용도라
    기본은 **빈칸만** 채운다. IR에서 온 값은 연한 파랑으로 칠해 출처가 보이게 한다.
    """
    import openpyxl
    from openpyxl.styles import PatternFill

    if not os.path.exists(path):
        return None, f'{path} 가 없습니다'
    wb = openpyxl.load_workbook(path)          # 수식 유지 (data_only 쓰지 않는다)
    ws = wb['DATA']

    col_to_excel = {}
    for ec in range(1, ws.max_column + 1):
        try:
            col_to_excel[int(ws.cell(1, ec).value)] = ec
        except (TypeError, ValueError):
            pass

    short_to_row = {}
    for r in range(6, ws.max_row + 1):
        if str(ws.cell(r, 2).value) != str(period):
            continue
        short = resolve_short(ws, r)
        if short:
            short_to_row[short] = r

    fill = PatternFill('solid', start_color=IR_FILL_COLOR, end_color=IR_FILL_COLOR)
    written, skipped = 0, 0
    for short, cols in values.items():
        row = short_to_row.get(short)
        if not row:
            continue
        for col, val in cols.items():
            ec = col_to_excel.get(col)
            if not ec:
                continue
            cell = ws.cell(row, ec)
            if cell.value not in (None, '') and not overwrite:
                skipped += 1
                continue
            cell.value = val
            cell.fill = fill
            written += 1
    wb.save(path)
    return (written, skipped), None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--period', required=True)
    ap.add_argument('--dry', action='store_true')
    ap.add_argument('--inplace', action='store_true')
    ap.add_argument('--overwrite', action='store_true')
    ap.add_argument('--excel', action='store_true',
                    help='작업 엑셀에도 직접 채운다 (IR 값은 연한 파랑)')
    ap.add_argument('--excel-path', default=EXCEL,
                    help='채울 엑셀 파일 (기본: $WORK_XLSX 또는 DATA_작업_빈칸.xlsx)')
    a = ap.parse_args()

    values, notes = collect(a.period)
    print(f'=== {a.period} IR 팩트시트 추출 ===')

    for short in sorted(values):
        cols = values[short]
        print(f'  {short:<8} {len(cols)}개 컬럼: ' +
              ', '.join(f'Col{c}={v:,.0f}' for c, v in sorted(cols.items())))
    for n in notes:
        print('  ! ' + n)
    if a.dry:
        return
    rows, fields, added, filled = merge(a.period, values, a.overwrite)
    dst = (os.path.join('data', f'data_sheet_{a.period}.csv') if a.inplace
           else os.path.join('data', f'data_sheet_{a.period}_ir.csv'))
    with open(dst, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)
    print(f'\n저장: {dst}  (신규컬럼 {added}개, 채운 셀 {filled}개)')

    if a.excel:
        res, err = write_excel(a.period, values, a.overwrite, a.excel_path)
        if err:
            print(f'엑셀 채우기 건너뜀: {err}')
        else:
            w, sk = res
            print(f'엑셀: {a.excel_path} 에 {w}셀 기록'
                  f'{f" (이미 값이 있어 건너뛴 셀 {sk}개)" if sk else ""}'
                  f' — IR 값은 연한 파랑으로 표시됩니다')


if __name__ == '__main__':
    main()
