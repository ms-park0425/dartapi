# -*- coding: utf-8 -*-
"""DATA 시트 안에서 컬럼끼리 항상 성립하는 항등식을 찾고, 그걸로 빈칸을 메운다.

IR 자료를 아무리 뒤져도 안 나오는 칸이라도 같은 행의 다른 칸으로
계산되는 경우가 많다(합계=부분합, 예실차=예상−실제, 비율=b÷a 등).
정답 시트 전 기간·전 회사에서 예외 없이 성립하는 관계만 채택한다.
"""
import sys, itertools, argparse
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

ANSWER = '202512_동업사 공시비교_DATA_작업_260320.xlsx'
REL_TOL = 2e-3
MIN_ROWS = 20        # 이만큼은 검증돼야 우연이 아니다
MATERIAL = 0.02      # 더하는 항이 대상의 2% 도 안 되면 '있으나 마나' → 잡음


def load_rows(path=ANSWER, sheet='DATA'):
    ws = openpyxl.load_workbook(path, data_only=True)[sheet]
    c2e, names = {}, {}
    for ec in range(1, ws.max_column + 1):
        try:
            c2e[int(ws.cell(1, ec).value)] = ec
        except (TypeError, ValueError):
            pass
    for col, ec in c2e.items():
        parts = []
        for r in (3, 4, 5):
            v = None
            for c2 in range(ec, 0, -1):
                v = ws.cell(r, c2).value
                if v is not None:
                    break
            if v is not None:
                parts.append(str(v).replace('\n', ' ').strip())
        names[col] = ' > '.join(dict.fromkeys(parts))
    rows = []
    for r in range(6, ws.max_row + 1):
        p, s = ws.cell(r, 2).value, ws.cell(r, 3).value
        if not p or not s:
            continue
        rows.append(((str(p), str(s).strip()),
                     {c: ws.cell(r, e).value for c, e in c2e.items()
                      if isinstance(ws.cell(r, e).value, (int, float))}))
    return rows, names


def close(a, b):
    return abs(a - b) <= max(abs(a), abs(b), 1e-9) * REL_TOL


def apply_rule(rule, d):
    op = rule[0]
    try:
        if op == 'eq':
            return d[rule[1]]
        if op == 'neg':
            return -d[rule[1]]
        a, b = d[rule[1]], d[rule[2]]
        if op == '+':
            return a + b
        if op == '-':
            return a - b
        if op == '/':
            return a / b if b else None
    except KeyError:
        return None
    return None


def _check(rows, t, rule, material=True):
    """(검증 행 수, 위배 여부)"""
    n = 0
    for _, d in rows:
        if t not in d:
            continue
        exp = apply_rule(rule, d)
        if exp is None:
            continue
        n += 1
        if not close(d[t], exp):
            return n, True
        if material and rule[0] in '+-':
            # 더한 항이 실질적으로 기여하는지 (t ≈ a 인 사실상 복사 관계 배제)
            other = d[rule[2]]
            if abs(other) >= abs(d[t]) * MATERIAL:
                material = False
    return (0 if material and rule[0] in '+-' else n), False


def find(rows):
    allc = sorted({c for _, d in rows for c in d})

    def nonzero(c):
        vs = [d[c] for _, d in rows if c in d]
        return sum(1 for v in vs if abs(v) > 1e-9) >= max(MIN_ROWS, len(vs) * 0.5)
    terms = [c for c in allc if nonzero(c)]

    found = []
    for t in allc:
        best = None
        for s in terms:
            if s == t:
                continue
            for op in ('eq', 'neg'):
                n, bad = _check(rows, t, (op, s))
                if not bad and n >= MIN_ROWS:
                    best = (t, (op, s), n)
                    break
            if best:
                break
        if best:
            found.append(best)
            continue
        for a, b in itertools.combinations([c for c in terms if c != t], 2):
            hit = None
            for op in ('+', '-', '/'):
                for x, y in ((a, b), (b, a)):
                    if op == '+' and (x, y) != (a, b):
                        continue
                    n, bad = _check(rows, t, (op, x, y))
                    if not bad and n >= MIN_ROWS:
                        hit = (t, (op, x, y), n)
                        break
                if hit:
                    break
            if hit:
                found.append(hit)
                break
    return found


def fmt(rule):
    if rule[0] == 'eq':
        return f'Col{rule[1]}'
    if rule[0] == 'neg':
        return f'-Col{rule[1]}'
    return f'Col{rule[1]} {rule[0]} Col{rule[2]}'


if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('--answer', default=ANSWER)
    a = ap.parse_args()
    rows, names = load_rows(a.answer)
    print(f'행 {len(rows)}개')
    for t, rule, n in find(rows):
        print(f'Col{t:<4} = {fmt(rule):<22} ({n}행 검증)  {names.get(t,"")[:52]}')


# ------------------------------------------------------------ 빈칸 메우기
from ir_fill import resolve_short


def fill_blanks(path, sheet='DATA', rules=None, answer=ANSWER, dry=True,
                out=None, color='D9F2D0', periods=None, skip_cols=(31, 45, 46, 47,
                64, 72, 89, 90, 91)):
    """항등식으로 채울 수 있는 빈칸을 채운다. 채운 칸 목록을 돌려준다.

    periods 를 주면 그 기간만 손댄다. 이미 완성된 과거 기간까지 건드리면
    검산(CHECK) 컬럼이 한쪽만 채워져 0 이 아니게 되는 부작용이 있다.
    skip_cols 는 검산 전용 컬럼 — 계산해서 넣을 대상이 아니다.
    """
    from openpyxl.styles import PatternFill
    if rules is None:
        arows, _ = load_rows(answer)
        rules = {t: r for t, r, _ in find(arows)}
    wb = openpyxl.load_workbook(path)      # 수식 유지
    ws = wb[sheet]
    c2e = {}
    for ec in range(1, ws.max_column + 1):
        try:
            c2e[int(ws.cell(1, ec).value)] = ec
        except (TypeError, ValueError):
            pass
    fill = PatternFill('solid', fgColor=color)
    filled = []
    for r in range(6, ws.max_row + 1):
        p = ws.cell(r, 2).value
        s = resolve_short(ws, r)
        if not p or not s:
            continue
        if periods and str(p) not in periods:
            continue
        d = {c: ws.cell(r, e).value for c, e in c2e.items()
             if isinstance(ws.cell(r, e).value, (int, float))}
        for _ in range(4):        # 연쇄 적용 (A→B→C)
            new = 0
            for t, rule in rules.items():
                if t in d or t not in c2e or t in skip_cols:
                    continue
                v = apply_rule(rule, d)
                if v is None:
                    continue
                filled.append((str(p), str(s).strip(), t, v, fmt(rule)))
                if not dry:
                    cell = ws.cell(r, c2e[t])
                    cell.value = v
                    cell.fill = fill
                d[t] = v
                new += 1
            if not new:
                break
    if not dry:
        wb.save(out or path)
    return filled
