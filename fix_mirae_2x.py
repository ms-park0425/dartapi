# -*- coding: utf-8 -*-
"""미래에셋 2603·2606 의 CSM 2배 집계 잔재를 IR 팩트시트 기준으로 바로잡는다.

XBRL TypesOfContractsAxis 에 보험종류별·배당유무별 두 분류가 함께 실려
합산하면 정확히 2배가 되는 버그(parse_annual.py 에서 이미 수정됨)로 채워진 값이
2603·2606 행에 남아 있다. DART 를 다시 부를 수 없으므로 팩트시트 값으로 덮는다.
"""
import sys, argparse
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import PatternFill
from ir_parse import load
from ir_fill import resolve_short

FS = 'data/미래에셋생명/IR/2606/미래에셋생명_2606_팩트시트.xlsx'
BN = 1000.0                      # 팩트시트 CSM 시트는 십억원
FIX_COLOR = 'FFE0B2'             # 주황 — 손으로 바로잡은 칸

# 2배로 들어가 있던 BEL·RA (원래 값의 절반). 팩트시트에 같은 정의의 항목이 없어
# 직접 못 뽑고, 전 기간 추세와 팩트시트 잔여보장 BEL/RA 로 2배임을 확인했다.
HALVED = {'2603': {73: 23629496.0, 74: 419290.0},
          '2606': {73: 25576329.0, 74: 446976.0}}


def factsheet_csm():
    """CSM 시트에서 {기간: {컬럼: 백만원}} — 왼쪽 블록은 반기누계, 오른쪽은 2분기 단독"""
    ws = load(FS)['CSM']
    cum, q2 = 5, 10              # 열 위치(반기누계 / 2분기 단독)
    row = {'기시': 7, '신계약': 8, '이자': 13, '가정변경': 14, '상각': 15, '기말': 16}
    g = lambda r, c: ws.cell(r, c).value * BN
    h = {k: g(r, cum) for k, r in row.items()}
    s = {k: g(r, q2) for k, r in row.items()}
    return {
        # 1분기 = 반기누계 − 2분기 단독. 기말은 2분기의 '기시'가 곧 1분기 기말.
        '2603': {75: s['기시'], 76: h['신계약'] - s['신계약'],
                 77: -(h['상각'] - s['상각']), 78: h['가정변경'] - s['가정변경'],
                 79: h['이자'] - s['이자'], 80: h['기시']},
        '2606': {75: h['기말'], 76: h['신계약'], 77: -h['상각'],
                 78: h['가정변경'], 79: h['이자'], 80: h['기시']},
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx')
    ap.add_argument('--dry', action='store_true')
    a = ap.parse_args()

    want = factsheet_csm()
    wb = openpyxl.load_workbook(a.xlsx)
    ws = wb['DATA']
    c2e = {}
    for ec in range(1, ws.max_column + 1):
        try:
            c2e[int(ws.cell(1, ec).value)] = ec
        except (TypeError, ValueError):
            pass
    fill = PatternFill('solid', fgColor=FIX_COLOR)
    changed = 0
    for r in range(6, ws.max_row + 1):
        p = str(ws.cell(r, 2).value)
        if resolve_short(ws, r) != '미래' or p not in want:
            continue
        # Col73(BEL)·Col74(RA)는 팩트시트에 DATA 정의(순부채기준)와 같은 항목이
        # 없다. 원래 값이 정확히 2배인 것을 확인했으므로 그 절반을 박아 넣는다.
        # (÷2 를 두 번 돌리면 4분의 1이 되므로 절대값으로 둔다)
        for col, v in HALVED[p].items():
            cur = ws.cell(r, c2e[col]).value
            same = isinstance(cur, (int, float)) and abs(cur - v) <= max(abs(v), 1) * 0.002
            print(f'  {p} Col{col:<3} {(cur if isinstance(cur,(int,float)) else 0):>16,.0f} → {v:>16,.0f}  {"OK" if same else "고침(÷2)"}')
            if same:
                continue
            if not a.dry:
                ws.cell(r, c2e[col]).value = v
                ws.cell(r, c2e[col]).fill = fill
            changed += 1
        for col, v in sorted(want[p].items()):
            cur = ws.cell(r, c2e[col]).value
            same = isinstance(cur, (int, float)) and abs(cur - v) <= max(abs(v), 1) * 0.002
            mark = 'OK' if same else '고침'
            print(f'  {p} Col{col:<3} {(cur if isinstance(cur,(int,float)) else 0):>16,.0f} → {v:>16,.0f}  {mark}')
            if same:
                continue
            if not a.dry:
                ws.cell(r, c2e[col]).value = v
                ws.cell(r, c2e[col]).fill = fill
            changed += 1
    if not a.dry:
        wb.save(a.xlsx)
    print(f'\n바로잡은 칸: {changed}')


if __name__ == '__main__':
    main()
