# -*- coding: utf-8 -*-
"""CSM movement 검산으로 어긋난 칸을 찾아 바로잡는다.

    기시(80) + 신계약(76) + 이자부리(79) − 상각(77) + 조정(78) = 기말(75)

이 항등식은 정의상 반드시 성립한다. 안 맞으면 그 행 어딘가가 틀린 것이고,
어긋난 양이 어떤 항의 정확히 2배면 그 항의 부호가 뒤집힌 것이다.
정답 시트가 없는 기간(2603·2606 등)을 검증할 수 있는 유일한 자체 검산이다.
"""
import sys, argparse
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import PatternFill
from ir_fill import resolve_short, collect

MOVE = dict(end=75, new=76, amort=77, adj=78, interest=79, begin=80)
TOL = 0.002
FIX_COLOR = 'FFE0B2'


def check_row(v):
    """(계산된 기말, 상대오차)"""
    calc = v[80] + v[76] + v[79] - v[77] + v[78]
    return calc, (calc - v[75]) / max(abs(v[75]), 1)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx')
    ap.add_argument('--periods', default='2603,2606')
    ap.add_argument('--dry', action='store_true')
    a = ap.parse_args()
    periods = a.periods.split(',')

    ir = {p: collect(p)[0] for p in periods}
    wb = openpyxl.load_workbook(a.xlsx)
    ws = wb['DATA']
    c2e = {}
    for ec in range(1, ws.max_column + 1):
        try:
            c2e[int(ws.cell(1, ec).value)] = ec
        except (TypeError, ValueError):
            pass
    fill = PatternFill('solid', fgColor=FIX_COLOR)
    fixed = 0

    for r in range(6, ws.max_row + 1):
        p = str(ws.cell(r, 2).value)
        s = resolve_short(ws, r)
        if p not in periods or not s:
            continue
        g = lambda c: (ws.cell(r, c2e[c]).value
                       if c in c2e and isinstance(ws.cell(r, c2e[c]).value, (int, float))
                       else None)
        v = {c: g(c) for c in MOVE.values()}
        if any(x is None for x in v.values()):
            continue
        calc, d = check_row(v)
        if abs(d) < TOL:
            continue
        print(f'{p} {s}: 검산 {calc:,.0f} ≠ 기말 {v[75]:,.0f} ({d:+.2%})')

        # ① 어떤 항의 부호가 뒤집혔나 — 그 항의 2배만큼 어긋난다
        done = False
        for col in (78, 79, 76, 77):
            if abs(calc - 2 * v[col] - v[75]) <= max(abs(v[75]), 1) * TOL:
                print(f'   → Col{col} 부호 반전: {v[col]:,.0f} → {-v[col]:,.0f}')
                if not a.dry:
                    ws.cell(r, c2e[col]).value = -v[col]
                    ws.cell(r, c2e[col]).fill = fill
                fixed += 1
                done = True
                break
        if done:
            continue

        # ② 기말(75)만 틀린 경우 — IR 팩트시트가 검산값과 같으면 그쪽을 믿는다
        irv = ir.get(p, {}).get(s, {}).get(75)
        if irv and abs(irv - calc) <= max(abs(calc), 1) * TOL:
            print(f'   → Col75 기말 CSM: {v[75]:,.0f} → {irv:,.0f} (검산·IR 팩트시트 일치)')
            if not a.dry:
                ws.cell(r, c2e[75]).value = irv
                ws.cell(r, c2e[75]).fill = fill
            fixed += 1
            continue
        print('   → 자동 판정 불가. 수동 확인 필요')

    if not a.dry:
        wb.save(a.xlsx)
    print(f'\n바로잡은 칸: {fixed}')


if __name__ == '__main__':
    main()
