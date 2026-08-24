# -*- coding: utf-8 -*-
"""작업 파일이 비어 있는데 정답 시트에는 값이 있는 칸(=진짜 MISS)을 센다."""
import sys, argparse, collections
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from ir_identity import load_rows, ANSWER
from ir_fill import resolve_short


def work_rows(path, sheet='DATA'):
    # recalc.py 로 LibreOffice 계산을 거친 사본을 넣어야 수식 칸까지 제대로 센다.
    ws = openpyxl.load_workbook(path, data_only=True)[sheet]
    c2e = {}
    for ec in range(1, ws.max_column + 1):
        try:
            c2e[int(ws.cell(1, ec).value)] = ec
        except (TypeError, ValueError):
            pass
    out = {}
    for r in range(6, ws.max_row + 1):
        p = ws.cell(r, 2).value
        s = resolve_short(ws, r)
        if not p or not s:
            continue
        out[(str(p), s)] = {c: ws.cell(r, e).value for c, e in c2e.items()}
    return out, c2e


def score(path, answer=ANSWER, periods=None):
    arows, names = load_rows(answer)
    ans = {k: d for k, d in arows}
    work, _ = work_rows(path)
    miss = collections.Counter()
    wrong = []
    have = collections.Counter()
    for key, d in work.items():
        if periods and key[0] not in periods:
            continue
        a = ans.get(key)
        if not a:
            continue
        for c, av in a.items():
            v = d.get(c)
            if v in (None, ''):
                miss[key[0]] += 1
                miss['col%d' % c] += 1
            elif isinstance(v, (int, float)):
                have[key[0]] += 1
                if abs(v - av) > max(abs(av), 1e-9) * 0.002:
                    wrong.append((key[0], key[1], c, v, av))
    return miss, have, wrong, names


if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('path', nargs='?', default='DATA_작업_빈칸.xlsx')
    ap.add_argument('--periods', default='')
    ap.add_argument('--detail', action='store_true')
    a = ap.parse_args()
    per = a.periods.split(',') if a.periods else None
    miss, have, wrong, names = score(a.path, periods=per)
    tot = 0
    for p in sorted(k for k in miss if not str(k).startswith('col')):
        print(f'{p}  MISS {miss[p]:>5}   채워짐 {have[p]:>5}')
        tot += miss[p]
    print(f'합계 MISS {tot},  틀린 값 {len(wrong)}')
    if a.detail:
        cols = [(int(k[3:]), v) for k, v in miss.items() if str(k).startswith('col')]
        for c, n in sorted(cols, key=lambda x: -x[1])[:30]:
            print(f'  Col{c:<4} {n:>4}  {names.get(c,"")[:60]}')
        for w in wrong[:20]:
            print('  틀림', w)
