# -*- coding: utf-8 -*-
"""정답 시트를 열쇠 삼아 IR 팩트시트의 매핑을 자동으로 찾아낸다.

원리: 어떤 기간의 정답값을 이미 알고 있으므로, 팩트시트의 모든 숫자를 훑어
그 값과 일치하는 칸을 찾으면 "이 라벨이 그 컬럼"이라는 매핑이 역으로 드러난다.
단위 배수(1 / 100 / 1000)와 부호까지 같이 추정한다.

사용:
  python ir_discover.py 삼성생명 "<팩트시트.xlsx>" 2512
  python ir_discover.py 삼성생명 "<팩트시트.xlsx>" 2512 --only 21,22,23,95,96,97
"""
import re, sys, argparse
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from ir_parse import load, period_columns, period_to_yq

SHORT = {'미래에셋생명': '미래', '삼성생명': '삼성', '한화생명': '한화',
         '동양생명': '동양', '삼성화재': '삼성화재', '현대해상': '현대해상',
         '메리츠화재': '메리츠화재', 'KB손해보험': 'KB손보', 'KB생명': 'KB라이프',
         '교보생명': '교보', '신한라이프생명': '신한라이프', 'DB손해보험': 'DB손보'}
ANSWER_FILE = '202512_동업사 공시비교_DATA_작업_260320.xlsx'
UNITS = [(1000, '십억원'), (100, '억원'), (1, '백만원'), (0.001, '원')]
TOL = 0.001        # 정밀 일치
LOOSE = 0.01       # 근사 일치 (반올림 자릿수가 다른 경우)


def answer_row(period, short, path=ANSWER_FILE):
    ws = openpyxl.load_workbook(path, data_only=True)['DATA']
    c2e = {}
    for ec in range(1, ws.max_column + 1):
        try:
            c2e[int(ws.cell(1, ec).value)] = ec
        except (TypeError, ValueError):
            pass
    names = {}
    for col, ec in c2e.items():
        parts = []
        for r in (3, 4, 5):
            v = None
            for c2 in range(ec, 0, -1):
                v = ws.cell(r, c2).value
                if v is not None:
                    break
            if v is not None:
                parts.append(str(v).replace('\n', ' ').strip())
        names[col] = ' > '.join(dict.fromkeys(parts))
    for r in range(6, ws.max_row + 1):
        if str(ws.cell(r, 2).value) == period and str(ws.cell(r, 3).value).strip() == short:
            return ({c: ws.cell(r, e).value for c, e in c2e.items()
                     if isinstance(ws.cell(r, e).value, (int, float))}, names)
    return {}, names


def left_label(ws, row, col, back=8):
    """숫자 칸 왼쪽에서 가장 가까운 텍스트 = 그 행의 이름"""
    for c in range(col - 1, max(col - back, 0), -1):
        v = ws.cell(row, c).value
        if isinstance(v, str) and v.strip() and not v.strip().replace('.', '').isdigit():
            return v.strip()[:40]
    return ''


def discover_pairs(path, period, targets, window=12):
    """단일 셀로 안 맞는 컬럼을, **같은 열 안 두 행의 합**으로 다시 찾는다.

    팩트시트가 한 항목을 여러 줄로 쪼개 놓는 경우가 있다
    (미래에셋 FVOCI = 평가손익 + 손실충당금).
    행 간격이 멀면 우연의 일치가 늘어나므로 위아래 window 행 안에서만 본다.
    """
    wb = load(path)
    year, q = period_to_yq(period)
    hits = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.max_row > 1000:
            continue
        pcols = period_columns(ws, default_year=year)
        want = {c for (yy, pp), c in pcols.items() if yy == year and pp in (q, 'FY')}
        if not want:
            continue
        for col_x in want:
            vals = []
            for r in range(1, min(ws.max_row, 400) + 1):
                v = ws.cell(r, col_x).value
                if isinstance(v, (int, float)) and v:
                    vals.append((r, v))
            for i, (r1, v1) in enumerate(vals):
                for r2, v2 in vals[i + 1:i + 1 + window]:
                    tot = v1 + v2
                    for col, exp in targets.items():
                        if not exp:
                            continue
                        for mult, uname in UNITS:
                            if abs(tot * mult - exp) <= abs(exp) * TOL:
                                hits.setdefault(col, []).append(
                                    (sn, r1, r2,
                                     left_label(ws, r1, col_x),
                                     left_label(ws, r2, col_x), uname))
    return hits


def discover(path, period, targets, tol=TOL):
    wb = load(path)
    year, q = period_to_yq(period)
    hits = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.max_row > 3000:
            continue
        pcols = period_columns(ws, default_year=year)
        # 이 기간에 해당하는 열만 본다 (없으면 시트 전체)
        want = {c for (yy, pp), c in pcols.items()
                if yy == year and pp in (q, 'FY')}
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 400)):
            for cell in row:
                v = cell.value
                if not isinstance(v, (int, float)) or v == 0:
                    continue
                if want and cell.column not in want:
                    continue
                for col, exp in targets.items():
                    if exp is None or exp == 0:
                        continue
                    for mult, uname in UNITS:
                        for sign in (1, -1):
                            got = v * mult * sign
                            if abs(got - exp) <= abs(exp) * tol:
                                hits.setdefault(col, []).append(
                                    (sn, cell.row, cell.column,
                                     left_label(ws, cell.row, cell.column),
                                     uname, sign))
    return hits


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('company'); ap.add_argument('factsheet'); ap.add_argument('period')
    ap.add_argument('--only', default='', help='쉼표로 구분한 컬럼 번호')
    a = ap.parse_args()
    ans, names = answer_row(a.period, SHORT[a.company])
    if a.only:
        keep = {int(x) for x in a.only.split(',')}
        ans = {c: v for c, v in ans.items() if c in keep}
    hits = discover(a.factsheet, a.period, ans)
    print(f'=== {a.company} {a.period}: 정답 {len(ans)}개 컬럼 중 {len(hits)}개 매칭')
    for col in sorted(hits):
        print(f'\nCol{col}  {names.get(col, "")[:58]}   정답={ans[col]:,.3f}')
        seen = set()
        for sn, r, c, lab, uname, sign in hits[col][:4]:
            k = (sn, lab)
            if k in seen:
                continue
            seen.add(k)
            sg = '' if sign == 1 else ' (부호반전)'
            print(f'    {sn:<26} r{r:<4} c{c:<3} "{lab}"  단위={uname}{sg}')
    misses = [c for c in ans if c not in hits]
    if misses:
        print('\n못 찾은 컬럼:', ', '.join(f'Col{c}' for c in sorted(misses)))


if __name__ == '__main__':
    main()


# ---------------------------------------------------------------- PDF 탐색
def _line_matches(line, colname):
    """그 줄의 글자가 컬럼 이름과 실제로 겹치는지.

    PDF 는 페이지당 숫자가 수백 개라 값만 맞춰서는 우연의 일치가 압도적이다
    (교보 경영공시 638쪽에서 89개 컬럼 중 88개가 '일치'했다).
    줄에 붙은 항목명이 컬럼 이름과 겹칠 때만 후보로 본다.
    """
    stop = set('기준 등 및 계 합계 전기 당기 b-a 2~5년 22년말 ifrs4 수정함 '
               '순부채기준 24.12월부터 최종 재보제외 잠정치 회사명'.split())

    def toks(t):
        t = re.sub(r'\(.*?\)', '', str(t))
        return {w for w in re.split(r'[^0-9A-Za-z가-힣%\-]+', t)
                if len(w) > 1 and w not in stop}

    return bool(toks(line) & toks(colname))


def discover_pdf(path, targets, tol=TOL, names=None):
    """PDF 본문의 숫자를 훑어 정답값과 맞는 곳을 찾는다.

    PDF 는 셀 구조가 없으므로 '몇 페이지 몇 번째 줄'과 그 줄 전체를 돌려준다.
    표를 눈으로 확인해 매핑을 정할 때 쓰는 실마리다.
    """
    import pdfplumber
    from ir_pdf import _nums

    hits = {}
    with pdfplumber.open(path) as pdf:
        for pno, page in enumerate(pdf.pages, 1):
            text = page.extract_text() or ''
            for ln, line in enumerate(text.split('\n'), 1):
                nums = _nums(line)
                if not nums:
                    continue
                for col, exp in targets.items():
                    if not exp:
                        continue
                    for v in nums:
                        if not v:
                            continue
                        for mult, uname in UNITS:
                            for sign in (1, -1):
                                if abs(v * mult * sign - exp) <= abs(exp) * tol:
                                    if names and not _line_matches(line, names.get(col, '')):
                                        continue
                                    hits.setdefault(col, []).append(
                                        (pno, ln, line.strip()[:70], uname, sign))
                                    break
    return hits
