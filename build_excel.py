"""
엑셀 복사본을 만들어 XBRL에서 뽑을 수 있는 값으로 채우는 스크립트
단위: 억원 (원 -> /1e8)
"""
import os, csv, sys
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

EXCEL_SRC = "(보고) 202512_동업사 공시비교_20260401.xlsx"
EXCEL_DST = "data/동업사_공시비교_2026Q1.xlsx"
FINANCIALS_CSV = "data/financials_2026Q1.csv"

# 엑셀 헤더 회사 순서 (엑셀 기준)
COMPANIES_EXCEL = ["미래에셋생명", "삼성생명", "교보생명", "한화생명", "신한라이프", "동양생명", "KB라이프", "삼성화재", "현대해상", "메리츠화재", "KB손보", "DB손보"]

# financials CSV의 회사명 -> 엑셀 회사명 매핑
CORP_MAP = {
    "미래에셋생명": "미래에셋생명",
    "삼성생명": "삼성생명",
    "교보생명": "교보생명",
    "한화생명": "한화생명",
    "신한라이프생명": "신한라이프",
    "동양생명": "동양생명",
    "KB생명": "KB라이프",
    "삼성화재": "삼성화재",
    "현대해상": "현대해상",
    "메리츠화재": "메리츠화재",
    "KB손해보험": "KB손보",
    "DB손해보험": "DB손보",
}

UNIT = 1e8  # 원 -> 억원


def load_financials():
    """financials_2026Q1.csv 로드: {account_id: {col: value}}"""
    data = {}
    with open(FINANCIALS_CSV, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            aid = row["account_id"]
            data[aid] = {k: v for k, v in row.items() if k not in ("account_id", "label_kor")}
    return data


def get_val(data, account_id, corp_csv, col_suffix):
    """corp_csv 회사의 특정 account_id 값 반환 (억원). col_suffix 예: '연결_당기말'"""
    row = data.get(account_id, {})
    col = f"{corp_csv}_{col_suffix}"
    v = row.get(col, "")
    if v:
        try:
            return round(float(v) / UNIT, 2)
        except ValueError:
            return ""
    return ""


def get_val_sum(data, account_id, corp_csv, col_suffixes):
    """여러 컬럼 합산 (없으면 첫 번째 있는 것 반환)"""
    for suffix in col_suffixes:
        v = get_val(data, account_id, corp_csv, suffix)
        if v != "":
            return v
    return ""


def build_sheet(ws_src, ws_dst, data):
    """시트를 복사하되 숫자 셀은 XBRL 값으로 교체"""

    # 회사명 헤더 행 찾기 (col index)
    header_row = None
    corp_col_map = {}  # 엑셀명 -> col index
    for row in ws_src.iter_rows():
        for cell in row:
            if cell.value == "미래에셋생명":
                header_row = cell.row
                # 같은 행에서 모든 회사 위치 파악
                for c in ws_src[header_row]:
                    if c.value in COMPANIES_EXCEL:
                        corp_col_map[c.value] = c.column
                break
        if header_row:
            break

    print(f"  헤더 행: {header_row}, 회사 컬럼: {corp_col_map}")

    # 역매핑: 엑셀회사명 -> csv회사명
    excel_to_csv = {v: k for k, v in CORP_MAP.items()}

    # XBRL 값으로 채울 항목 정의
    # (엑셀 행 레이블 키워드, account_id, col_suffix 우선순위 리스트)
    FILL_MAP = [
        # 재무상태표
        ("자산",     "ifrs-full_Assets",      ["연결_당기말", "별도_당기말"]),
        ("부채",     "ifrs-full_Liabilities",  ["연결_당기말", "별도_당기말"]),
        ("CSM",      "ifrs-full_ContractualServiceMargin", ["연결_당기말", "별도_당기말"]),
        ("이익잉여금", "ifrs-full_RetainedEarnings", ["연결_당기말", "별도_당기말"]),
        ("기타포괄손익", "ifrs-full_OtherComprehensiveIncome", ["연결_당기_누적", "별도_당기_누적"]),
        # 자기자본
        ("자기자본(신종자본증권제외)", "ifrs-full_Equity", ["연결_당기말", "별도_당기말"]),
        # 손익
        ("보험손익",   "ifrs-full_InsuranceServiceResult",  ["연결_당기_누적", "별도_당기_누적"]),
        ("투자손익",   "dart_InsuranceFinanceIncomeFromInsuranceContractsIssuedRecognisedInProfitOrLoss", ["연결_당기_누적", "별도_당기_누적"]),
        ("세전이익",   "ifrs-full_ProfitLossBeforeTax",     ["연결_당기_누적", "별도_당기_누적"]),
        ("당기순이익", "ifrs-full_ProfitLoss",              ["연결_당기_누적", "별도_당기_누적"]),
        ("영업이익",   "ifrs-full_ProfitLossFromOperatingActivities", ["연결_당기_누적", "별도_당기_누적"]),
    ]

    # 레이블 -> 행 번호 매핑
    label_rows = {}
    for row in ws_src.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                label_rows.setdefault(cell.value.strip(), []).append(cell.row)

    filled = 0
    for label_kw, account_id, suffixes in FILL_MAP:
        target_rows = []
        for lbl, rows in label_rows.items():
            if lbl == label_kw or (label_kw in lbl and len(lbl) < len(label_kw) + 10):
                target_rows.extend(rows)

        if not target_rows:
            print(f"  ⚠️  '{label_kw}' 행 못 찾음")
            continue

        for r in set(target_rows):
            for excel_corp, col_idx in corp_col_map.items():
                csv_corp = excel_to_csv.get(excel_corp)
                if not csv_corp:
                    continue
                val = get_val_sum(data, account_id, csv_corp, suffixes)
                if val != "":
                    ws_dst.cell(row=r, column=col_idx, value=val)
                    filled += 1

    print(f"  채운 셀: {filled}개")


def main():
    print("엑셀 로드 중...")
    wb_src = load_workbook(EXCEL_SRC, data_only=True)
    wb_dst = load_workbook(EXCEL_SRC, data_only=True)

    data = load_financials()
    print(f"XBRL 데이터 로드: {len(data)}개 account_id")

    # 숫자 셀 초기화 (2512 시트)
    ws_src = wb_src["2512"]
    ws_dst = wb_dst["2512"]

    # 숫자 값 모두 지우기 (회사 데이터 영역만)
    print("숫자 초기화 중...")
    header_row = None
    corp_cols = set()
    for row in ws_src.iter_rows():
        for cell in row:
            if cell.value == "미래에셋생명":
                header_row = cell.row
                for c in ws_src[header_row]:
                    if c.value in COMPANIES_EXCEL:
                        corp_cols.add(c.column)
                break
        if header_row:
            break

    for row in ws_dst.iter_rows(min_row=header_row + 1):
        for cell in row:
            if cell.column in corp_cols and isinstance(cell.value, (int, float)):
                cell.value = None

    print("XBRL 값 채우기 중...")
    build_sheet(ws_src, ws_dst, data)

    os.makedirs("data", exist_ok=True)
    wb_dst.save(EXCEL_DST)
    print(f"\n저장 완료: {EXCEL_DST}")


if __name__ == "__main__":
    main()
