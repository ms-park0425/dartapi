# -*- coding: utf-8 -*-
"""아직 못 채운 셀에 대해 팩트시트를 넓게 훑는다.

1) 단일 셀 정밀(TOL) — 이미 ir_discover.discover 가 하는 것
2) 두 행 합(discover_pairs)
3) 느슨한 허용오차(LOOSE) — 반올림 자릿수가 다른 표

의미 없는 우연의 일치를 걸러내려고 _line_matches 로 라벨/컬럼명 토큰이
겹치는 후보만 남긴다.
"""
import sys, argparse, glob, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from ir_discover import (answer_row, discover, discover_pairs, SHORT,
                         _line_matches, LOOSE, TOL)
from ir_parse import SPECS

WORK = 'DATA_작업_빈칸.xlsx'


def miss_cols(period, short, path=WORK):
    """작업 파일에서 아직 비어 있는 컬럼 번호"""
    ws = openpyxl.load_workbook(path, data_only=False)['DATA']
    c2e = {}
    for ec in range(1, ws.max_column + 1):
        try:
            c2e[int(ws.cell(1, ec).value)] = ec
        except (TypeError, ValueError):
            pass
    for r in range(6, ws.max_row + 1):
        if str(ws.cell(r, 2).value) == period and str(ws.cell(r, 3).value).strip() == short:
            return {c for c, e in c2e.items() if ws.cell(r, e).value in (None, '')}
    return set()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--companies', default='')
    ap.add_argument('--periods', default='2503,2506,2509,2512')
    a = ap.parse_args()
    comps = a.companies.split(',') if a.companies else sorted(SPECS)

    for comp in comps:
        mapped = {it.col for it in SPECS[comp]['items']}
        mapped |= set((SPECS[comp].get('derived') or {}))
        for period in a.periods.split(','):
            fs = sorted(glob.glob(f'data/{comp}/IR/*/*.xlsx'))
            if not fs:
                continue
            ans, names = answer_row(period, SHORT[comp])
            miss = miss_cols(period, SHORT[comp])
            todo = {c: v for c, v in ans.items()
                    if c in miss and c not in mapped and v}
            if not todo:
                continue
            for f in fs:
                try:
                    h1 = discover(f, period, todo, tol=LOOSE)
                except Exception as e:
                    print(f'  !! {os.path.basename(f)} {e}'); continue
                try:
                    h2 = discover_pairs(f, period, todo)
                except Exception:
                    h2 = {}
                for col in sorted(set(h1) | set(h2)):
                    lines = []
                    for sn, r, c, lab, un, sg in h1.get(col, [])[:6]:
                        if _line_matches(f'{sn} {lab}', names.get(col, '')):
                            lines.append(f'      단일 {sn} r{r} "{lab}" {un}{"" if sg==1 else " 부호반전"}')
                    for sn, r1, r2, l1, l2, un in h2.get(col, [])[:6]:
                        if _line_matches(f'{sn} {l1} {l2}', names.get(col, '')):
                            lines.append(f'      합   {sn} r{r1}+r{r2} "{l1}"+"{l2}" {un}')
                    if lines:
                        print(f'\n[{comp} {period}] Col{col} {names.get(col,"")[:55]} = {ans[col]:,.3f}')
                        print(f'    ({os.path.basename(f)})')
                        for l in dict.fromkeys(lines):
                            print(l)


if __name__ == '__main__':
    main()
