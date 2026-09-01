"""
data_sheet_{period}.csv vs DATA 시트 정답 비교 검증 스크립트
"""
import csv
import os
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import sys
import argparse

sys.stdout.reconfigure(encoding="utf-8")

parser = argparse.ArgumentParser()
parser.add_argument("--period", default="2512")
args, _ = parser.parse_known_args()

PERIOD = args.period
ANSWER_FILE = "202512_동업사 공시비교_DATA_작업_260320.xlsx"
OUTPUT_CSV = f"data/data_sheet_{PERIOD}_ir.csv" if os.path.exists(f"data/data_sheet_{PERIOD}_ir.csv") else f"data/data_sheet_{PERIOD}.csv"


def load_answer():
    """Excel DATA 시트에서 2512 정답 로드"""
    wb = openpyxl.load_workbook(ANSWER_FILE, data_only=True)
    ws = wb["DATA"]

    # Row 1 has column numbers
    col_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None:
            col_map[c] = int(v)

    answer = {}
    for r in range(5, ws.max_row + 1):
        key = ws.cell(row=r, column=1).value
        if not key or str(PERIOD) not in str(key):
            continue
        key = str(key).strip()
        row_data = {}
        for c, col_num in col_map.items():
            if col_num == 1:
                continue
            v = ws.cell(row=r, column=c).value
            if v is not None and isinstance(v, (int, float)):
                row_data[col_num] = v
        answer[key] = row_data

    return answer


def load_output():
    """CSV 출력 파일 로드"""
    output = {}
    with open(OUTPUT_CSV, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            key = row["key"]
            data = {}
            for k, v in row.items():
                if k.startswith("Col") and v:
                    try:
                        col_num = int(k.replace("Col", ""))
                        data[col_num] = float(v)
                    except ValueError:
                        pass
            output[key] = data
    return output


def compare(answer, output, tol=0.01):
    """비교: OK/FAIL/MISS 집계"""
    ok, fail, miss = 0, 0, 0
    fail_details = []
    miss_details = {}
    miss_cells = []  # (short, col_num) for painting

    for csv_key, out in output.items():
        if csv_key not in answer:
            print(f"  WARNING: no answer for {csv_key}")
            continue
        ans = answer[csv_key]

        CHECK_COLS = {31, 45, 46, 47, 54, 64, 72, 82, 89, 90, 92, 93, 94, 159}
        for col_num, expected in ans.items():
            if col_num in (1, 2, 3):
                continue
            if col_num in CHECK_COLS:
                continue

            actual = out.get(col_num)

            if actual is None:
                miss += 1
                miss_details.setdefault(col_num, []).append(csv_key[4:])
                miss_cells.append((csv_key[4:], col_num))
                continue

            # Compare
            abs_diff = abs(actual - expected)
            if abs_diff <= 1:  # absolute tolerance: diff <= 1 is OK
                ok += 1
            elif abs(expected) < tol:  # effectively zero
                if abs(actual) < tol:
                    ok += 1
                else:
                    fail += 1
                    fail_details.append((csv_key[4:], col_num, expected, actual))
            else:
                rel_err = abs_diff / max(abs(expected), 1e-10)
                if rel_err < 0.001:  # 0.1% tolerance
                    ok += 1
                else:
                    fail += 1
                    fail_details.append((csv_key[4:], col_num, expected, actual))

    return ok, fail, miss, fail_details, miss_details, miss_cells


ORANGE_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
EXCEL_OUT = "DATA_작업_빈칸.xlsx"


def paint_fails(fail_details, miss_cells=None):
    """FAIL/MISS 셀을 주황색으로 칠하기"""
    import os
    if miss_cells is None:
        miss_cells = []
    if not os.path.exists(EXCEL_OUT):
        print(f"\n[SKIP] {EXCEL_OUT} 없음 — 색칠 생략")
        return
    if not fail_details and not miss_cells:
        print("\n[색칠] FAIL/MISS 없음")
        return

    wb = openpyxl.load_workbook(EXCEL_OUT)
    ws = wb["DATA"]

    # row1: col_num → excel열 역매핑
    col_to_excel = {}
    for ec in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=ec).value
        if v is not None:
            try:
                col_to_excel[int(v)] = ec
            except (ValueError, TypeError):
                pass

    # col2=기간코드, col3=회사명 → 행 매핑
    period_col = col_to_excel.get(2, 2)
    name_col = col_to_excel.get(3, 3)
    short_to_row = {}
    for r in range(6, ws.max_row + 1):
        v2 = ws.cell(row=r, column=period_col).value
        v3 = ws.cell(row=r, column=name_col).value
        if str(v2) == str(PERIOD) and v3:
            short_to_row[str(v3).strip()] = r

    painted = 0
    for short, col_num, _exp, _act in fail_details:
        row_idx = short_to_row.get(short)
        ec = col_to_excel.get(col_num)
        if row_idx and ec:
            ws.cell(row=row_idx, column=ec).fill = ORANGE_FILL
            painted += 1
    for short, col_num in miss_cells:
        row_idx = short_to_row.get(short)
        ec = col_to_excel.get(col_num)
        if row_idx and ec:
            ws.cell(row=row_idx, column=ec).fill = ORANGE_FILL
            painted += 1

    wb.save(EXCEL_OUT)
    print(f"\n[색칠] {painted}셀 주황색 표시 완료 → {EXCEL_OUT}")


def main():
    print("Loading answer from Excel...")
    answer = load_answer()
    print(f"  {len(answer)} companies loaded")

    print("Loading output CSV...")
    output = load_output()
    print(f"  {len(output)} companies loaded")

    print("\nComparing...")
    ok, fail, miss, fail_details, miss_details, miss_cells = compare(answer, output)

    total = ok + fail + miss
    print(f"\n{'='*60}")
    print(f"TOTAL: OK={ok}, FAIL={fail}, MISS={miss} (total={total})")
    print(f"{'='*60}")

    if fail_details:
        print(f"\n--- FAIL details (top 30) ---")
        for key, col, exp, act in fail_details[:30]:
            pct = abs(act - exp) / max(abs(exp), 1e-10) * 100 if exp != 0 else abs(act)
            print(f"  {key} Col{col}: expected={exp:.4f}, got={act:.4f} (err={pct:.2f}%)")

    if miss_details:
        # Summarize by column
        print(f"\n--- MISS summary by column (top 20) ---")
        sorted_miss = sorted(miss_details.items(), key=lambda x: -len(x[1]))
        for col, keys in sorted_miss[:20]:
            print(f"  Col{col} ({len(keys)} miss): {', '.join(keys[:4])}{'...' if len(keys)>4 else ''}")

    paint_fails(fail_details, miss_cells)


if __name__ == "__main__":
    main()
