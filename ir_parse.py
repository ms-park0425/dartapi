# -*- coding: utf-8 -*-
"""
IR 팩트시트 → DATA 시트 컬럼 추출 모듈

XBRL / 원문XML / FISIS 로 채우지 못하는 항목(예실차·CSM변동 등)을
각 보험사 IR 팩트시트(xlsx)에서 추출한다.

핵심 규칙 (미래에셋생명 2503~2512 정답 대조로 검증):
  * 단위: 팩트시트 십억원(BN KRW) → DATA 시트 백만원  = ×1000
  * flow 항목(예실차·신계약CSM·상각·이자·가정변경)은 **연간 누계** = Q1..Qn 합
    - 2512(FY)는 FY 컬럼 그대로
  * stock 항목(기말CSM)은 해당 분기 **시점값**
  * Col80(전기CSM)은 **직전 연도 FY 기말CSM**
  * CSM상각(Col77)은 팩트시트가 음수, DATA는 양수 → 부호 반전
  * 팩트시트는 후속 분기에 과거 수치를 재작성(restate)한다.
    → 각 기간은 **그 기간에 공시된 팩트시트**를 쓰는 것이 원칙.
"""
import io, os, re, glob, shutil, zipfile
import openpyxl


def load(path):
    """openpyxl 로 열되, 손상된 docProps/custom.xml 때문에 실패하면 그 부분만 빼고 다시 연다.

    (삼성화재 팩트시트가 name 이 비어 있는 custom property 를 갖고 있어 그대로는 열리지 않음)
    """
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except TypeError:
        buf = io.BytesIO()
        with zipfile.ZipFile(path) as zin, zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
            for it in zin.infolist():
                if it.filename == 'docProps/custom.xml':
                    continue
                zout.writestr(it, zin.read(it.filename))
        buf.seek(0)
        return openpyxl.load_workbook(buf, data_only=True)

MILLION_PER_BILLION = 1000.0   # 십억원 → 백만원

# ---------------------------------------------------------------- 헤더 인식
def _norm_period(raw):
    if isinstance(raw, str):
        t = raw.strip().upper()
        if t == 'FY': return 'FY'
        if t in ('Q1', 'Q2', 'Q3', 'Q4'): return t
        if t in ('1Q', '2Q', '3Q', '4Q'): return 'Q' + t[0]
        if t == 'YTD': return 'FY'   # 연간누계 열
        if t in ('1', '2', '3', '4'): return 'Q' + t
    elif isinstance(raw, int) and 1 <= raw <= 4:
        return 'Q%d' % raw
    return None


def _combined(v):
    """한 칸이 연도와 기간을 함께 담은 경우 → (연도, 기간).

    '25.4Q' / 'FY25.2Q' / '25.4Q\n누계'(=연간누계) / 날짜셀(2025-06-30) 을 인식한다.
    """
    import datetime as _dt
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.year, 'Q%d' % ((v.month - 1) // 3 + 1)
    if isinstance(v, (int, float)):
        # 25.12 / 24.09 처럼 'YY.MM' 을 숫자로 적어둔 헤더
        yy = int(v)
        if 20 <= yy <= 99:
            mm = int(round((v - yy) * 100))
            if mm in (3, 6, 9, 12):
                return 2000 + yy, 'Q%d' % (mm // 3)
        return None
    if not isinstance(v, str):
        return None
    t = v.strip().lstrip("'\u2019")
    m = re.match(r'^(?:FY)?\s*(\d{2,4})\s*[.\-/]\s*([1-4])\s*Q', t, re.I)
    if not m:
        # '2025 Q4' / '2025Q4' 형태 (분기 표기가 뒤에 오는 서식)
        m2 = re.match(r'^(?:FY)?\s*(\d{2,4})\s*Q\s*([1-4])\s*$', t, re.I)
        if not m2:
            m3 = re.match(r'^(\d{2})\.(03|06|09|12)$', t)
            if m3:
                return 2000 + int(m3.group(1)), 'Q%d' % (int(m3.group(2)) // 3)
            # '2025. 12월 말' / '2025.12월' 형태
            m4 = re.match(r'^(20\d\d)\s*[.\-]\s*(\d{1,2})\s*월', t)
            if m4:
                mm = int(m4.group(2))
                if mm in (3, 6, 9, 12):
                    return int(m4.group(1)), 'Q%d' % (mm // 3)
            # 'Mar. 25' / 'Dec.25' 형태 (KB금융 팩트북)
            m5 = re.match(r'^(Mar|Jun|Sep|Dec)\w*\.?\s*(\d{2,4})$', t, re.I)
            if m5:
                mm = {'mar': 3, 'jun': 6, 'sep': 9, 'dec': 12}[m5.group(1).lower()]
                yy = int(m5.group(2))
                return (2000 + yy if yy < 100 else yy), 'Q%d' % (mm // 3)
            return None
        n2 = int(m2.group(1))
        return (2000 + n2 if n2 < 100 else n2), 'Q' + m2.group(2)
    n = int(m.group(1))
    year = 2000 + n if n < 100 else n
    # '누계' 표기가 붙으면 연간누계 열
    return year, ('FY' if '누계' in t or 'YTD' in t.upper() else 'Q' + m.group(2))


def _norm_year(v):
    """2025 / 'FY25' / 'FY2025' / 'FY25.4Q' → 2025"""
    if isinstance(v, int) and 2000 < v < 2100:
        return v
    if isinstance(v, str):
        m = re.match(r'^\s*FY\s*(\d{2,4})', v.strip(), re.I)
        if m:
            n = int(m.group(1))
            return 2000 + n if n < 100 else n
    return None


def period_columns(ws, maxscan=12, default_year=None):
    """{(연도, 'FY'|'Q1'..'Q4'): 엑셀 열번호}

    기간 라벨이 3개 이상인 행을 기간행으로 보고, 그 바로 위 행을 연도행으로 읽는다.
    연도 셀이 병합되어 있으면 병합 범위 전체에 그 연도를 전파한다.
    """
    # ① 연도+기간을 한 칸에 담은 헤더(예: '25.4Q', 날짜셀)를 먼저 훑는다.
    combined = {}
    for r in range(1, maxscan + 1):
        text_hits, num_hits = {}, {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            yp = _combined(v)
            if not yp:
                continue
            (num_hits if isinstance(v, (int, float)) else text_hits)[c] = yp
        # 문자열 헤더('25.4Q', '2025 Q4', 날짜셀)는 한 개만 있어도 믿는다.
        # 숫자형(25.12)은 본문 숫자와 구분이 안 되므로 한 행에 2개 이상일 때만 인정한다.
        for c, yp in text_hits.items():
            combined.setdefault(c, yp)
        if len(num_hits) >= 2:
            for c, yp in num_hits.items():
                combined.setdefault(c, yp)

    # ② 연도행 + 기간행 방식
    prow = None
    for r in range(1, maxscan + 1):
        n = sum(1 for c in range(1, ws.max_column + 1)
                if _norm_period(ws.cell(r, c).value))
        if n >= 3:
            prow = r
            break
    if not prow:
        out = {}
        for c, (y, p) in sorted(combined.items()):
            out.setdefault((y, p), c)
        if not out:
            # 분기 없이 연도만 적힌 표(삼성생명 Ⅰ-1 의 'FY24 | FY25')는 FY 열로 본다.
            for r in range(1, maxscan + 1):
                ys = {c: _norm_year(ws.cell(r, c).value)
                      for c in range(1, ws.max_column + 1)}
                ys = {c: y for c, y in ys.items() if y}
                if len(ys) >= 2:
                    for c, y in sorted(ys.items()):
                        out.setdefault((y, 'FY'), c)
                    break
        return out
    yrow = prow - 1
    ymap = {}
    for c in range(1, ws.max_column + 1):
        y = _norm_year(ws.cell(yrow, c).value)
        if y:
            ymap[c] = y
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= yrow <= mr.max_row:
            y = _norm_year(ws.cell(mr.min_row, mr.min_col).value)
            if y:
                for c in range(mr.min_col, mr.max_col + 1):
                    ymap.setdefault(c, y)
    # 연도 행이 아예 없는 시트(삼성생명 Ⅰ-3 처럼 분기만 적힌 경우)는
    # 그 팩트시트가 다루는 연도로 간주한다.
    if not ymap and default_year:
        for c in range(1, ws.max_column + 1):
            ymap[c] = default_year

    out, year = {}, None
    for c in range(1, ws.max_column + 1):
        if c in combined:
            out.setdefault(combined[c], c)
            year = combined[c][0]      # 이후 '2Q','3Q' 처럼 연도 없는 칸에 물려준다
            continue
        if c in ymap:
            year = ymap[c]
        p = _norm_period(ws.cell(prow, c).value)
        if year and p:
            out.setdefault((year, p), c)
    return out


def label_row(ws, labels, label_cols=(1, 2, 3, 4, 5, 6, 7, 8, 9), row_min=1, row_max=400):
    """라벨 문자열 → 행번호.

    라벨 열 중 어느 한 칸이 정확히 일치하거나, 앞 칸들을 이어붙인 문자열이
    일치하면 그 행으로 본다(영문 병기 열이 뒤에 붙는 서식 대응).
    """
    def norm(x):
        return re.sub(r'\s+', '', str(x)).lower()

    targets = {L: norm(L) for L in labels}
    found = {}
    for r in range(row_min, min(ws.max_row, row_max) + 1):
        cells = [ws.cell(r, c).value for c in label_cols]
        singles = [norm(v) for v in cells if v is not None and str(v).strip()]
        if not singles:
            continue
        joined = norm(' '.join(str(v) for v in cells[:2] if v is not None))
        for L, t in targets.items():
            if L in found:
                continue
            if t in singles or joined == t:
                found[L] = r
    return found


# ---------------------------------------------------------------- 추출 규칙
CUM, POINT, PRIOR_FY, OPEN = 'cum', 'point', 'prior_fy', 'open'

class Item:
    """팩트시트 한 항목 → DATA 컬럼 하나.

    section: 같은 라벨이 여러 블록에 나올 때(예: 한화 '이자부리'가 BEL·CSM 양쪽)
             이 앵커 라벨이 나온 행 **아래**에서만 label 을 찾는다.
    """
    __slots__ = ('sheet', 'label', 'col', 'mode', 'sign', 'section', 'unit')
    def __init__(self, sheet, label, col, mode=CUM, sign=1, section=None, unit=None):
        self.sheet, self.label, self.col = sheet, label, col
        self.mode, self.sign, self.section = mode, sign, section
        self.unit = unit   # 시트마다 단위가 다를 때만 지정 (예: SAP-BS 는 백만원)

# 회사별 매핑. unit = 팩트시트 단위 → 백만원 배수
#   십억원(BN KRW) = 1000, 억원 = 100
SPECS = {
    # SAP-BS 는 감독회계 재무상태표 시트. FISIS 에서 받던 항목들이 여기 그대로 있다.
    '미래에셋생명': {'unit': 1000,
                     'derived': {25: (250, 251, '+'), 107: (109, 108, '/')},
                     'items': [
        Item('SAP-BS', 'I. 운용자산',                      22, POINT, unit=1),
        Item('SAP-BS', '4. 보험계약자산(부채)순금융손익',   26, POINT, unit=1),
        Item('SAP-BS', '5. 재보험계약자산(부채)순금융손익', 27, POINT, unit=1),
        Item('SAP-BS', '8. 위험회피파생상품평가손익',       28, POINT, unit=1),
        Item('SAP-BS', '9. 재평가잉여금',                    29, POINT, unit=1),
        Item('SAP-BS', '10. 보험수리적손익',                 30, POINT, unit=1),
        Item('SAP-BS', '6. 해약환급금준비금',                99, POINT, unit=1),
        Item('SAP-BS', '1. 기타포괄손익-공정가치측정금융상품평가손익', 250, POINT, unit=1),
        Item('SAP-BS', '2. 기타포괄손익-공정가치측정채무상품손실충당금', 251, POINT, unit=1),
        Item('EFF',    '위험보험료',                       108),
        Item('EFF',    '보험금',                           109),
        Item('PROFIT', '예실차',     111),
        Item('PROFIT', '보험금차',   112),
        Item('PROFIT', '예상보험금', 113),
        Item('PROFIT', '실제보험금', 114),
        Item('PROFIT', '사업비차',   115),
        Item('PROFIT', '예상사업비', 116),
        Item('PROFIT', '실제사업비', 117),
        Item('CSM',    '기말CSM',     75, POINT),
        Item('CSM',    '신계약CSM',   76),
        Item('CSM',    'CSM상각',     77, CUM, -1),
        Item('CSM',    '가정변경 등', 78),
        Item('CSM',    '이자비용',    79),
        Item('CSM',    '기말CSM',     80, PRIOR_FY),
    ]},

    # 한화생명 팩트시트는 분기 파일마다 시트명이 '요약 손익계산서' ↔ '(별도) 요약
    # 손익계산서' 로 바뀐다 → 같은 컬럼을 두 이름으로 모두 걸어 둔다(하나만 맞으면 됨).
    '한화생명': {'unit': 1000, 'items': [
        Item('운용자산',        '자산계',           21, POINT),
        Item('요약 손익계산서', '예실차',          111),
        Item('요약 손익계산서', '보험금 예실차',   112),
        Item('요약 손익계산서', '사업비 예실차',   115),
        Item('(별도) 요약 손익계산서', '예실차',        111),
        Item('(별도) 요약 손익계산서', '보험금 예실차', 112),
        Item('(별도) 요약 손익계산서', '사업비 예실차', 115),
        Item('(별도) 요약 손익계산서', '보험손익',      48),
        Item('(별도) 요약 손익계산서', '당기순이익',    44),
        Item('요약 손익계산서', '보험손익',   48),
        Item('요약 손익계산서', '당기순이익', 44),
        Item('(별도) 요약 재무상태표', '자산계',           21, POINT),
        Item('(별도) 요약 재무상태표', '위험조정(RA)',     74, POINT),
        Item('요약 재무상태표', '위험조정(RA)',     74, POINT),
        # Col49·Col50(투자·영업외손익), Col73(BEL), Col109(사고보험금)은 팩트시트가
        # DATA 정의와 범위가 달라(별도 SAP vs 감독기준) 매핑에서 뺐다.
        Item('효율지표', '위험보험료', 108),
        Item('보험부채 Movement', '기말 보험계약마진', 75, POINT,  section='■ 보험계약마진(CSM)'),
        Item('보험부채 Movement', '신계약',           76, CUM,     section='■ 보험계약마진(CSM)'),
        Item('보험부채 Movement', '상각',             77, CUM, -1, section='■ 보험계약마진(CSM)'),
        Item('보험부채 Movement', '가정변경 등 조정', 78, CUM,     section='■ 보험계약마진(CSM)'),
        Item('보험부채 Movement', '이자부리',         79, CUM,     section='■ 보험계약마진(CSM)'),
        Item('보험부채 Movement', '기말 보험계약마진', 80, PRIOR_FY, section='■ 보험계약마진(CSM)'),
    ]},

    '메리츠화재': {'unit': 100, 'items': [
        Item('CSM', '기말 CSM',  75, POINT),
        Item('CSM', '신계약 CSM', 76),
        Item('CSM', 'CSM 상각',  77, CUM, -1),
        Item('CSM', '경험조정 등', 78),
        Item('CSM', '이자비용',   79),
        Item('CSM', '기말 CSM',  80, PRIOR_FY),
    ]},

    '삼성생명': {'unit': 1000,
                 'derived': {107: (109, 108, '/')},
                 'items': [
        Item('Ⅰ-6', '위험보험료',     108),
        Item('Ⅰ-6', '사고보험금1)',   109),
        Item('Ⅰ-5', '기말 CSM',        75, POINT),
        Item('Ⅰ-5', '신계약 CSM',      76),
        Item('Ⅰ-3', '상각',            77, CUM, -1, section='CSM'),
        Item('Ⅰ-3', '가정변경等 조정',  78, CUM,     section='CSM'),
        Item('Ⅰ-3', '이자부리',        79, CUM,     section='CSM'),
        Item('Ⅰ-5', '기말 CSM',        80, PRIOR_FY),
        Item('Ⅰ-2', '예실차',         111),
        Item('Ⅰ-3', '기말 BEL',       73, POINT),
        Item('Ⅰ-3', '기말 RA',        74, POINT),
    ]},

    '동양생명': {'unit': 1000,
                 'derived': {92: (93, 94, '/')},
                 'items': [
        Item('Stability', '지급여력금액 (Available Capital) (A)', 93, POINT),
        Item('Stability', '지급여력기준금액 (Required Capital) (B)', 94, POINT),
        Item('CSM', '기말 CSM (Ending CSM)',            75, POINT,    section='CSM'),
        Item('CSM', '신계약 (New business)',            76, CUM,      section='CSM'),
        Item('CSM', 'CSM 상각 (Amortization of CSM)',   77, CUM, -1,  section='CSM'),
        Item('CSM', '기타 (Others)',                    78, CUM,      section='CSM'),
        Item('CSM', '이자부리 (Unwinding)',             79, CUM,      section='CSM'),
        Item('CSM', '기말 CSM (Ending CSM)',            80, PRIOR_FY, section='CSM'),
        # 동양생명 FS 시트는 연결 기준이라 자산·자본·순이익은 DATA(별도)와 3% 안팎
        # 차이가 난다 → 별도/연결 차이가 없는 항목만 쓴다.
        Item('FS', 'ⅰ. 보험계약부채 (Insurance contracts issued that are liabilities)', 13, POINT),
        Item('FS', '보험손익 (Insurance service result)', 48),
        Item('FS', '예실차 (Difference between estimated and actual)', 111),
        Item('FS', '보험금예실차 (Claims experience variance)',        112),
        Item('FS', '사업비예실차 (Expense experience variance)',       115),
    ]},

    # 현대해상 팩트시트는 백만원 단위로 적혀 있다(unit=1).
    # Col77(CSM상각)은 팩트시트가 장기부문만 담고 있어 DATA 기준(전사)과 1.4% 차이가 난다
    # → IR 소스에서 제외하고 기존 XBRL 값을 쓴다.
    # Col114·Col117 은 팩트시트에 '발생손해액/실제사업비'로 실리는데 DATA 정의(예상−예실차)와
    # 범위가 달라, 정의대로 계산해서 채운다.
    '현대해상': {'unit': 1,
                 'derived': {114: (113, 112), 117: (116, 115), 23: (22, 21, '/')},
                 'items': [
        Item('(6) BS',         '총   자   산',        21, POINT),
        Item('(4) Investment', '운용자산 계',          22, POINT),
        Item('(6) BS',         '- 보험계약부채',       13, POINT),
        Item('(6) BS',         '이익잉여금',           18, POINT),
        Item('(6) BS',         '기타포괄손익',         63, POINT),
        Item('(6) BS',         '기타포괄손익',         55, PRIOR_FY),
        Item('(6) BS',         '이익잉여금',           71, POINT),
        Item('(6) BS',         '이익잉여금',           65, PRIOR_FY),
        Item('(6) BS',         '- 해약환급금준비금',   99, POINT),
        Item('(3) UW Income', '예상대비 실제 차이 (+)', 111, section='장기'),
        Item('(3) UW Income', '- 보험금 예실차',        112, section='장기'),
        Item('(3) UW Income', '예상손해액',             113, section='장기'),
        Item('(3) UW Income', '- 사업비 예실차',        115, section='장기'),
        Item('(3) UW Income', '예상사업비',             116, section='장기'),
    ]},

    # DB손해보험 팩트시트는 백만원 단위(unit=1)이고, CSM/BEL movement 와 보종별
    # 손익을 모두 '분기 단독'으로 싣는다 → CUM 모드가 Q1..Qn 을 더해 연간누계를 만든다.
    # CSM(기시)는 항상 1분기 열에서 읽어야 전기말 잔액이 된다(OPEN).
    # Col73(BEL)·Col74(RA)는 팩트시트가 잔여보장부채만 담아 DATA 기준과 달라 제외.
    'DB손해보험': {'unit': 1,
                   'derived': {111: (112, 115, '+')},
                   'items': [
        Item('BEL,CSM변동', 'CSM(기말)',   75, POINT),
        Item('BEL,CSM변동', '신계약 유입', 76, CUM,  section='(2) CSM'),
        Item('BEL,CSM변동', '상각',        77, CUM, -1),
        Item('BEL,CSM변동', 'CSM 조정',    78, CUM),
        Item('BEL,CSM변동', '이자부리',    79, CUM,  section='(2) CSM'),
        Item('BEL,CSM변동', 'CSM(기시)',   80, OPEN),
        Item('보험손익', '보험금 예실차(발생사고요소조정 포함)', 112, CUM),
        Item('보험손익', '사업비 예실차',                        115, CUM),
    ]},

    '삼성화재': {'unit': 100,
                 'derived': {95: (96, 97, '/'), 92: (93, 94, '/')},
                 'items': [
        Item('Capital Ratio',      '가용자본',        96, POINT),
        Item('Capital Ratio',      '요구자본',        97, POINT),
        Item('K-ICS Ratio',        '가용자본',        96, POINT),   # 분기 파일 시트명
        Item('K-ICS Ratio',        '요구자본',        97, POINT),
        Item('Capital Ratio',      '가용자본',        93, POINT),
        Item('Capital Ratio',      '요구자본',        94, POINT),
        Item('K-ICS Ratio',        '가용자본',        93, POINT),
        Item('K-ICS Ratio',        '요구자본',        94, POINT),
        Item('Liability structure', '해약환급금준비금', 99, POINT),
        Item('Liability structure', '기타포괄손익-공정가치측정상품', 25, POINT),
        Item('Liability structure', '장기', 73, POINT, section='최선추정#-1'),
        Item('Liability structure', '장기', 74, POINT, section='위험조정#-1'),
        Item('CSM', '기말 CSM',   75, POINT),
        Item('CSM', '신계약 CSM', 76),
        Item('CSM', 'CSM 상각',   77, CUM, -1),
        Item('CSM', 'CSM 조정',   78),
        Item('CSM', '이자부리',   79),
        Item('CSM', '기말 CSM',   80, PRIOR_FY),
        Item('Profit & Loss Breakdown', '보험금 예실차', 112),
        Item('Profit & Loss Breakdown', '사업비 예실차', 115),
    ]},
}


# ---------------------------------------------------------------- 추출 엔진
QORDER = ['Q1', 'Q2', 'Q3', 'Q4']

def period_to_yq(period):
    """'2506' → (2025, 'Q2'), '2512' → (2025, 'FY')"""
    yy, mm = int(period[:2]), int(period[2:])
    year = 2000 + yy
    return year, ('FY' if mm == 12 else 'Q%d' % (mm // 3))


def extract(factsheet_path, period, items, unit=1000.0):
    """팩트시트 1개에서 해당 기간의 {col: 백만원} 추출"""
    wb = load(factsheet_path)
    year, q = period_to_yq(period)
    result, misses = {}, []

    cache = {}
    for it in items:
        if it.sheet not in wb.sheetnames:
            misses.append((it.col, it.label, 'no sheet'))
            continue
        key = (it.sheet, it.section)
        if key not in cache:
            ws = wb[it.sheet]
            start = 1
            if it.section:
                # '최선추정#2' 처럼 뒤에 #N 을 붙이면 N번째 등장을 앵커로 삼는다.
                # (삼성화재 'Liability structure' 는 '최선추정' 이 두 번 나온다)
                name, _, nth = it.section.partition('#')
                try:
                    nth = int(nth)
                except ValueError:
                    nth = 1
                hits = []
                for r0 in range(1, min(ws.max_row, 400) + 1):
                    if label_row(ws, [name], row_min=r0, row_max=r0).get(name):
                        hits.append(r0)
                if hits:
                    # 음수는 뒤에서부터. 목차/요약에 같은 말이 먼저 나오는 시트가 많아
                    # '마지막 등장'(#-1)이 실제 표를 가리키는 경우가 흔하다.
                    idx = nth - 1 if nth > 0 else nth
                    if -len(hits) <= idx < len(hits):
                        start = hits[idx]
            same = [x.label for x in items
                    if x.sheet == it.sheet and x.section == it.section]
            cache[key] = (ws, period_columns(ws, default_year=year),
                          label_row(ws, same, row_min=start))
        ws, pcols, rows = cache[key]
        r = rows.get(it.label)
        if not r:
            misses.append((it.col, it.label, 'no row'))
            continue

        def cell(yy, pp):
            c = pcols.get((yy, pp))
            v = ws.cell(r, c).value if c else None
            return v if isinstance(v, (int, float)) else None

        val = None
        if it.mode == OPEN:
            # 당해 연도 1분기 열의 '기시' 값 = 전기말 잔액
            # (DB손해보험처럼 분기단독 movement 만 싣는 팩트시트용)
            val = cell(year, 'Q1')
        elif it.mode == PRIOR_FY:
            # 전기말 잔액 = 직전 연도 FY(없으면 4분기말)
            val = cell(year - 1, 'FY')
            if val is None:
                val = cell(year - 1, 'Q4')
        elif it.mode == POINT:
            # 시점값. 연간이면 4분기말 우선(YTD 열이 잔액이 아닌 경우가 있음)
            val = cell(year, 'Q4') if q == 'FY' else cell(year, q)
            if val is None and q == 'FY':
                # 분기 열이 없는 연간 표에서만 FY 열로 넘어간다.
                # (분기 요청에 연말값을 돌려주면 조용히 틀린 값이 들어간다)
                val = cell(year, 'FY')
        else:  # CUM — 연간누계
            if q == 'FY':
                val = cell(year, 'FY')
            if val is None:
                n = 4 if q == 'FY' else QORDER.index(q) + 1
                parts = [cell(year, QORDER[i]) for i in range(n)]
                val = sum(parts) if all(p is not None for p in parts) else None

        if val is None:
            misses.append((it.col, it.label, 'no value'))
        else:
            result[it.col] = val * it.sign * (it.unit if it.unit else unit)
    wb.close()
    # 같은 컬럼을 여러 시트에서 시도했을 때(시트명이 파일마다 다른 경우)
    # 하나라도 성공했으면 나머지 실패는 잡음이므로 지운다.
    misses = [m for m in misses if m[0] not in result]
    return result, misses


def extract_company(company, factsheet_path, period):
    spec = SPECS.get(company)
    if not spec:
        raise KeyError(f'매핑 미정의: {company}')
    res, miss = extract(factsheet_path, period, spec['items'], float(spec['unit']))
    # 파생 컬럼: {대상: (a, b)} → a − b,  {대상: (a, b, '/')} → a ÷ b
    for tgt, rule in (spec.get('derived') or {}).items():
        a, b = rule[0], rule[1]
        op = rule[2] if len(rule) > 2 else '-'
        if a in res and b in res and (op != '/' or res[b]):
            res[tgt] = ({'-': lambda x, y: x - y,
                         '+': lambda x, y: x + y,
                         '/': lambda x, y: x / y}[op])(res[a], res[b])
        else:
            miss.append((tgt, f'파생({a}{op}{b})', 'no source'))
    # 200번대는 파생 계산에만 쓰는 임시 컬럼이라 결과에서 뺀다
    res = {c: v for c, v in res.items() if c < 200}
    miss = [m for m in miss if m[0] < 200]
    return res, miss


if __name__ == '__main__':
    import argparse, json
    ap = argparse.ArgumentParser(description='IR 팩트시트 추출')
    ap.add_argument('company')
    ap.add_argument('factsheet')
    ap.add_argument('period', help='예: 2512')
    a = ap.parse_args()
    res, miss = extract_company(a.company, a.factsheet, a.period)
    print(json.dumps({'values': {f'Col{k}': v for k, v in sorted(res.items())},
                      'misses': miss}, ensure_ascii=False, indent=1))
