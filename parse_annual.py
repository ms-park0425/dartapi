"""
사업보고서 XBRL에서 DATA 시트에 필요한 값을 추출하는 스크립트.
단위: 백만원 (원 / 1e6)

DATA 시트 구조:
- Row 1: 번호 (1,2,3,...)
- Row 2: 제목
- Row 3: 카테고리 (B/S, P/L, ...)
- Row 4: 대분류
- Row 5: 소분류
- Row 6~: 데이터 (기간+회사 조합)

각 행의 Col1 = "YYMM회사약칭" (예: 2512미래)
Col2 = 기간코드 (예: 2512)
Col3 = 회사약칭

열 매핑 (Col → XBRL account_id + context):
"""
import os
import re
import sys
import csv
import xml.etree.ElementTree as ET
from glob import glob
from itertools import combinations

sys.stdout.reconfigure(encoding="utf-8")

NS_XBRLI = "http://www.xbrl.org/2003/instance"
NS_XBRLDI = "http://xbrl.org/2006/xbrldi"

COMPANIES = [
    {"name": "미래에셋생명", "short": "미래"},
    {"name": "삼성생명", "short": "삼성"},
    {"name": "교보생명", "short": "교보"},
    {"name": "한화생명", "short": "한화"},
    {"name": "신한라이프생명", "short": "신한라이프"},
    {"name": "동양생명", "short": "동양"},
    {"name": "KB생명", "short": "KB라이프"},
    {"name": "삼성화재", "short": "삼성화재"},
    {"name": "현대해상", "short": "현대해상"},
    {"name": "메리츠화재", "short": "메리츠화재"},
    {"name": "KB손해보험", "short": "KB손보"},
    {"name": "DB손해보험", "short": "DB손보"},
]

# ====== XBRL → DATA 열 매핑 ======
# BS 항목 (별도, 당기말 instant context)
BS_MAPPING = {
    4: "ifrs-full_Assets",
    5: "ifrs-full_FinancialAssetsAtFairValueThroughProfitOrLoss",
    6: "ifrs-full_FinancialAssetsAtFairValueThroughOtherComprehensiveIncome",  # dart_ fallback: SecuritiesAtFairValueThroughOtherComprehensiveIncome
    # Col7: 상각후원가 자산 - LoansAtAmortisedCost + SecuritiesAtAmortisedCost 합산도 후보
    7: "ifrs-full_FinancialAssetsAtAmortisedCost",
    8: "ifrs-full_InvestmentsInSubsidiariesJointVenturesAndAssociates",
    # fallback in extract_company_data: InvestmentAccountedForUsingEquityMethod
    12: "ifrs-full_Liabilities",
    13: "ifrs-full_InsuranceContractsIssuedThatAreLiabilities",
    15: "ifrs-full_Equity",  # 총자본
    16: "dart_HybridBonds",  # 신종자본증권
    17: "ifrs-full_AccumulatedOtherComprehensiveIncome",
    18: "ifrs-full_RetainedEarnings",
}

# PL 항목 (별도, 당기 duration context)
PL_MAPPING = {
    19: "dart_EquityAtBeginningOfPeriod",  # 기초자본
    20: "dart_EquityAtBeginningOfPeriod",  # 기초자본(신종자본증권제외) - 같은 값
    32: "ifrs-full_InsuranceServiceResult",
    33: "dart_OperatingIncomeInsurance",  # 또는 ifrs-full_OperatingIncomeInsurance
    34: "dart_OperatingExpenseInsurance",  # 또는 ifrs-full_OperatingExpenseInsurance
    35: "dart_InvestmentIncomeExpenses",  # 또는 ifrs-full_InvestmentIncomeExpenses
    39: "ifrs-full_SellingGeneralAndAdministrativeExpense",
    41: "ifrs-full_ProfitLossFromOperatingActivities",
    43: "ifrs-full_ProfitLossBeforeTax",
    44: "ifrs-full_ProfitLoss",
}

# 계산 항목 (두 PL 항목의 차이)
CALC_MAPPING = {
    36: ("dart_InsuranceFinanceIncomeFromInsuranceContractsIssuedRecognisedInProfitOrLoss",
         "dart_InsuranceFinanceExpensesFromInsuranceContractsIssuedRecognisedInProfitOrLoss",
         "subtract"),  # income - expense
    37: ("dart_FinanceIncomeFromReinsuranceContractsHeldRecognisedInProfitOrLoss",
         "dart_FinanceExpensesFromReinsuranceContractsHeldRecognisedInProfitOrLoss",
         "subtract"),
    42: ("ifrs-full_NonOperatingProfitLoss", None, "direct"),
    50: ("ifrs-full_NonOperatingProfitLoss", None, "direct"),
}

# OCI 항목 (별도, 당기 duration) - 각 col에 여러 후보 태그 (순서대로 시도)
OCI_MAPPING = {
    56: [
        "ifrs-full_OtherComprehensiveIncomeNetOfTaxFinancialAssetsMeasuredAtFairValueThroughOtherComprehensiveIncome",
        "ifrs-full_OtherComprehensiveIncomeNetOfTaxFinancialAssetsMeasuredAtFairValueThroughOtherComprehensiveIncomeTotal",
        "ifrs-full_OtherComprehensiveIncomeNetOfTaxChangeInFairValueOfFinancialAssetsMeasuredAtFairValueThroughOtherComprehensiveIncome",
    ],
    # Col59는 OCI_SUM_MAPPING으로 처리 (CashFlowHedges + GainsLossesOnHedgingInstruments 합산)
    60: [
        "ifrs-full_OtherComprehensiveIncomeNetOfTaxGainsLossesOnRevaluation",
        "ifrs-full_OtherComprehensiveIncomeNetOfTaxGainsLossesOnRevaluationOfPropertyPlantAndEquipment",
    ],
    57: ["ifrs-full_OtherComprehensiveIncomeNetOfTaxCreditLossesOfFinancialAssetsMeasuredAtFairValueThroughOtherComprehensiveIncome"],
    61: ["ifrs-full_OtherComprehensiveIncomeNetOfTaxGainsLossesOnRemeasurementsOfDefinedBenefitPlans"],
    62: ["ifrs-full_OtherComprehensiveIncomeNetOfTaxExchangeDifferencesOnTranslation"],
}

# 자본변동표 항목 (1-dim Separate PL duration) - 배당/자기주식/신종자본증권
# Col66: 결산배당 (음수로 저장) - DividendsPaidClassifiedAsFinancingActivities가 총합(Annual+Interim)
DIVIDEND_TAGS = [
    "dart_DividendsPaidClassifiedAsFinancingActivities",
    "ifrs-full_DividendsPaidClassifiedAsFinancingActivities",
    "dart_AnnualDividendsPaid",
    "ifrs-full_AnnualDividendsPaid",
    "ifrs-full_DividendsPaid",
]
# Col70: 신종자본증권 배당 (음수로 저장)
HYBRID_DIVIDEND_TAGS = [
    "dart_DividendToHybridBond",
    "dart_DividendsPaidToHybridBond",
    "ifrs-full_DividendToHybridBond",
    "dart_InterestExpenseForHybridBond",
    "ifrs-full_InterestExpenseForHybridBond",
]
# Col67: 자기주식 소각/처분 (RetainedEarnings에서 차감 → 음수)
TREASURY_TAGS = [
    "ifrs-full_CancellationOfTreasuryShares",
    "dart_CancellationOfTreasuryShares",
]

# Col58 = 보험OCI + 재보험OCI (두 항목 합산)
OCI_SUM_MAPPING = {
    58: (
        "ifrs-full_OtherComprehensiveIncomeNetOfTaxInsuranceFinanceIncomeExpensesFromInsuranceContractsIssuedExcludedFromProfitOrLossThatWillBeReclassifiedToProfitOrLoss",
        "ifrs-full_OtherComprehensiveIncomeNetOfTaxFinanceIncomeExpensesFromReinsuranceContractsHeldExcludedFromProfitOrLoss",
    ),
    59: (
        "ifrs-full_OtherComprehensiveIncomeNetOfTaxCashFlowHedges",
        "dart_OtherComprehensiveIncomeNetOfTaxGainsLossesOnHedgingInstrument",
    ),
}

# 전기말 BS 항목
PRIOR_BS_MAPPING = {
    55: "ifrs-full_AccumulatedOtherComprehensiveIncome",  # 전기말 기타포괄손익누계액
    19: "ifrs-full_Equity",                              # 기초자본 (전기말)
    65: "ifrs-full_RetainedEarnings",                    # 이익잉여금 기초 (전기말)
    71: "ifrs-full_RetainedEarnings",                    # 이익잉여금 기말 (당기말 BS와 동일)
    # 전기말 신종자본증권 (Col20 계산용)
    "_prior_hybrid": "dart_HybridBonds",
}

# OCI 누계 항목 - 당기말 BS context에서 ChangesInAccumulatedOCI axis 멤버 키워드로 탐색
# (Col24~30): 3-dim context (Separate + OtherComprehensiveIncomeLossAccumulatedAmountMember + entity-specific)
OCI_ACCUM_MAPPING = {
    # col: list of keywords (member 이름에 하나라도 포함되면 해당 col에 매핑)
    # 더 구체적인 키워드를 앞에 배치, 순서가 중요함
    24: ["OtherComprehensiveIncome"],
    25: ["FinancialAssets", "GainsAndLossesOnFinancialAssets", "GainsAndLossesFromInvestments",
         "GainsAndLossesOnDebtAndEquity"],  # FVOCI 관련 합산
    26: ["InsuranceFinanceIncomeExpenses", "InsuranceContract"],
    27: ["ReinsuranceFinance", "ReinsuranceContract"],
    28: ["CashFlowHedges", "CashFlow", "Hedging", "ExchangeDifferences"],
    29: ["Revaluation"],          # 전기말 사용
    30: ["RemeasurementsOfDefinedBenefit", "DefinedBenefit"],
}

# 계리적 가정 변경 민감도 분석 (Col146~158)
# 3-dim duration: Separate + InsuranceContractsIssuedMember + BEL or CSM component
ACTUARIAL_MAPPING = {
    # (col, component_keyword): tag_local
    (146, "BEL"): "dart_ChangeInCancellationRateAssumptionOfEffectOfChangingAssumptionsOfChangesInInsuranceLiabilitiesBasedOnActuarialAssumptionsTableOfItems",
    (147, "BEL"): "dart_ChangeInRiskRateAssumptionOfEffectOfChangingAssumptionsOfChangesInInsuranceLiabilitiesBasedOnActuarialAssumptionsTableOfItems",
    (148, "BEL"): "dart_ChangeInProjectRatioAssumptionOfEffectOfChangingAssumptionsOfChangesInInsuranceLiabilitiesBasedOnActuarialAssumptionsTableOfItems",
    (149, "BEL"): "dart_OtherAssumptionChangesOfEffectOfChangingAssumptionsOfChangesInInsuranceLiabilitiesBasedOnActuarialAssumptionsTableOfItems",
    (150, "BEL"): "dart_FluctuationsDueToVolumeDifferencesAndDifferencesInInvestmentFactorsExpectedAndActualOfChangesInFutureServicesAttributableToFactorsOtherThanNewContractsOfChangesInInsuranceLiabilitiesBasedOnActuarialAssumptionsTableOfItems",
    (153, "CSM"): "dart_ChangeInCancellationRateAssumptionOfEffectOfChangingAssumptionsOfChangesInInsuranceLiabilitiesBasedOnActuarialAssumptionsTableOfItems",
    (154, "CSM"): "dart_ChangeInRiskRateAssumptionOfEffectOfChangingAssumptionsOfChangesInInsuranceLiabilitiesBasedOnActuarialAssumptionsTableOfItems",
    (155, "CSM"): "dart_ChangeInProjectRatioAssumptionOfEffectOfChangingAssumptionsOfChangesInInsuranceLiabilitiesBasedOnActuarialAssumptionsTableOfItems",
    (156, "CSM"): "dart_OtherAssumptionChangesOfEffectOfChangingAssumptionsOfChangesInInsuranceLiabilitiesBasedOnActuarialAssumptionsTableOfItems",
    (157, "CSM"): "dart_FluctuationsDueToVolumeDifferencesAndDifferencesInInvestmentFactorsExpectedAndActualOfChangesInFutureServicesAttributableToFactorsOtherThanNewContractsOfChangesInInsuranceLiabilitiesBasedOnActuarialAssumptionsTableOfItems",
    (158, "CSM"): "dart_FluctuationsDueToLossFactorsOfChangesInFutureServicesAttributableToFactorsOtherThanNewContractsOfChangesInInsuranceLiabilitiesBasedOnActuarialAssumptionsTableOfItems",
}

# 추가 PL 항목
EXTRA_PL_MAPPING = {
    68: "ifrs-full_ProfitLoss",          # 순이익 (Col44와 같음)
    48: "ifrs-full_InsuranceServiceResult",  # 보험손익 (Col32와 같음)
    49: "ifrs-full_InvestmentIncomeExpenses",  # 투자손익 (Col35와 같음)
}

# Col40: 기타투자손익 = FeeIncome + RentalIncome + OtherIncomeInvestment - OtherExpenseInvestment
# 태그는 회사마다 다름, 아래 각 항목에 여러 후보
COL40_FEE = ["dart_FeeIncomeOfInvestmentIncome", "dart_FeesAndCommissionIncomeOfInvestmentIncome",
             "ifrs-full_FeeAndCommissionIncome"]
COL40_RENT = ["dart_RentalReturnOfInvestmentIncome", "ifrs-full_RentalIncome"]
COL40_OTHER_INC = ["ifrs-full_OtherOperatingIncomeInvestment", "dart_OtherOperatingIncomeInvestment"]
# 투자비용: OtherOperatingExpenseInvestment + PropertyManagementExpense + udf 계열 모두 포함
# udf_ + OfOperatingExpenseInvestment는 _get_pl_by_keyword가 아닌 별도 처리
COL40_OTHER_EXP_KEYWORDS = ["OtherOperatingExpenseInvestment", "PropertyManagementExpense"]
# DB손보 패턴: udf_IS_ + OfOperatingExpenseInvestment (당해연도 신규 정의)
COL40_UDF_EXP_PREFIX = "OfOperatingExpenseInvestment"

# 복합 축 항목 (CSM 변동표 등)
# 별도 + InsuranceContractsIssuedMember + 각 ComponentMember
INSURANCE_COMPONENT_MAPPING = {
    # Col: (component_member, tag_name) - duration context에서 기말잔액 합산
    # 이건 instant context에서 세부 disaggregation별 합산 필요
}


def parse_actuarial_sensitivity(xbrl_path, target_year="2025", ref_date=None, start_date=None):
    """
    계리적 가정 변경 민감도 분석 파싱.
    3-dim duration: Separate + InsuranceContractsIssuedMember + BEL or CSM component
    반환: {(col, component): value_won}
    """
    import calendar as _cal
    if ref_date is None:
        ref_date = f"{target_year}-12-31"
    if start_date is None:
        start_date = f"{target_year}-01-01"

    tree = ET.parse(xbrl_path)
    root = tree.getroot()

    bel_member = "EstimatesOfPresentValueOfFutureCashFlowsMember"
    csm_members = {
        "ContractualServiceMarginMember",
        "ContractualServiceMarginRelatedToContractsThatExistedAtTransitionDateToWhichModifiedRetrospectiveApproachHasBeenAppliedMember",
        "ContractualServiceMarginNotRelatedToContractsThatExistedAtTransitionDateToWhichModifiedRetrospectiveApproachOrFairValueApproachHasBeenAppliedMember",
        "ContractualServiceMarginRelatedToContractsThatExistedAtTransitionDateToWhichFairValueApproachHasBeenAppliedMember",
    }

    # Build target context dict (3-dim 이상 지원)
    target_ctxs = {}  # ctx_id -> "BEL" or "CSM"
    for ctx in root.findall(f"{{{NS_XBRLI}}}context"):
        period = ctx.find(f"{{{NS_XBRLI}}}period")
        start = period.find(f"{{{NS_XBRLI}}}startDate")
        end = period.find(f"{{{NS_XBRLI}}}endDate")
        if start is None or start.text != start_date:
            continue
        if end is None or end.text != ref_date:
            continue
        members = ctx.findall(f".//{{{NS_XBRLDI}}}explicitMember")
        dims = {}
        for m in members:
            dim_name = m.get("dimension", "").split(":")[-1]
            val = (m.text or "").split(":")[-1]
            dims[dim_name] = val
        if "SeparateMember" != dims.get("ConsolidatedAndSeparateFinancialStatementsAxis", ""):
            continue
        if "InsuranceContractsIssuedMember" != dims.get("DisaggregationOfInsuranceContractsAxis", ""):
            continue
        comp = dims.get("InsuranceContractsByComponentsAxis", "")
        if comp == bel_member:
            target_ctxs[ctx.get("id", "")] = "BEL"
        elif comp == "RiskAdjustmentForNonfinancialRiskMember":
            target_ctxs[ctx.get("id", "")] = "RA"
        elif comp in csm_members:
            target_ctxs[ctx.get("id", "")] = "CSM"

    # tag_local -> list of (col, expected_component_type)
    tag_to_col = {}
    for (col, comp), tag in ACTUARIAL_MAPPING.items():
        tag_local = tag.split("_", 1)[1]
        tag_to_col.setdefault(tag_local, []).append((col, comp))

    # 패턴2: 키워드 기반 context 탐색 (삼성생명 등 entity-specific 태그 지원)
    # 2-dim (Separate + BEL/CSM via DisaggregationAxis or InsuranceContractsByComponentsAxis) + duration
    SENS_KEYWORDS = {
        "BEL": {
            146: ["CancellationRate", "SurrenderRate", "SurrenderRatio", "LapseRate"],
            147: ["RiskRate", "RiskRatio", "RiskFactor", "MortalityRate"],
            148: ["ExpenseRate", "OperatingExpenseRate", "ProjectRatio", "ProjectExpenseRate",
                  "OperatingExpense"],
            149: ["OtherAssumption"],  # OtherChange 제외 (너무 광범위)
            150: ["VolumeDifferences", "QuantityDifferences", "FluctuationsDueTo",
                  "DifferenceOfQuantity", "DifferenceInQuantity", "VolumeDifference"],
        },
        "CSM": {
            153: ["CancellationRate", "SurrenderRate", "SurrenderRatio", "LapseRate"],
            154: ["RiskRate", "RiskRatio", "RiskFactor", "MortalityRate"],
            155: ["ExpenseRate", "OperatingExpenseRate", "ProjectRatio"],
            156: ["OtherAssumption"],
            157: ["VolumeDifferences", "QuantityDifferences", "FluctuationsDueTo",
                  "DifferenceOfQuantity", "DifferenceInQuantity"],
            158: ["LossComponent", "LossFactor"],
        },
    }

    # alt_ctxs: 패턴2 context (삼성생명 2-dim + 현대해상/DB손보/한화 5~6-dim entity-specific)
    alt_ctxs = {}  # ctx_id -> "BEL" or "CSM"
    for ctx in root.findall(f"{{{NS_XBRLI}}}context"):
        period = ctx.find(f"{{{NS_XBRLI}}}period")
        start = period.find(f"{{{NS_XBRLI}}}startDate")
        end = period.find(f"{{{NS_XBRLI}}}endDate")
        if start is None or end is None: continue
        if start.text != start_date or end.text != ref_date: continue
        members = ctx.findall(f".//{{{NS_XBRLDI}}}explicitMember")
        dims = {}
        for m in members:
            dim_name = m.get("dimension", "").split(":")[-1]
            val = (m.text or "").split(":")[-1]
            dims[dim_name] = val
        if "SeparateMember" != dims.get("ConsolidatedAndSeparateFinancialStatementsAxis", ""):
            continue
        # 이미 target_ctxs에 있으면 skip
        ctx_id = ctx.get("id", "")
        if ctx_id in target_ctxs: continue

        comp = dims.get("InsuranceContractsByComponentsAxis", "")
        disagg = dims.get("DisaggregationOfInsuranceContractsAxis", "")

        # 패턴A: 2-dim, DisaggregationAxis=BEL/RA/CSM (삼성생명)
        if len(members) == 2 and not comp:
            if bel_member in disagg:
                alt_ctxs[ctx_id] = "BEL"
            elif "RiskAdjustmentForNonfinancialRiskMember" in disagg:
                alt_ctxs[ctx_id] = "RA"
            elif any(m in disagg for m in csm_members):
                alt_ctxs[ctx_id] = "CSM"
        # 패턴B: 5~6-dim, InsuranceContractsIssuedMember + BEL/CSM component (현대해상/DB손보/한화)
        elif len(members) >= 4 and "InsuranceContractsIssuedMember" == disagg:
            if comp == bel_member:
                alt_ctxs[ctx_id] = "BEL"
            elif comp == "RiskAdjustmentForNonfinancialRiskMember":
                alt_ctxs[ctx_id] = "RA"
            elif comp in csm_members:
                alt_ctxs[ctx_id] = "CSM"

    # Accumulate: standard ACTUARIAL_MAPPING (미래에셋 패턴)
    result = {}
    for elem in root:
        ctx_ref = elem.get("contextRef", "")
        if ctx_ref not in target_ctxs or not elem.text:
            continue
        tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
        if tag not in tag_to_col:
            continue
        try:
            v = float(elem.text.strip())
        except ValueError:
            continue
        comp_type = target_ctxs[ctx_ref]
        for col, expected_comp in tag_to_col[tag]:
            if expected_comp == "BEL" and comp_type in ("BEL", "RA"):
                result[(col, "BEL")] = result.get((col, "BEL"), 0.0) + v
            elif expected_comp == comp_type:
                result[(col, comp_type)] = result.get((col, comp_type), 0.0) + v

    # 키워드 기반 탐색: target_ctxs + alt_ctxs 모두 탐색
    # (ACTUARIAL_MAPPING에 없는 entity-specific 태그 지원)
    keyword_seen = set()  # (ctx_id, col) -> 중복 방지
    all_keyword_ctxs = {**target_ctxs, **alt_ctxs}
    for elem in root:
        ctx_ref = elem.get("contextRef", "")
        if ctx_ref not in all_keyword_ctxs or not elem.text:
            continue
        comp_type = all_keyword_ctxs[ctx_ref]
        tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
        # ACTUARIAL_MAPPING에 있는 태그는 이미 처리됨
        if tag in tag_to_col:
            continue
        try:
            v = float(elem.text.strip())
        except ValueError:
            continue
        if v == 0:
            continue
        # 키워드 매핑 (BEL과 RA는 같은 col에 합산)
        effective_comp = "BEL" if comp_type in ("BEL", "RA") else comp_type
        for col, kwords in SENS_KEYWORDS.get(effective_comp, {}).items():
            if any(k.lower() in tag.lower() for k in kwords):
                pair = (ctx_ref, col)
                if pair not in keyword_seen:
                    keyword_seen.add(pair)
                    result[(col, effective_comp)] = result.get((col, effective_comp), 0.0) + v
                break
    return result


def parse_prior_csm(xbrl_path, target_year="2025"):
    """
    전기말 CSM 합계 파싱. 여러 패턴 지원.
    패턴A(미래에셋): 3-dim entity-disagg + ContractualServiceMarginMember, tag=InsuranceContractsIssuedThatAreLiabilities
    패턴B(삼성생명): 2-dim (Separate+IssuedMember), tag=ContractualServiceMargin
    패턴C(삼성생명 4-dim): TypesOfContracts + BEL/RA 방식에서 total 도출
    """
    prev_year = str(int(target_year) - 1)
    tree = ET.parse(xbrl_path)
    root = tree.getroot()

    # 패턴A: 3-dim entity-specific disagg + ContractualServiceMarginMember
    total_a = 0.0
    target_ctxs_a = set()
    # 패턴B: 2-dim (Separate+IssuedMember) + ContractualServiceMargin tag
    total_b = None

    for ctx in root.findall(f"{{{NS_XBRLI}}}context"):
        period = ctx.find(f"{{{NS_XBRLI}}}period")
        instant = period.find(f"{{{NS_XBRLI}}}instant")
        if instant is None or instant.text != f"{prev_year}-12-31":
            continue
        members = ctx.findall(f".//{{{NS_XBRLDI}}}explicitMember")
        dims = {}
        for m in members:
            dim_name = m.get("dimension", "").split(":")[-1]
            val = (m.text or "").split(":")[-1]
            dims[dim_name] = val
        if "SeparateMember" != dims.get("ConsolidatedAndSeparateFinancialStatementsAxis", ""):
            continue

        # 패턴A: 3-dim entity-specific disagg
        if len(members) == 3:
            if "ContractualServiceMarginMember" != dims.get("InsuranceContractsByComponentsAxis", ""):
                continue
            disagg = dims.get("DisaggregationOfInsuranceContractsAxis", "")
            if disagg in ("InsuranceContractsIssuedMember", "ReinsuranceContractsHeldMember", ""):
                continue
            target_ctxs_a.add(ctx.get("id", ""))

        # 패턴B: 2-dim (Separate + IssuedMember)
        elif len(members) == 2:
            if "InsuranceContractsIssuedMember" == dims.get("DisaggregationOfInsuranceContractsAxis", ""):
                ctx_id = ctx.get("id", "")
                for elem in root:
                    if elem.get("contextRef") != ctx_id or not elem.text:
                        continue
                    tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
                    if tag == "ContractualServiceMargin":
                        try:
                            total_b = float(elem.text.strip())
                        except ValueError:
                            pass

    for elem in root:
        ctx_ref = elem.get("contextRef", "")
        if ctx_ref not in target_ctxs_a or not elem.text:
            continue
        tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
        if tag == "InsuranceContractsIssuedThatAreLiabilities":
            try:
                total_a += float(elem.text.strip())
            except ValueError:
                pass

    if total_a > 0:
        return total_a
    if total_b is not None and total_b > 0:
        return total_b
    return None


def parse_fvoci_oci(xbrl_path, target_year="2025", ref_date=None, start_date=None):
    """
    FVOCI 금융자산 평가손익 OCI 파싱. 여러 context 패턴 지원:
    1. 2-dim (Separate + ReportedAmountMember) + dart_ChangesInOtherComprehensiveIncome...
    2. 2-dim (Separate + AccumulatedOtherComprehensiveIncomeMember) + ifrs-full_...ChangeInFairValue...
    """
    if ref_date is None:
        ref_date = f"{target_year}-12-31"
    if start_date is None:
        start_date = f"{target_year}-01-01"
    tree = ET.parse(xbrl_path)
    root = tree.getroot()

    dart_tag = "ChangesInOtherComprehensiveIncomeFairValueMeasurementGainAndLossOnValuationOfFinancialAssetsEtcOfSignificantTransactionsNotAffectingCashFlowsOfSignificantTransactionsNotAffectingCashFlowsTableOfItems"
    ifrs_tags = [
        "OtherComprehensiveIncomeNetOfTaxChangeInFairValueOfFinancialAssetsMeasuredAtFairValueThroughOtherComprehensiveIncome",
        "OtherComprehensiveIncomeNetOfTaxFinancialAssetsMeasuredAtFairValueThroughOtherComprehensiveIncome",
        "OtherComprehensiveIncomeNetOfTaxFinancialAssetsMeasuredAtFairValueThroughOtherComprehensiveIncomeTotal",
    ]

    for ctx in root.findall(f"{{{NS_XBRLI}}}context"):
        period = ctx.find(f"{{{NS_XBRLI}}}period")
        start = period.find(f"{{{NS_XBRLI}}}startDate")
        end = period.find(f"{{{NS_XBRLI}}}endDate")
        if start is None or start.text != start_date:
            continue
        if end is None or end.text != ref_date:
            continue
        members = ctx.findall(f".//{{{NS_XBRLDI}}}explicitMember")
        if len(members) != 2:
            continue
        dims = {}
        for m in members:
            dim_name = m.get("dimension", "").split(":")[-1]
            val = (m.text or "").split(":")[-1]
            dims[dim_name] = val
        if "SeparateMember" != dims.get("ConsolidatedAndSeparateFinancialStatementsAxis", ""):
            continue

        second_dim_vals = [v for k, v in dims.items() if k != "ConsolidatedAndSeparateFinancialStatementsAxis"]
        is_reported = any("ReportedAmountMember" in v for v in second_dim_vals)
        is_accum_oci = any("AccumulatedOtherComprehensiveIncomeMember" in v for v in second_dim_vals)

        if not is_reported and not is_accum_oci:
            continue

        ctx_id = ctx.get("id", "")
        for elem in root:
            if elem.get("contextRef") != ctx_id or not elem.text:
                continue
            tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
            if is_reported and tag == dart_tag:
                try:
                    return float(elem.text.strip())
                except ValueError:
                    pass
            if is_accum_oci and tag in ifrs_tags:
                try:
                    return float(elem.text.strip())
                except ValueError:
                    pass
    return None


def parse_oci_accumulated(xbrl_path, target_year="2025", ref_date=None):
    """
    OCI 누계 항목 파싱. 세 가지 구조 지원:
    - 패턴A (미래에셋): 3-dim (Separate + OtherComprehensiveIncomeLossAccumulatedAmountMember + entity-specific), tag=Equity
    - 패턴B (삼성생명): 2-dim (Separate + ComponentsOfEquityAxis member), tag=OtherComprehensiveIncomeLossAccumulatedAmount
    - 패턴C (DB손보/현대해상): 2-dim (Separate + CarryingAmountAxis→ReportedAmountMember), individual ifrs-full reserve tags
    반환: {"cur": {keyword: value_won}, "prior": {keyword: value_won},
           "cur_direct": {col: value_won}, "prior_direct": {col: value_won}}
    cur_direct/prior_direct는 패턴C에서 col번호로 직접 매핑된 값.
    """
    if ref_date is None:
        ref_date = f"{target_year}-12-31"
    prev_year = str(int(target_year) - 1)
    prior_date = f"{prev_year}-12-31"

    tree = ET.parse(xbrl_path)
    root = tree.getroot()

    result = {"cur": {}, "prior": {}, "cur_direct": {}, "prior_direct": {}}

    # 패턴C: CarryingAmountAxis → ReportedAmountMember 에서 개별 ifrs-full reserve 태그 직접 읽기
    # (DB손보, 현대해상 패턴)
    # tag → col 매핑 (ifrs-full 로컬 이름 → col번호)
    PATC_TAG_TO_COL = {
        # Col26: 보험계약 금융손익 OCI
        "ReserveOfInsuranceFinanceIncomeExpensesFromInsuranceContractsIssuedExcludedFromProfitOrLossThatWillBeReclassifiedToProfitOrLoss": 26,
        # Col27: 재보험계약 금융손익 OCI
        "ReserveOfFinanceIncomeExpensesFromReinsuranceContractsHeldExcludedFromProfitOrLoss": 27,
        # Col28: 현금흐름위험회피 OCI + 해외사업환산 (OCI_ACCUM_MAPPING col28 keywords 포함)
        "ReserveOfCashFlowHedges": 28,
        "ReserveOfCashFlowHedgesContinuingHedges": 28,
        "ReserveOfExchangeDifferencesOnTranslation": 28,
        # Col29: 재평가잉여금
        "RevaluationSurplus": 29,
        # Col30: 확정급여부채 재측정
        "ReserveOfRemeasurementsOfDefinedBenefitPlans": 30,
    }
    # Col25 FVOCI: entity-specific dart tag with keyword → col 25
    PATC_FVOCI_KEYWORDS = (
        "ReserveOfGainsAndLossesOnFinancialAssetsAtFairValueThroughOtherComprehensiveIncome",
        "ReserveOfGainsAndLossesOnDebtInstrumentsAtFairValueThroughOtherComprehensiveIncome",
        "ReserveOfGainsAndLossesFromInvestmentsInEquityInstruments",
        "ReserveOfGainsAndLossesOnFinancialAssets",
    )
    # dart hedge reserve tags for Col28 fallback
    PATC_HEDGE_KEYWORDS = (
        "ReserveOfGainsAndLossesOnHedgingDerivatives",
        "ReserveOfGainsAndLossesOnHedgingInstrument",
        "HedgingReserve",
    )

    for ctx in root.findall(f"{{{NS_XBRLI}}}context"):
        period = ctx.find(f"{{{NS_XBRLI}}}period")
        instant = period.find(f"{{{NS_XBRLI}}}instant")
        if instant is None:
            continue
        if instant.text == ref_date:
            bucket = "cur"
        elif instant.text == prior_date:
            bucket = "prior"
        else:
            continue

        members = ctx.findall(f".//{{{NS_XBRLDI}}}explicitMember")
        dims = {}
        for m in members:
            dim_name = m.get("dimension", "").split(":")[-1]
            val = (m.text or "").split(":")[-1]
            dims[dim_name] = val
        if "SeparateMember" != dims.get("ConsolidatedAndSeparateFinancialStatementsAxis", ""):
            continue

        # 패턴A: OtherComprehensiveIncomeLossAccumulatedAmountMember 포함 + 3-dim
        if len(members) == 3 and any("OtherComprehensiveIncomeLossAccumulatedAmountMember" in v for v in dims.values()):
            change_member = ""
            for k, v in dims.items():
                if "Changes" in k or ("Accumulated" in k and "Separate" not in k):
                    change_member = v
            if change_member:
                ctx_id = ctx.get("id", "")
                for elem in root:
                    if elem.get("contextRef") != ctx_id or not elem.text:
                        continue
                    tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
                    if tag == "Equity":
                        try:
                            result[bucket][change_member] = float(elem.text.strip())
                        except ValueError:
                            pass

        # 패턴B: 2-dim with ComponentsOfEquityAxis (OCI 관련 reserve members) - 반드시 2-dim
        elif len(members) == 2 and instant is not None:
            oci_member = ""
            for k, v in dims.items():
                if k in ("ComponentsOfEquityAxis", "ReservesWithinEquityAxis"):
                    oci_member = v
            if not oci_member:
                # 패턴C: CarryingAmountAxis → ReportedAmountMember
                carry_val = dims.get("CarryingAmountAccumulatedDepreciationAmortisationAndImpairmentAndGrossCarryingAmountAxis", "")
                if "ReportedAmountMember" not in carry_val:
                    continue
                ctx_id = ctx.get("id", "")
                direct_bucket = f"{bucket}_direct"
                fvoci_total = 0.0
                hedge_val = 0.0
                for elem in root:
                    if elem.get("contextRef") != ctx_id or not elem.text:
                        continue
                    tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
                    try:
                        v = float(elem.text.strip())
                    except ValueError:
                        continue
                    if tag in PATC_TAG_TO_COL:
                        col = PATC_TAG_TO_COL[tag]
                        # Col28은 여러 항목 합산 (CFH + 해외사업환산 등)
                        if col == 28:
                            result[direct_bucket][col] = result[direct_bucket].get(col, 0.0) + v
                        elif col not in result[direct_bucket]:
                            result[direct_bucket][col] = v
                    elif any(kw in tag for kw in PATC_FVOCI_KEYWORDS):
                        fvoci_total += v
                    elif any(kw in tag for kw in PATC_HEDGE_KEYWORDS):
                        hedge_val += v
                if fvoci_total != 0.0 and 25 not in result[direct_bucket]:
                    result[direct_bucket][25] = fvoci_total
                if hedge_val != 0.0:
                    # Col28에 hedge_val 합산 (ExchangeDifferences가 이미 있을 수 있음)
                    result[direct_bucket][28] = result[direct_bucket].get(28, 0.0) + hedge_val
                continue

            # Reserve of FVOCI / Insurance OCI / Reinsurance OCI / CFH / Revaluation / DefinedBenefit
            if not any(kw in oci_member for kw in ("GainsAndLossesOnFinancialAssets", "InsuranceFinance",
                                                     "ReinsuranceFinance", "CashFlowHedges",
                                                     "Revaluation", "RemeasurementsOfDefinedBenefit",
                                                     "ReserveOfGainsAndLosses", "ExchangeDifferences",
                                                     "Hedging")):
                continue
            ctx_id = ctx.get("id", "")
            # ComponentsOfEquityAxis가 있는 context의 값을 우선
            oci_axis = dims.get("ComponentsOfEquityAxis") or dims.get("ReservesWithinEquityAxis", "")
            PREF_TAGS_B = (
                "OtherComprehensiveIncomeLossAccumulatedAmount",
                "AccumulatedOtherComprehensiveIncome",
                "Equity",
            )
            best_val = None
            best_prio = 99
            comp_axis_used = "ComponentsOfEquityAxis" in dims
            for elem in root:
                if elem.get("contextRef") != ctx_id or not elem.text:
                    continue
                tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
                try:
                    v = float(elem.text.strip())
                except ValueError:
                    continue
                if tag in PREF_TAGS_B:
                    prio = PREF_TAGS_B.index(tag)
                    if prio < best_prio:
                        best_val = v
                        best_prio = prio
            if best_val is not None:
                # ComponentsOfEquityAxis 버전을 우선 (덮어쓰기 허용)
                if comp_axis_used or oci_member not in result[bucket]:
                    result[bucket][oci_member] = best_val
    return result


def parse_csm_movement(xbrl_path, target_year="2025", ref_date=None, start_date=None):
    """
    CSM 변동 항목 파싱. 여러 dim 구조 지원 (5-dim, 4-dim, 3-dim, 2-dim).
    반환: {tag_local_name: total_value_won}
    """
    if ref_date is None:
        ref_date = f"{target_year}-12-31"
    if start_date is None:
        start_date = f"{target_year}-01-01"
    tree = ET.parse(xbrl_path)
    root = tree.getroot()

    csm_tags = {
        "InsuranceRevenueContractualServiceMarginRecognisedInProfitOrLossBecauseOfTransferOfServices": "csm_amortization",
        "IncreaseDecreaseThroughChangesInEstimatesThatAdjustContractualServiceMarginInsuranceContractsLiabilityAsset": "csm_adjustment",
        "InsuranceFinanceIncomeExpensesFromInsuranceContractsIssuedRecognisedInProfitOrLoss": "csm_finance",
        "IncreaseDecreaseThroughEffectsOfContractsInitiallyRecognisedInPeriodInsuranceContractsLiabilityAsset": "new_csm",
        # 삼성생명 패턴
        "IncreaseDecreaseThroughRecognitionOfContractualServiceMarginInProfitOrLossToReflectTransferOfServicesInsuranceContractsLiabilityAsset": "csm_amortization",
        # 이자부리 대체 태그 (한화/교보/현대해상 등)
        "IncreaseDecreaseThroughInsuranceFinanceIncomeOrExpensesInsuranceContractsLiabilityAsset": "csm_finance",
        "IncreaseDecreaseThroughEffectsOfContractsAcquiredInPeriodInsuranceContractsLiabilityAsset": "new_csm",
    }
    component_keys = {
        "ContractualServiceMarginMember",
        "ContractualServiceMarginRelatedToContractsThatExistedAtTransitionDateToWhichModifiedRetrospectiveApproachHasBeenAppliedMember",
        "ContractualServiceMarginNotRelatedToContractsThatExistedAtTransitionDateToWhichModifiedRetrospectiveApproachOrFairValueApproachHasBeenAppliedMember",
        "ContractualServiceMarginRelatedToContractsThatExistedAtTransitionDateToWhichFairValueApproachHasBeenAppliedMember",
    }

    # 어떤 dim 구조인지 먼저 파악 (4-dim or 3-dim)
    has_4dim = False
    has_5dim = False
    has_3dim = False
    has_6dim = False
    for ctx in root.findall(f"{{{NS_XBRLI}}}context"):
        period = ctx.find(f"{{{NS_XBRLI}}}period")
        start = period.find(f"{{{NS_XBRLI}}}startDate")
        end = period.find(f"{{{NS_XBRLI}}}endDate")
        if start is None or start.text != start_date: continue
        if end is None or end.text != ref_date: continue
        members = ctx.findall(f".//{{{NS_XBRLDI}}}explicitMember")
        dims = {m.get("dimension", "").split(":")[-1]: (m.text or "").split(":")[-1] for m in members}
        if "SeparateMember" != dims.get("ConsolidatedAndSeparateFinancialStatementsAxis", ""): continue
        if "InsuranceContractsIssuedMember" != dims.get("DisaggregationOfInsuranceContractsAxis", ""): continue
        if dims.get("InsuranceContractsByComponentsAxis", "") not in component_keys: continue
        if len(dims) == 6: has_6dim = True
        elif len(dims) == 5: has_5dim = True
        elif len(dims) == 4: has_4dim = True
        elif len(dims) == 3: has_3dim = True

    # dim 선택 전략:
    # 6-dim 있으면: 6-dim만 (삼성화재 패턴)
    # 4-dim 있으면: 4-dim (상각/신계약) + 5-dim (조정/이자부리) 병행 - 미래에셋 패턴
    # 3-dim만 있으면: 3-dim (삼성생명 패턴)
    target_ndims = set()
    if has_6dim:
        target_ndims.add(6)
    elif has_4dim:
        target_ndims.add(4)
        if has_5dim:
            target_ndims.add(5)  # 미래에셋: 5-dim에도 일부 태그 있음
    elif has_3dim:
        target_ndims.add(3)
    if not target_ndims:
        target_ndims = {3, 4, 5}

    target_ctxs = set()
    for ctx in root.findall(f"{{{NS_XBRLI}}}context"):
        period = ctx.find(f"{{{NS_XBRLI}}}period")
        start = period.find(f"{{{NS_XBRLI}}}startDate")
        end = period.find(f"{{{NS_XBRLI}}}endDate")
        if start is None or start.text != start_date: continue
        if end is None or end.text != ref_date: continue
        members = ctx.findall(f".//{{{NS_XBRLDI}}}explicitMember")
        dims = {m.get("dimension", "").split(":")[-1]: (m.text or "").split(":")[-1] for m in members}
        if "SeparateMember" != dims.get("ConsolidatedAndSeparateFinancialStatementsAxis", ""): continue
        if "InsuranceContractsIssuedMember" != dims.get("DisaggregationOfInsuranceContractsAxis", ""): continue
        comp = dims.get("InsuranceContractsByComponentsAxis", "")
        if comp not in component_keys: continue
        if len(dims) in target_ndims:
            target_ctxs.add(ctx.get("id", ""))

    # 태그별 합산: 각 컴포넌트 레이블에서 합산
    # 4-dim+5-dim 병행 시 상각/신계약 태그는 4-dim만, 조정/이자 태그는 5-dim에 있는 경우 해당값 사용
    AMORT_TAGS = {
        "InsuranceRevenueContractualServiceMarginRecognisedInProfitOrLossBecauseOfTransferOfServices",
        "IncreaseDecreaseThroughRecognitionOfContractualServiceMarginInProfitOrLossToReflectTransferOfServicesInsuranceContractsLiabilityAsset",
        "IncreaseDecreaseThroughChangesThatRelateToCurrentServiceInsuranceContractsLiabilityAsset",
    }
    NEW_CSM_TAGS = {
        "IncreaseDecreaseThroughEffectsOfContractsInitiallyRecognisedInPeriodInsuranceContractsLiabilityAsset",
        "IncreaseDecreaseThroughEffectsOfContractsAcquiredInPeriodInsuranceContractsLiabilityAsset",
    }

    sums = {v: 0.0 for v in csm_tags.values()}
    # 상각/신계약: 소규모 dim에서 합산 (4 또는 가장 작은 target_ndim)
    # 조정/이자: 모든 target_ndim에서 합산 (단, 상쇄 방지)
    min_ndim = min(target_ndims) if target_ndims else 4
    min_ctxs = set()
    all_ctxs = set()

    for ctx_ref, comp_key in [(c, None) for c in target_ctxs]:
        pass  # rebuild below

    # ctx_id → ndim 매핑 (target_ctxs + adj용 all_sep_ctxs)
    ctx_ndim = {}
    # csm_adjustment 전용: 다양한 context 패턴 지원
    all_sep_ctxs = set()  # adj_tag용
    for ctx in root.findall(f"{{{NS_XBRLI}}}context"):
        period = ctx.find(f"{{{NS_XBRLI}}}period")
        start = period.find(f"{{{NS_XBRLI}}}startDate")
        end = period.find(f"{{{NS_XBRLI}}}endDate")
        if start is None or start.text != start_date: continue
        if end is None or end.text != ref_date: continue
        ctx_id = ctx.get("id", "")
        members = ctx.findall(f".//{{{NS_XBRLDI}}}explicitMember")
        dims_t = {m.get("dimension", "").split(":")[-1]: (m.text or "").split(":")[-1] for m in members}
        if "SeparateMember" != dims_t.get("ConsolidatedAndSeparateFinancialStatementsAxis", ""): continue
        # 패턴1: DisaggregationAxis=InsuranceContractsIssuedMember + InsuranceContractsByComponentsAxis=CSM
        # 패턴2(삼성생명): DisaggregationAxis=ContractualServiceMarginMember (2-dim)
        is_pattern1 = (
            "InsuranceContractsIssuedMember" == dims_t.get("DisaggregationOfInsuranceContractsAxis", "")
            and dims_t.get("InsuranceContractsByComponentsAxis", "") in component_keys
        )
        is_pattern2 = (
            "ContractualServiceMarginMember" == dims_t.get("DisaggregationOfInsuranceContractsAxis", "")
            and "InsuranceContractsByComponentsAxis" not in dims_t
        )
        if is_pattern1:
            ctx_ndim[ctx_id] = len(members)
            all_sep_ctxs.add(ctx_id)
            if ctx_id in target_ctxs and len(members) == min_ndim:
                min_ctxs.add(ctx_id)
        elif is_pattern2:
            ctx_ndim[ctx_id] = len(members)
            all_sep_ctxs.add(ctx_id)

    # (context_id, tag) 쌍으로 중복 방지
    # 추가: 같은 tag + comp_key 조합에서 중복 합산 방지 (낮은 ndim 우선)
    seen = set()
    # comp_key별로 값 수집: (comp_key, value) 리스트
    comp_values = {v: [] for v in csm_tags.values()}  # comp_key -> [(ndim, value)]

    adj_tag = "IncreaseDecreaseThroughChangesInEstimatesThatAdjustContractualServiceMarginInsuranceContractsLiabilityAsset"

    for elem in root:
        ctx_ref = elem.get("contextRef", "")
        if not elem.text: continue
        tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
        if tag not in csm_tags: continue
        # adj_tag는 all_sep_ctxs에서 수집, 나머지는 target_ctxs에서
        if tag == adj_tag:
            if ctx_ref not in all_sep_ctxs: continue
        else:
            if ctx_ref not in target_ctxs: continue
        pair = (ctx_ref, tag)
        if pair in seen: continue
        seen.add(pair)
        comp_key = csm_tags[tag]
        ndim = ctx_ndim.get(ctx_ref, 99)
        try:
            v = float(elem.text.strip())
        except ValueError:
            continue
        # 같은 context에서 같은 comp_key는 한 번만 합산
        ctx_comp_pair = (ctx_ref, comp_key)
        if ctx_comp_pair in seen: continue
        if tag in AMORT_TAGS or tag in NEW_CSM_TAGS:
            if ctx_ref in min_ctxs:
                seen.add(ctx_comp_pair)
                comp_values[comp_key].append((ndim, v))
        else:
            seen.add(ctx_comp_pair)
            comp_values[comp_key].append((ndim, v))

    # 각 comp_key에 대해 합산 전략:
    # - csm_adjustment: 가장 낮은 ndim의 합산만 사용 (상위 집계값이 정답)
    # - 나머지: 같은 절댓값 중복 제거 후 합산
    adj_tag_key = "csm_adjustment"
    for comp_key, vals in comp_values.items():
        if not vals:
            continue
        if comp_key == adj_tag_key:
            # 조정: 최소 ndim에서 ContractualServiceMarginMember(집계) 값이 있으면 그것 하나만,
            # 없으면 최소 ndim의 sub-component 합산
            min_nd = min(nd for nd, _ in vals)
            min_nd_vals = [(nd, v) for nd, v in vals if nd == min_nd]
            # ContractualServiceMarginMember context는 별도로 표시됨
            # 집계값(단일)과 세부값들이 같은 ndim에 있을 때 → 집계값 하나만 사용
            # 집계값 = 세부값들의 합과 같음 → 집계값이 있으면 집계값만
            total_min = sum(v for _, v in min_nd_vals)
            # 값이 2개 이상이고 그 중 하나가 전체합과 같으면 그것 하나만 사용
            for _, v in min_nd_vals:
                if abs(v - total_min / 2) < abs(total_min) * 0.01 and len(min_nd_vals) > 1:
                    # 약 절반인 값이 있으면 집계값 = 절반 → 그것만 사용
                    sums[comp_key] = v
                    break
            else:
                sums[comp_key] = total_min
        elif comp_key in ("new_csm", "csm_amortization"):
            # 신계약/상각: 최소 ndim에서 합산하되,
            # 단일 집계값이 세부값들의 합과 같으면 세부값들만 사용
            if vals:
                min_nd = min(nd for nd, _ in vals)
                min_vals = [v for nd, v in vals if nd == min_nd]
                total_min = sum(min_vals)
                # 집계값이 세부값 합과 같으면(같은 ndim에 중복 저장) → 더 작은 세부값들만 합산
                # 예: [2037561, 1386, 2036175] → 2037561 = 1386+2036175 → 세부값 합 사용
                # 하나의 값이 나머지 합과 같으면 그것은 집계값
                non_dup = []
                for v in min_vals:
                    rest_sum = sum(x for x in min_vals if x != v or min_vals.count(x) > 1)
                    if abs(abs(v) - abs(total_min - v)) < abs(v) * 0.001:
                        # v ≈ total - v → v ≈ total/2 → 중복 없음 (이 경우는 단순 합산)
                        pass
                    # 하나가 나머지 합과 같은지 체크
                if len(min_vals) > 1:
                    for i, v in enumerate(min_vals):
                        others_sum = sum(x for j, x in enumerate(min_vals) if j != i)
                        if abs(abs(v) - abs(others_sum)) < abs(v) * 0.01:
                            # v = sum of others → v is the aggregate, exclude it
                            non_dup = [x for j, x in enumerate(min_vals) if j != i]
                            break
                    else:
                        non_dup = min_vals
                else:
                    non_dup = min_vals
                sums[comp_key] = sum(non_dup)
        elif comp_key == "csm_finance":
            # 이자부리: 최소 ndim에서만 합산 (여러 ndim에 같은 값이 중복될 수 있음)
            if vals:
                min_nd = min(nd for nd, _ in vals)
                for nd, v in vals:
                    if nd == min_nd:
                        sums[comp_key] += v
        else:
            # 나머지: 단순 합산
            for _, v in vals:
                sums[comp_key] += v
    return sums


def find_xbrl_file(corp_name, report_type="사업보고서", year="2025"):
    """XBRL 파일 경로 찾기. 없으면 DART API에서 자동 다운로드."""
    base = os.path.join("data", corp_name, report_type, year)
    files = glob(os.path.join(base, "*.xbrl"))
    if files:
        return files[0]

    # 자동 다운로드
    corp_code = ALL_CORP_CODES.get(corp_name)
    if not corp_code:
        return None
    try:
        from dart_api import get_disclosures, download_xbrl
        reprt_code_map = {"사업보고서": "11011", "반기보고서": "11012",
                          "1분기보고서": "11013", "3분기보고서": "11014"}
        reprt_code = reprt_code_map.get(report_type, "11011")
        # 보고서 제출 시기별 검색 범위
        if report_type == "사업보고서":
            next_year = str(int(year) + 1)
            data = get_disclosures(corp_code=corp_code, bgn_de=f"{next_year}0101",
                                   end_de=f"{next_year}0731", page_count=100)
        elif report_type == "1분기보고서":
            # 1분기보고서: 4~6월 제출
            data = get_disclosures(corp_code=corp_code, bgn_de=f"{year}0401",
                                   end_de=f"{year}0630", page_count=100)
        elif report_type == "반기보고서":
            # 반기보고서: 7~9월 제출
            data = get_disclosures(corp_code=corp_code, bgn_de=f"{year}0701",
                                   end_de=f"{year}0930", page_count=100)
        elif report_type == "3분기보고서":
            # 3분기보고서: 10~12월 제출
            data = get_disclosures(corp_code=corp_code, bgn_de=f"{year}1001",
                                   end_de=f"{year}1231", page_count=100)
        else:
            data = get_disclosures(corp_code=corp_code, bgn_de=f"{year}0101",
                                   end_de=f"{year}1231", page_count=100)
        items = data.get("list", [])
        # 보고서 유형별 키워드 매칭
        # reprt_code로 필터링: DART API list에 보고서코드 없음 → 제목으로 구분
        month_kw = {
            "사업보고서":  ["사업보고서"],
            "반기보고서":  ["반기보고서"],
            "1분기보고서": ["분기보고서 (2025.03)", "분기보고서 (2024.03)", "분기보고서 (2023.03)",
                           "분기보고서 (2022.03)", f"분기보고서 ({year}.03)"],
            "3분기보고서": ["분기보고서 (2025.09)", "분기보고서 (2024.09)", "분기보고서 (2023.09)",
                           f"분기보고서 ({year}.09)"],
        }
        kws = month_kw.get(report_type, [report_type])
        candidates = [i for i in items if any(kw in i.get("report_nm", "") for kw in kws)]
        originals = [c for c in candidates if "정정" not in c.get("report_nm", "")]
        selected = (originals or candidates)
        if not selected:
            return None
        selected = max(selected, key=lambda x: x.get("rcept_dt", ""))
        rcept_no = selected["rcept_no"]
        print(f"    XBRL 다운로드: {corp_name} {selected['report_nm']}")
        save_path = download_xbrl(rcept_no, corp_name=corp_name,
                                  reprt_name=report_type, year=year,
                                  reprt_code=reprt_code)
        files = glob(os.path.join(save_path, "*.xbrl"))
        return files[0] if files else None
    except Exception as e:
        print(f"    XBRL 다운로드 실패 ({corp_name}): {e}")
        return None


def parse_simple_contexts(xbrl_path, target_year="2025", ref_date=None, month="12"):
    """
    보고서 XBRL 파싱.
    ref_date: BS 기준일 (기본 {year}-{month}-말일)
    month: 결산월 ("03","06","09","12")

    반환: {"bs": {account_id: value_won}, "pl": {account_id: value_won}, "prior_bs": {...}}
    """
    import calendar as _cal
    if ref_date is None:
        last_day = _cal.monthrange(int(target_year), int(month))[1]
        ref_date = f"{target_year}-{month}-{last_day:02d}"

    # PL 기간: 1분기=1~3월, 반기=1~6월, 3분기=1~9월, 연간=1~12월
    pl_start = f"{target_year}-01-01"
    pl_end = ref_date

    # 전기말: 직전 연도 12-31 (또는 동일연도 직전 분기말)
    prev_year = str(int(target_year) - 1)
    prior_date = f"{prev_year}-12-31"

    tree = ET.parse(xbrl_path)
    root = tree.getroot()

    # 단순 context (ConsolidatedAndSeparate 축 1개만) 찾기
    bs_ctx_id = None  # instant 당기말
    pl_ctx_id = None  # duration 누적
    prior_bs_ctx_id = None  # instant 전기말

    for ctx in root.findall(f"{{{NS_XBRLI}}}context"):
        ctx_id = ctx.get("id", "")
        members = ctx.findall(f".//{{{NS_XBRLDI}}}explicitMember")
        if len(members) != 1:
            continue
        dim = members[0].get("dimension", "")
        val = (members[0].text or "").strip()
        is_separate_axis = (
            ("ConsolidatedAndSeparateFinancialStatements" in dim and "Separate" in val)
            or ("StatementInformationAxis" in dim and "SeparateMember" in val)
        )
        if not is_separate_axis:
            continue

        period = ctx.find(f"{{{NS_XBRLI}}}period")
        instant = period.find(f"{{{NS_XBRLI}}}instant")
        start = period.find(f"{{{NS_XBRLI}}}startDate")
        end = period.find(f"{{{NS_XBRLI}}}endDate")

        if instant is not None:
            if instant.text == ref_date:
                if bs_ctx_id is None or "ConsolidatedAndSeparate" in ctx_id:
                    bs_ctx_id = ctx_id
            elif instant.text == prior_date:
                if prior_bs_ctx_id is None or "ConsolidatedAndSeparate" in ctx_id:
                    prior_bs_ctx_id = ctx_id
        elif start is not None and end is not None:
            if start.text == pl_start and end.text == pl_end:
                if pl_ctx_id is None or "ConsolidatedAndSeparate" in ctx_id:
                    pl_ctx_id = ctx_id

    if not bs_ctx_id:
        print(f"  WARNING: BS context not found in {xbrl_path}")
    if not pl_ctx_id:
        print(f"  WARNING: PL context not found in {xbrl_path}")

    bs_data = {}
    pl_data = {}
    prior_bs_data = {}

    target_ctxs = {bs_ctx_id, pl_ctx_id, prior_bs_ctx_id} - {None}

    for elem in root:
        tag = elem.tag
        ctx_ref = elem.get("contextRef", "")
        if not elem.text or not elem.text.strip():
            continue
        if ctx_ref not in target_ctxs:
            continue

        if "}" not in tag:
            continue
        ns_uri, local = tag[1:].split("}", 1)
        if "ifrs-full" in ns_uri or "ifrs_full" in ns_uri:
            prefix = "ifrs-full"
        elif "dart" in ns_uri and "dart-gcd" not in ns_uri:
            prefix = "dart"
        else:
            continue

        account_id = f"{prefix}_{local}"
        val = elem.text.strip()
        try:
            v = float(val)
        except ValueError:
            continue

        if ctx_ref == bs_ctx_id:
            bs_data[account_id] = v
        elif ctx_ref == pl_ctx_id:
            pl_data[account_id] = v
        elif ctx_ref == prior_bs_ctx_id:
            prior_bs_data[account_id] = v

    return {"bs": bs_data, "pl": pl_data, "prior_bs": prior_bs_data}


def parse_insurance_components(xbrl_path, target_year="2025", ref_date=None):
    """
    보험계약부채 구성요소별 잔액 합산 (CSM, BEL, RA).
    여러 XBRL 구조 지원:
    - 패턴A (미래에셋): 5-dim (Separate+IssuedMember+InsContractsAxis+TypesOfContracts+Component), tag=InsuranceContractsThatAreLiabilities
    - 패턴B (삼성생명): 4-dim (Separate+IssuedMember+TypesOfContracts+Component), tag=InsuranceContractsThatAreLiabilities
                       + 2-dim CSM (Separate+IssuedMember), tag=ContractualServiceMargin
    """
    if ref_date is None:
        ref_date = f"{target_year}-12-31"
    tree = ET.parse(xbrl_path)
    root = tree.getroot()

    component_sums = {"CSM": 0.0, "BEL": 0.0, "RA": 0.0}

    component_map = {
        "ContractualServiceMarginMember": "CSM",
        "ContractualServiceMarginRelatedToContractsThatExistedAtTransitionDateToWhichModifiedRetrospectiveApproachHasBeenAppliedMember": "CSM",
        "ContractualServiceMarginNotRelatedToContractsThatExistedAtTransitionDateToWhichModifiedRetrospectiveApproachOrFairValueApproachHasBeenAppliedMember": "CSM",
        "ContractualServiceMarginRelatedToContractsThatExistedAtTransitionDateToWhichFairValueApproachHasBeenAppliedMember": "CSM",
        "EstimatesOfPresentValueOfFutureCashFlowsMember": "BEL",
        "RiskAdjustmentForNonfinancialRiskMember": "RA",
    }

    # 분류 축 이름들 (회사별로 다름)
    CLASSIFICATION_AXES = {
        "TypesOfContractsAxis",
        "TypesOfInsuranceContractsAxis",
        "ClassificationOfInsuranceProductsAxisOfDisclosureOfReconciliationOfChangesInInsuranceContractsIssuedByComponentsTableOfAxis",
        "ClassificationOfInsuranceContractsByPresenceOrAbsenceOfParticipationFeaturesAxisOfDisclosureOfDetailedInformationAboutConcentrationsOfRiskThatArisesFromContractsWithinScopeOfIfrs17OfAxis",
        "NatureOfInsuranceContractsAxisOfDisclosureOfReconciliationOfChangesInInsuranceContractsIssuedByComponentsTableOfAxis",
        "PortfolioClassificationAxisOfDisclosureOfInsuranceLiabilitiesByAccountingModelAndPortfolioOfAxis",
        "TypesOfInsuranceContractsByDividendOfDisclosureOfReconciliationOfChangesInInsuranceContractsIssuedByRemainingCoverageAndIncurredClaimsTableOfAxis",
    }
    # 분류 축이지만 중복 계층(집계 레벨)을 만드는 축들 - 제외
    EXCLUDE_AXES = {
        "InsuranceContractsByRemainingCoverageAndIncurredClaimsAxis",
    }

    target_ctxs = {}
    csm_2dim_ctx = None  # 2-dim CSM (삼성생명 패턴)

    for ctx in root.findall(f"{{{NS_XBRLI}}}context"):
        ctx_id = ctx.get("id", "")
        period = ctx.find(f"{{{NS_XBRLI}}}period")
        instant = period.find(f"{{{NS_XBRLI}}}instant")
        if instant is None or instant.text != ref_date:
            continue

        members = ctx.findall(f".//{{{NS_XBRLDI}}}explicitMember")
        dims = {}
        for m in members:
            dim_name = m.get("dimension", "").split(":")[-1]
            val = (m.text or "").split(":")[-1]
            dims[dim_name] = val

        if "SeparateMember" != dims.get("ConsolidatedAndSeparateFinancialStatementsAxis", ""):
            continue
        if "InsuranceContractsIssuedMember" != dims.get("DisaggregationOfInsuranceContractsAxis", ""):
            continue

        component_val = dims.get("InsuranceContractsByComponentsAxis", "")

        # 공통: InsuranceContractsIssuedMember + 분류축 1개 이상 + ComponentAxis
        # dim 수는 4~6개, 반드시 분류축이 1개 이상 있어야 함
        has_classification = any(ax in dims for ax in CLASSIFICATION_AXES)

        # 중복 계층 축이 있으면 제외
        has_exclude = any(ax in dims for ax in EXCLUDE_AXES)
        if component_val in component_map and has_classification and not has_exclude:
            target_ctxs[ctx_id] = (component_map[component_val], len(dims))

        # 패턴B CSM: 2-dim (Separate + IssuedMember), tag=ContractualServiceMargin
        elif len(dims) == 2 and component_val == "":
            csm_2dim_ctx = ctx_id

    # BEL/RA/CSM 합산: 올바른 dim 레벨 자동 선택
    # InsuranceContractsThatAreLiabilities 우선, 없으면 InsuranceContractsLiabilityAsset
    PREF_TAG = "InsuranceContractsThatAreLiabilities"
    FALLBACK_TAG = "InsuranceContractsLiabilityAsset"

    # dim별, 태그별로 분리 (InsuranceContractsThatAreAssets는 절댓값으로 합산)
    ASSET_TAG = "InsuranceContractsThatAreAssets"
    by_ndim = {}  # ndim -> {"pref": {comp: sum}, "fallback": {comp: sum}}
    for ctx_ref, (comp_key, ndim) in target_ctxs.items():
        if ndim not in by_ndim:
            by_ndim[ndim] = {
                "pref": {"CSM": 0.0, "BEL": 0.0, "RA": 0.0},
                "fallback": {"CSM": 0.0, "BEL": 0.0, "RA": 0.0},
            }
        for elem in root:
            if elem.get("contextRef") != ctx_ref or not elem.text:
                continue
            tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
            try:
                v = float(elem.text.strip())
                if tag == PREF_TAG:
                    by_ndim[ndim]["pref"][comp_key] += v
                elif tag == ASSET_TAG:
                    # Assets는 별도 누적 (나중에 pref와 합산 여부 결정)
                    by_ndim[ndim].setdefault("assets", {"CSM": 0.0, "BEL": 0.0, "RA": 0.0})
                    by_ndim[ndim]["assets"][comp_key] += abs(v)
                elif tag == FALLBACK_TAG:
                    by_ndim[ndim]["fallback"][comp_key] += v
            except ValueError:
                pass

    if by_ndim:
        # pref 태그로 합산이 있는 ndim 중 최솟값 선택
        chosen = None
        chosen_ndim = None
        for ndim in sorted(by_ndim.keys()):
            pref = by_ndim[ndim]["pref"]
            if any(pref[k] != 0 for k in ["BEL", "RA", "CSM"]):
                chosen = dict(pref)
                chosen_ndim = ndim
                break
        if chosen is None:
            for ndim in sorted(by_ndim.keys()):
                fallback = by_ndim[ndim]["fallback"]
                if any(fallback[k] != 0 for k in ["BEL", "RA", "CSM"]):
                    chosen = dict(fallback)
                    chosen_ndim = ndim
                    break

        if chosen is not None:
            assets = by_ndim.get(chosen_ndim, {}).get("assets", {"CSM": 0.0, "BEL": 0.0, "RA": 0.0})
            # 각 component별로 Assets 추가 여부를 독립적으로 결정
            # pref + assets가 정답이 되는 조건: assets < pref * 6% (DB손보 CSM 패턴)
            # pref가 이미 정답인 경우: assets 제외
            for key in ["BEL", "RA", "CSM"]:
                pref_v = chosen[key]
                asset_v = assets.get(key, 0.0)
                if asset_v == 0 or pref_v == 0:
                    component_sums[key] = pref_v
                elif abs(asset_v) < abs(pref_v) * 0.06:
                    # assets가 pref의 6% 미만 → 별도 항목 (DB손보 CSM: 667924/11537419≈0.058)
                    component_sums[key] = pref_v + asset_v
                else:
                    # assets가 더 큰 비율 → 중복 제외
                    component_sums[key] = pref_v

    # 패턴B CSM: ContractualServiceMargin 태그 (2-dim context)
    # 이미 패턴A/B에서 CSM을 합산했으면 이 값은 무시 (중복 방지)
    if component_sums["CSM"] == 0.0 and csm_2dim_ctx:
        for elem in root:
            if elem.get("contextRef") != csm_2dim_ctx or not elem.text:
                continue
            tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
            if tag == "ContractualServiceMargin":
                try:
                    component_sums["CSM"] = float(elem.text.strip())
                except ValueError:
                    pass

    # 패턴B CSM fallback: 2-dim (Separate+IssuedMember) ContractualServiceMargin 합산
    if component_sums["CSM"] == 0.0:
        for ctx in root.findall(f"{{{NS_XBRLI}}}context"):
            ctx_id = ctx.get("id", "")
            period = ctx.find(f"{{{NS_XBRLI}}}period")
            instant = period.find(f"{{{NS_XBRLI}}}instant")
            if instant is None or instant.text != ref_date:
                continue
            members = ctx.findall(f".//{{{NS_XBRLDI}}}explicitMember")
            dims = {}
            for m in members:
                dim_name = m.get("dimension", "").split(":")[-1]
                val = (m.text or "").split(":")[-1]
                dims[dim_name] = val
            if "SeparateMember" != dims.get("ConsolidatedAndSeparateFinancialStatementsAxis", ""):
                continue
            if "InsuranceContractsIssuedMember" != dims.get("DisaggregationOfInsuranceContractsAxis", ""):
                continue
            if len(dims) != 2:
                continue
            for elem in root:
                if elem.get("contextRef") != ctx_id or not elem.text:
                    continue
                tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
                if tag == "ContractualServiceMargin":
                    try:
                        v = float(elem.text.strip())
                        if v != 0:
                            component_sums["CSM"] = v
                    except ValueError:
                        pass

    return component_sums


# 1Q document.xml fallback용 corp_code 매핑 (XBRL에 component axis 없는 회사)
DOC_FALLBACK_CORP_CODES = {
    "교보생명": "00112882",
    "신한라이프생명": "00137517",
    "KB생명": "00160393",
    "메리츠화재": "00117744",
    "KB손해보험": "00120216",
}

# 손보사 여부 (PAA 제외 대상)
NON_LIFE_COMPANIES = {"삼성화재", "현대해상", "메리츠화재", "KB손해보험", "DB손해보험"}

# 사업보고서 document.xml에서 CSM 변동 추출이 필요한 회사 (XBRL에 axis 없음)
# 같은 corp_code 사용
DOC_CSM_MOVEMENT_COMPANIES = DOC_FALLBACK_CORP_CODES.copy()

# 전체 회사 corp_code (사업보고서 doc.xml 다운로드용)
ALL_CORP_CODES = {
    "미래에셋생명": "00112332",
    "삼성생명":    "00126256",
    "교보생명":    "00112882",
    "한화생명":    "00113058",
    "신한라이프생명": "00137517",
    "동양생명":    "00117267",
    "KB생명":     "00160393",
    "삼성화재":    "00139214",
    "현대해상":    "00164973",
    "메리츠화재":  "00117744",
    "KB손해보험":  "00120216",
    "DB손해보험":  "00159102",
}


def parse_insurance_components_from_doc(corp_name, target_sum, year="2025"):
    """
    1분기보고서 document.xml에서 BEL/RA/CSM 추출 (XBRL fallback).
    target_sum: Col13 보험계약부채 (백만원 단위, 1e6 곱해서 원 단위로 비교)
    반환: {"BEL": float, "RA": float, "CSM": float} (원 단위)
    """
    corp_code = DOC_FALLBACK_CORP_CODES.get(corp_name)
    if not corp_code:
        return {"CSM": 0.0, "BEL": 0.0, "RA": 0.0}

    # 1Q report year = target_year + 1
    q1_year = str(int(year) + 1)
    cache_path = os.path.join("data", corp_name, f"doc_1q_{q1_year}.xml")
    os.makedirs(os.path.dirname(cache_path), exist_ok=True)

    if not os.path.exists(cache_path):
        try:
            from dart_api import find_rcept_no, _fetch_document_content
            rcept_no = find_rcept_no(corp_code, q1_year, "11013")
            content = _fetch_document_content(rcept_no)
            os.makedirs("data", exist_ok=True)
            with open(cache_path, "w", encoding="utf-8") as f:
                f.write(content)
            print(f"    document.xml 다운로드: {cache_path}")
        except Exception as e:
            print(f"    document.xml 다운로드 실패 ({corp_name}): {e}")
            return {"CSM": 0.0, "BEL": 0.0, "RA": 0.0}

    # target_sum은 백만원 단위 → 원 단위로 변환하지 않음 (doc.xml도 백만원)
    bel, ra, csm = _extract_components_from_doc(cache_path, target_sum)
    if bel is not None:
        return {"BEL": bel * 1e6, "RA": ra * 1e6, "CSM": csm * 1e6}
    return {"CSM": 0.0, "BEL": 0.0, "RA": 0.0}


def _extract_components_from_doc_q1(filepath):
    """분기보고서 document.xml에서 당기말 잔액 테이블의 BEL/RA/CSM 추출.

    패턴A (교보/신한라이프/KB생명/메리츠화재/KB손보):
      헤더: 구분|포트폴리오|일반모형|[변동수수료접근법]|[보험료배분접근법]|최선추정부채|위험조정|보험계약마진|...
      합계행: 합 계|BEL1|RA1|CSM1|[BEL2|RA2|CSM2]|[PAA_total]

    패턴B (한화생명/동양생명/삼성화재):
      헤더: 발행한 보험계약|보험료배분접근법 제외|구성요소별|BEL|RA|CSM|합계
      기말행: 기말 보험계약부채(자산) 또는 기말 장부금액 | BEL|RA|CSM|합계

    패턴C (삼성생명):
      헤더: 발행한 보험계약|일반모형|VFA|BEL_GM|RA_GM|CSM_GM|BEL_VFA|RA_VFA|CSM_VFA
      데이터행: 보험계약부채(자산)|BEL_GM|RA_GM|CSM_GM|BEL_VFA|RA_VFA|CSM_VFA

    패턴D (현대해상):
      헤더: 발행한 보험계약|잔여보장부채|구성요소별|BEL|RA|CSM
      데이터행: 보험계약부채(자산)|BEL|RA|CSM  (원 단위)

    패턴E (DB손해보험):
      테이블1 생명보험계약: 기말 장부금액|BEL|RA|PAA잔여|PAA손실|합계  (BEL/RA만)
      테이블2 비생명보험계약: 기말 장부금액|BEL|RA|CSM(하위분류들)|합계  (BEL+RA+CSM)
      두 테이블의 BEL/RA 합산, CSM은 비생명보험만

    패턴F (미래에셋생명):
      텍스트에서 '최선추정부채 N조 N,NNN 억원, 위험조정 N,NNN 억원, 보험계약마진 N조 N,NNN 억원' 파싱

    단위 감지:
      - 테이블 내부 (단위 : 억원) → × 100 → 백만원
      - 원 단위 (큰 숫자 또는 '단위 : 원') → / 1e6 → 백만원

    반환: (bel_백만원, ra_백만원, csm_백만원) 또는 (None, None, None)
    """
    try:
        with open(filepath, encoding="utf-8") as f:
            content = f.read()
    except Exception:
        return None, None, None

    table_pattern = re.compile(r"<TABLE[^>]*>(.*?)</TABLE>", re.DOTALL | re.IGNORECASE)

    def _detect_unit(text, tm_start, fallback_range=3000):
        """단위 감지: 1=백만원, 100=억원(×100→백만원), 1e-6=원(/1e6→백만원)"""
        unit_in_table = re.search(r"단위\s*[:\s]\s*(억원|백만원|원)", text)
        if unit_in_table:
            u = unit_in_table.group(1)
            if "억원" in u:
                return 100
            if "원" in u and "백만" not in u:
                return 1e-6
            return 1
        before = content[max(0, tm_start - fallback_range): tm_start]
        before_text = re.sub(r"<[^>]+>", " ", before)
        unit_before = re.search(r"단위\s*[:\s]\s*(억원|백만원|원)", before_text)
        if unit_before:
            u = unit_before.group(1)
            if "억원" in u:
                return 100
            if "원" in u and "백만" not in u:
                return 1e-6
        return 1

    def _parse_cells(row_fragment, stop_on_label=False):
        """셀 문자열 → 숫자 리스트 (None=빈칸/대시).
        stop_on_label=True: 한글/영문 레이블 셀을 만나면 파싱 중단.
        """
        cells = row_fragment.split("|")[1:]
        nums = []
        started = False  # 첫 숫자가 나왔는지 여부
        for cell in cells:
            cell = cell.strip()
            if not cell or cell in ("-", "- "):
                if stop_on_label and started:
                    nums.append(None)
                elif not stop_on_label:
                    nums.append(None)
                continue
            neg_m = re.match(r"^\(([0-9,]+)\)$", cell)
            pos_m = re.match(r"^-?([0-9,]+)$", cell)
            if neg_m:
                started = True
                nums.append(-int(neg_m.group(1).replace(",", "")))
            elif pos_m:
                started = True
                v = int(cell.replace(",", "").replace("-", "") or "0")
                if cell.startswith("-"):
                    v = -v
                nums.append(v)
            else:
                # 한글/영문 레이블 → stop_on_label이면 중단
                if stop_on_label and started:
                    break
                nums.append(None)
        return nums

    # ── 패턴A: 기존 로직 (교보/신한라이프/KB생명/메리츠화재/KB손보) ──────────
    # 연결+별도 두 테이블이 있을 경우 별도(마지막) 테이블을 써야 하므로 모두 수집 후 마지막 반환
    pattern_a_candidates = []
    for tm in table_pattern.finditer(content):
        tbl_raw = tm.group(1)
        text = re.sub(r"<[^>]+>", "|", tbl_raw)
        text = re.sub(r"\|[\s|]+\|", "|", text)
        text = re.sub(r"\s+", " ", text).strip()

        header = text[:600]
        if not ("최선추정부채" in header or "미래 현금흐름의 현재가치추정치" in header):
            continue
        if "재보험" in header[:300] or "출재" in header[:300]:
            continue
        if "보험계약마진" not in header:
            continue

        sum_row_match = re.search(r"합\s*계\s*\|([^|].*?)(?:\||$)", text)
        if not sum_row_match:
            continue

        unit_factor = _detect_unit(text, tm.start())

        row_start = text.find("합 계")
        if row_start == -1:
            row_start = text.find("합계")
        row_fragment = text[row_start:row_start + 400]
        nums = _parse_cells(row_fragment)
        valid_nums = [n for n in nums if n is not None]
        if len(valid_nums) < 3:
            continue

        has_vfa = "변동수수료" in header
        if has_vfa and len(valid_nums) >= 6:
            bel = valid_nums[0] + valid_nums[3]
            ra  = valid_nums[1] + valid_nums[4]
            csm = valid_nums[2] + valid_nums[5]
        else:
            bel = valid_nums[0]
            ra  = valid_nums[1]
            csm = valid_nums[2]

        if csm < 0:
            continue

        pattern_a_candidates.append((bel * unit_factor, ra * unit_factor, csm * unit_factor))

    if pattern_a_candidates:
        # 연결 테이블이 먼저, 별도 테이블이 나중에 오는 구조 → 마지막(별도) 사용
        return pattern_a_candidates[-1]

    # ── 패턴B: 기말 장부금액 / 기말 보험계약부채 행 파싱 ──────────────────────
    # (한화생명, 동양생명, 삼성화재)
    # 헤더에 '현재가치 추정치' (공백 포함) + '보험계약마진' + '기말'
    # 단일 테이블에 기초→기말 변동 구조. 첫 번째 매칭만 사용 (당분기).
    for tm in table_pattern.finditer(content):
        tbl_raw = tm.group(1)
        text = re.sub(r"<[^>]+>", "|", tbl_raw)
        text = re.sub(r"\|[\s|]+\|", "|", text)
        text = re.sub(r"\s+", " ", text).strip()

        header = text[:800]
        if "현재가치 추정치" not in header and "현재가치추정치" not in header:
            continue
        if "보험계약마진" not in header:
            continue
        if "재보험" in header[:300] or "출재" in header[:300]:
            continue
        # 기말 행이 있어야 함
        if "기말" not in text:
            continue
        # 기초 행도 있어야 변동표 (잔액 테이블 제외)
        if "기초" not in text:
            continue
        # 포트폴리오별 세부 테이블 제외 (미래에셋 패턴 - 계약의 유형 분류 있음)
        # 이런 테이블은 합계가 아닌 개별 포트폴리오 → 패턴F(텍스트)로 처리
        if "계약의 유형" in header[:500]:
            continue

        # 기말 보험계약부채(자산) 행 우선, 없으면 기말 장부금액 행
        # 우선순위: 기말 장부금액 > 기말 보험계약부채(자산) 집계행
        # (DB손보 패턴: '기말 보험계약부채' 뒤에 '자산인...' 레이블 행이 먼저 나와서
        #  stop_on_label로 자산값이 잡힐 수 있음 → '기말 장부금액' 직접 행이 더 안전)
        end_kw = None
        end_idx = -1
        for kw in ["기말 장부금액", "기말 보험계약부채"]:
            idx = text.rfind(kw)
            if idx >= 0:
                # stop_on_label로 첫 번째 숫자 시퀀스만 추출 (다음 레이블 행 제외)
                row_frag = text[idx:idx + 400]
                nums = _parse_cells(row_frag, stop_on_label=True)
                valid_nums = [n for n in nums if n is not None]
                if not valid_nums or max(abs(v) for v in valid_nums) < 1000:
                    continue  # 소액 행 (자산인 보험계약 등) → 건너뜀
                # 첫 번째 값이 양수여야 BEL (음수면 자산인 보험계약 값이 섞인 것)
                if valid_nums[0] < 0:
                    continue
                end_kw = kw
                end_idx = idx
                break
        if end_kw is None:
            continue

        unit_factor = _detect_unit(text, tm.start())

        row_frag = text[end_idx:end_idx + 400]
        nums = _parse_cells(row_frag, stop_on_label=True)
        valid_nums = [n for n in nums if n is not None]
        if len(valid_nums) < 3:
            continue

        # 구조: [BEL|RA|CSM_하위분류들|합계] 또는 [BEL|RA|CSM|합계]
        # 헤더에 VFA가 있으면 CSM 하위분류가 여럿 → 합산
        # 합계가 마지막 값
        has_vfa = "변동수수료" in header
        if has_vfa:
            # BEL=0, RA=1, CSM_sub들=2~-2, 합계=-1
            bel = valid_nums[0]
            ra  = valid_nums[1]
            total = valid_nums[-1]
            csm = total - bel - ra
        else:
            # [BEL|RA|CSM|합계] 또는 [BEL|RA|CSM하위1|CSM하위2|...|합계]
            bel = valid_nums[0]
            ra  = valid_nums[1]
            total = valid_nums[-1]
            csm = total - bel - ra

        if csm < 0:
            # 동양생명: CSM 하위분류가 수정소급법+공정가치법+기타 → 3개 합산
            # 구조: [BEL|RA|CSM_수정|CSM_공정|CSM_기타|합계]
            if len(valid_nums) >= 5:
                csm = sum(valid_nums[2:-1])
            if csm < 0:
                continue

        return bel * unit_factor, ra * unit_factor, csm * unit_factor

    # ── 패턴C: 삼성생명 (일반모형+VFA 잔액 테이블) ────────────────────────────
    # 헤더: 보험계약|일반모형|VFA|BEL|RA|CSM|BEL|RA|CSM
    # 데이터행: 보험계약부채(자산)|BEL_GM|RA_GM|CSM_GM|BEL_VFA|RA_VFA|CSM_VFA
    for tm in table_pattern.finditer(content):
        tbl_raw = tm.group(1)
        text = re.sub(r"<[^>]+>", "|", tbl_raw)
        text = re.sub(r"\|[\s|]+\|", "|", text)
        text = re.sub(r"\s+", " ", text).strip()

        header = text[:800]
        if "현재가치 추정치" not in header and "현재가치추정치" not in header:
            continue
        if "보험계약마진" not in header:
            continue
        if "재보험" in header[:300] or "출재" in header[:300]:
            continue
        # 삼성생명 패턴: 일반모형 + 변동수수료접근법 헤더 + 보험계약부채(자산) 데이터행
        if "일반모형" not in header or "변동수수료접근법" not in header:
            continue
        if "보험계약부채(자산)" not in text:
            continue

        # 계약 유형/약정 유형 등 세부 분류 없는 단순 합계 테이블
        if "계약의 유형" in header or "약정의 유형" in header:
            continue

        unit_factor = _detect_unit(text, tm.start())

        # '보험계약부채(자산)' 행 찾기
        idx = text.find("보험계약부채(자산)")
        if idx < 0:
            continue
        row_frag = text[idx:idx + 400]
        nums = _parse_cells(row_frag)
        valid_nums = [n for n in nums if n is not None]
        if len(valid_nums) < 6:
            continue

        # [BEL_GM|RA_GM|CSM_GM|BEL_VFA|RA_VFA|CSM_VFA]
        bel = valid_nums[0] + valid_nums[3]
        ra  = valid_nums[1] + valid_nums[4]
        csm = valid_nums[2] + valid_nums[5]

        if csm < 0:
            continue

        return bel * unit_factor, ra * unit_factor, csm * unit_factor

    # ── 패턴D: 현대해상 (원 단위 단일 잔액 테이블) ────────────────────────────
    # 헤더: 발행한 보험계약|잔여보장부채와 발생사고부채|잔여보장부채|구성요소별|BEL|RA|CSM
    # 데이터: 보험계약부채(자산)|BEL|RA|CSM  (숫자 정확히 3개)
    # 특징: '잔여보장부채와 발생사고부채' 포함 + 기초/기말 없음 + 정확히 3개 숫자
    for tm in table_pattern.finditer(content):
        tbl_raw = tm.group(1)
        text = re.sub(r"<[^>]+>", "|", tbl_raw)
        text = re.sub(r"\|[\s|]+\|", "|", text)
        text = re.sub(r"\s+", " ", text).strip()

        header = text[:800]
        if "현재가치 추정치" not in header and "현재가치추정치" not in header:
            continue
        if "보험계약마진" not in header:
            continue
        if "재보험" in header[:300] or "출재" in header[:300]:
            continue
        # 현대해상 특유 헤더: '잔여보장부채와 발생사고부채' 포함
        if "잔여보장부채와 발생사고부채" not in header:
            continue
        # 변동표가 아닌 잔액 테이블: 기초/기말 행 없음
        if "기초" in text or "기말" in text:
            continue
        if "보험계약부채(자산)" not in text:
            continue
        # 미래 현금유출액 테이블 제외 (미래에셋 스타일 세부 분류)
        if "미래 현금유출액의 현재가치 추정치" in header:
            continue

        idx = text.find("보험계약부채(자산)")
        if idx < 0:
            continue
        row_frag = text[idx:idx + 400]
        nums = _parse_cells(row_frag, stop_on_label=True)
        valid_nums = [n for n in nums if n is not None]
        if len(valid_nums) < 3:
            continue
        # 정확히 3개 또는 4개(합계포함) 숫자여야 함 (세부 분류 테이블 제외)
        if len(valid_nums) > 5:
            continue

        # 원 단위 감지 (큰 값)
        unit_factor = _detect_unit(text, tm.start())
        if unit_factor == 1:
            # 값이 1e10 이상이면 원 단위로 판단
            if max(abs(v) for v in valid_nums if v is not None) > 1e10:
                unit_factor = 1e-6

        bel = valid_nums[0]
        ra  = valid_nums[1]
        csm = valid_nums[2]

        if csm < 0:
            continue

        return bel * unit_factor, ra * unit_factor, csm * unit_factor

    # ── 패턴E: DB손해보험 (생명+비생명 두 테이블 합산) ────────────────────────
    # 생명보험계약 테이블: 기말 장부금액|BEL|RA|PAA잔여|PAA손실|합계 (CSM없음)
    # 비생명보험계약 테이블: 기말 장부금액|BEL|RA|CSM하위분류들|합계
    life_bel = life_ra = non_life_bel = non_life_ra = non_life_csm = None

    for tm in table_pattern.finditer(content):
        tbl_raw = tm.group(1)
        text = re.sub(r"<[^>]+>", "|", tbl_raw)
        text = re.sub(r"\|[\s|]+\|", "|", text)
        text = re.sub(r"\s+", " ", text).strip()

        header = text[:800]
        if "현재가치 추정치" not in header and "현재가치추정치" not in header:
            continue
        if "재보험" in header[:300] or "출재" in header[:300]:
            continue
        if "기말" not in text:
            continue

        is_life = "생명보험계약" in header[:500] and "비생명" not in header[:300]
        is_non_life = "비생명보험계약" in header[:500]
        if not is_life and not is_non_life:
            continue

        unit_factor = _detect_unit(text, tm.start())

        idx = text.rfind("기말 장부금액")
        if idx < 0:
            idx = text.rfind("기말 보험계약부채")
        if idx < 0:
            continue

        row_frag = text[idx:idx + 400]
        nums = _parse_cells(row_frag, stop_on_label=True)
        valid_nums = [n for n in nums if n is not None]
        if len(valid_nums) < 3:
            continue

        # 소액 행 (자산인 보험계약) 제외
        if max(abs(v) for v in valid_nums) < 1000:
            continue

        if is_life and life_bel is None:
            # 생명보험: BEL|RA|PAA잔여|PAA손실|합계 (CSM없음)
            life_bel = valid_nums[0] * unit_factor
            life_ra  = valid_nums[1] * unit_factor

        if is_non_life and non_life_bel is None:
            if "보험계약마진" not in header:
                continue
            # 비생명보험: BEL|RA|CSM하위분류들|합계
            bel_nl = valid_nums[0]
            ra_nl  = valid_nums[1]
            total_nl = valid_nums[-1]
            csm_nl = total_nl - bel_nl - ra_nl
            if csm_nl < 0:
                csm_nl = sum(valid_nums[2:-1])
            if csm_nl >= 0:
                non_life_bel = bel_nl * unit_factor
                non_life_ra  = ra_nl  * unit_factor
                non_life_csm = csm_nl * unit_factor

    if non_life_bel is not None:
        bel_total = (non_life_bel or 0) + (life_bel or 0)
        ra_total  = (non_life_ra  or 0) + (life_ra  or 0)
        csm_total = non_life_csm or 0
        return bel_total, ra_total, csm_total

    # ── 패턴F: 미래에셋 (텍스트에서 조억원 파싱) ──────────────────────────────
    # '보험부채는 최선추정부채 N조 N,NNN 억원, 위험조정 N,NNN 억원, 보험계약마진 N조 N,NNN 억원'
    # 또는 '최선추정부채 N조 N,NNN 억원' 단독
    # 억원 → 백만원: × 100
    def _parse_jo_ek(s):
        """'N조 N,NNN 억원' 또는 'N,NNN 억원' → 백만원"""
        m = re.search(r"([\d,]+)\s*조\s*([\d,]+)\s*억원", s)
        if m:
            return (int(m.group(1).replace(",", "")) * 10000
                    + int(m.group(2).replace(",", ""))) * 100
        m2 = re.search(r"([\d,]+)\s*억원", s)
        if m2:
            return int(m2.group(1).replace(",", "")) * 100
        return None

    idx = content.find("최선추정부채")
    while idx >= 0:
        region = content[idx:idx + 300]
        clean = re.sub(r"<[^>]+>", " ", region)
        clean = re.sub(r"\s+", " ", clean).strip()
        if "억원" in clean:
            bel_v = _parse_jo_ek(clean)
            ra_v  = None
            csm_v = None
            # 위험조정
            m_ra = re.search(r"위험조정\s+([\d,]+)\s*억원", clean)
            if m_ra:
                ra_v = int(m_ra.group(1).replace(",", "")) * 100
            # 보험계약마진
            m_csm = re.search(r"보험계약마진\s+([\d,]+)\s*조\s*([\d,]+)\s*억원", clean)
            if m_csm:
                csm_v = (int(m_csm.group(1).replace(",", "")) * 10000
                         + int(m_csm.group(2).replace(",", ""))) * 100
            else:
                m_csm2 = re.search(r"보험계약마진\s+([\d,]+)\s*억원", clean)
                if m_csm2:
                    csm_v = int(m_csm2.group(1).replace(",", "")) * 100
            if bel_v and ra_v and csm_v:
                return bel_v, ra_v, csm_v
        idx = content.find("최선추정부채", idx + 1)

    return None, None, None


def _get_q1_doc_path(corp_name, year):
    """분기보고서 doc.xml 경로 반환. 없으면 다운로드 시도. 실패시 None."""
    cache_path = os.path.join("data", corp_name, f"doc_1q_{year}.xml")
    if os.path.exists(cache_path):
        return cache_path

    # 다운로드 시도 (ALL_CORP_CODES에 있는 모든 회사)
    corp_code = ALL_CORP_CODES.get(corp_name)
    if not corp_code:
        return None
    try:
        from dart_api import find_rcept_no, _fetch_document_content
        os.makedirs(os.path.dirname(cache_path), exist_ok=True)
        rcept_no = find_rcept_no(corp_code, year, "11013")
        content = _fetch_document_content(rcept_no)
        with open(cache_path, "w", encoding="utf-8") as f:
            f.write(content)
        print(f"    doc_1q_{year}.xml 다운로드 완료: {cache_path}")
        return cache_path
    except Exception as e:
        print(f"    doc_1q_{year}.xml 다운로드 실패 ({corp_name}): {e}")
        return None


def _extract_components_from_doc(filepath, target_sum):
    """document.xml 변동표에서 기초 BEL/RA/CSM 추출. target_sum은 백만원 단위.
    생보사: target_sum ≈ grand합 (0.5% 이내)
    손보사: target_sum > grand합 (PAA 부분이 Col13에 포함되므로 20% 이내)

    Tables in document.xml come in 당분기/전분기 pairs. We take only 당분기 (first of each pair).
    The document may have 연결+별도 sections; we use the LAST occurrence of each section
    (별도 = separate FS, which is what Col13/73/74/75 reference).
    """
    try:
        with open(filepath, encoding="utf-8") as f:
            content = f.read()
    except Exception:
        return None, None, None

    table_pattern = re.compile(r'<TABLE[^>]*>(.*?)</TABLE>', re.DOTALL | re.IGNORECASE)
    all_matches = []

    for tm in table_pattern.finditer(content):
        tbl = tm.group(1)
        text = re.sub(r'<[^>]+>', '|', tbl)
        text = re.sub(r'\|[\s|]+\|', '|', text)
        text = re.sub(r'\s+', ' ', text)

        if '재보험' in text[:200] or '출재' in text[:200]:
            continue
        header = text[:600]
        if not ('현재가치' in header and '보험계약마진' in header):
            continue
        if '잔여보장' in header and '발생사고' in header:
            continue

        m = re.search(r'기초[^|]*순장부금액\|(.*?)(?:\|(?:미래|보험수익|보험서비스|신계약))', text)
        if not m:
            m = re.search(r'(?:1\.\s*)?기초 보험계약자산 및 부채\|(.*?)(?:\|\s*\(1\)|\|미래|\|보험수익)', text)
        if not m:
            m = re.search(r'기초\|보험계약부채\|(.*?)(?:\|미래|\|기말)', text)
        if not m:
            continue

        row_str = m.group(1)
        nums = [int(n.replace(',', '')) for n in re.findall(r'[\d,]+', row_str)]
        if len(nums) < 3:
            continue

        bel = nums[0]
        ra = nums[1]
        grand = nums[-1]
        csm = grand - bel - ra

        if csm < 0:
            for g in reversed(nums[2:-1]):
                c = g - bel - ra
                if c >= 0 and g > max(bel, ra):
                    csm = c
                    grand = g
                    break
            else:
                continue

        all_matches.append({'bel': bel, 'ra': ra, 'csm': csm, 'grand': grand, 'pos': tm.start()})

    if not all_matches:
        return None, None, None

    # Split into sections by large position gaps (연결 vs 별도)
    sections = []
    current_section = [all_matches[0]]
    for i in range(1, len(all_matches)):
        gap = all_matches[i]['pos'] - all_matches[i - 1]['pos']
        if gap > 500000:
            sections.append(current_section)
            current_section = [all_matches[i]]
        else:
            current_section.append(all_matches[i])
    sections.append(current_section)

    # Use last section (별도 재무제표)
    basis_rows = sections[-1]

    # Filter out reinsurance asset tables (small relative to section max)
    if len(basis_rows) > 2:
        max_grand = max(r['grand'] for r in basis_rows)
        basis_rows = [r for r in basis_rows if r['grand'] > max_grand * 0.03]

    # Section contains N tables: half are 당분기, half are 전분기.
    # Two common layouts:
    #   Alternating: [당A, 전A, 당B, 전B] → even indices = 당분기
    #   Grouped: [당A, 당B, 당C, 전A, 전B, 전C] → first half = 당분기
    n = len(basis_rows)
    half = n // 2 if n >= 4 else n

    # Generate candidate subsets (당분기 half)
    candidates = []
    if n >= 4:
        # Pattern 1: alternating → even indices
        even_idx = tuple(range(0, n, 2))
        candidates.append(even_idx)
        # Pattern 2: grouped → first half
        first_half = tuple(range(half))
        if first_half != even_idx:
            candidates.append(first_half)

    # Also try all combinations of size 1..half for robustness
    best = None
    best_err = float('inf')

    # Check structured candidates first
    for combo in candidates:
        g = sum(basis_rows[i]['grand'] for i in combo)
        err = abs(g - target_sum)
        if err < best_err:
            best_err = err
            best = combo

    # If structured candidates give exact match, use it
    if best and best_err < target_sum * 0.005:
        sel = [basis_rows[i] for i in best]
        return sum(r['bel'] for r in sel), sum(r['ra'] for r in sel), sum(r['csm'] for r in sel)

    # Exhaustive search for exact match (all sizes up to half)
    for size in range(1, half + 1):
        if size > n:
            continue
        for combo in combinations(range(n), size):
            g = sum(basis_rows[i]['grand'] for i in combo)
            err = abs(g - target_sum)
            if err < best_err:
                best_err = err
                best = combo
            if err < 100:
                break
        if best_err < 100:
            break

    if best and best_err < target_sum * 0.005:
        sel = [basis_rows[i] for i in best]
        return sum(r['bel'] for r in sel), sum(r['ra'] for r in sel), sum(r['csm'] for r in sel)

    # 손보사 fallback: target includes PAA, so exact match impossible.
    # Use structured candidates (even-idx or first-half) that are ≤ target.
    best_struct = None
    best_struct_sum = 0
    for combo in candidates:
        g = sum(basis_rows[i]['grand'] for i in combo)
        if g <= target_sum and g > best_struct_sum:
            best_struct_sum = g
            best_struct = combo

    if best_struct and best_struct_sum > target_sum * 0.80:
        sel = [basis_rows[i] for i in best_struct]
        return sum(r['bel'] for r in sel), sum(r['ra'] for r in sel), sum(r['csm'] for r in sel)

    return None, None, None


def parse_csm_movement_from_doc(corp_name, csm_end_target, year="2025"):
    """
    사업보고서 document.xml에서 CSM 변동표 파싱.
    csm_end_target: Col75 CSM 기말잔액 (백만원, 당기 half 식별용)
    반환: {76: 신계약CSM, 77: CSM상각, 78: CSM조정, 79: 보험금융, 80: 전기CSM} (백만원)
    """
    corp_code = DOC_CSM_MOVEMENT_COMPANIES.get(corp_name)
    if not corp_code:
        return {}

    cache_path = os.path.join("data", corp_name, f"doc_annual_{year}.xml")
    os.makedirs(os.path.dirname(cache_path), exist_ok=True)

    if not os.path.exists(cache_path):
        try:
            from dart_api import get_disclosures, _fetch_document_content
            # FY2025 사업보고서는 2026년에 제출됨
            next_year = str(int(year) + 1)
            data = get_disclosures(corp_code=corp_code, bgn_de=f"{next_year}0101",
                                   end_de=f"{next_year}0630", page_count=50)
            items = data.get("list", [])
            candidates = [i for i in items if "사업보고서" in i.get("report_nm", "")
                          and year in i.get("report_nm", "")]
            if not candidates:
                return {}
            # 원본 우선 (첨부정정/기재정정은 document.xml 없을 수 있음)
            originals = [c for c in candidates if "정정" not in c.get("report_nm", "")]
            selected = originals[0] if originals else candidates[-1]
            content = _fetch_document_content(selected["rcept_no"])
            os.makedirs("data", exist_ok=True)
            with open(cache_path, "w", encoding="utf-8") as f:
                f.write(content)
            print(f"    사업보고서 doc.xml 다운로드: {cache_path}")
        except Exception as e:
            print(f"    사업보고서 doc.xml 다운로드 실패 ({corp_name}): {e}")
            return {}

    return _extract_csm_movement(cache_path, csm_end_target)


def _parse_num(s):
    """백만원 단위 숫자 파싱 (괄호=음수, 하이픈=0)"""
    s = s.replace(',', '').replace('(', '-').replace(')', '').strip()
    if s in ('-', '', '- '):
        return 0
    try:
        return int(s)
    except ValueError:
        return 0


def _extract_csm_movement(filepath, csm_end_target):
    """사업보고서 document.xml에서 CSM 변동 행 추출 (범용)."""
    try:
        with open(filepath, encoding="utf-8") as f:
            content = f.read()
    except Exception:
        return {}

    table_pattern = re.compile(r'<TABLE[^>]*>(.*?)</TABLE>', re.DOTALL | re.IGNORECASE)
    all_tables = []

    for tm in table_pattern.finditer(content):
        tbl = tm.group(1)
        text = re.sub(r'<[^>]+>', '|', tbl)
        text = re.sub(r'\|[\s|]+\|', '|', text)
        text = re.sub(r'\s+', ' ', text)
        if '재보험' in text[:200] or '출재' in text[:200]:
            continue
        header = text[:600]
        if '보험계약마진' not in header:
            continue
        if '잔여보장' in header and '발생사고' in header:
            continue
        has_movement = ('기초' in text and
                        ('신계약' in text or '최초' in text) and
                        ('상각' in text or '서비스' in text))
        if not has_movement:
            continue
        all_tables.append((tm.start(), text))

    if not all_tables:
        return {}

    sections = [[all_tables[0]]]
    for i in range(1, len(all_tables)):
        if all_tables[i][0] - all_tables[i - 1][0] > 500000:
            sections.append([all_tables[i]])
        else:
            sections[-1].append(all_tables[i])

    target_section = sections[-1]

    def _find_end_row(lines):
        """Find 기말 row index - handles split labels like '기말|보험계약부채'."""
        for i, cell in enumerate(lines):
            cs = cell.strip()
            if '기말' in cs and ('순장부' in cs or '보험계약' in cs):
                return i
            if cs == '기말' and i + 1 < len(lines):
                nxt = lines[i + 1].strip()
                if '보험계약' in nxt or '순장부' in nxt:
                    return i + 1
        return -1

    n = len(target_section)

    def _table_csm_end_all(ti):
        """Get all possible CSM-end values for a table."""
        text = target_section[ti][1]
        lines = text.split('|')
        end_idx = _find_end_row(lines)
        if end_idx < 0:
            return {}
        row = [c.strip() for c in lines[end_idx + 1:end_idx + 10]]
        vals = [_parse_num(v) for v in row]
        results = {}
        for ci in range(min(7, len(vals))):
            if vals[ci] != 0:
                results[ci] = vals[ci]
        if len(vals) >= 5:
            s = sum(vals[2:5])
            if s != 0:
                results['s2_5'] = s
        if len(vals) >= 6:
            s = sum(vals[2:6])
            if s != 0:
                results['s2_6'] = s
        if len(vals) >= 7:
            results['col5'] = vals[5] if len(vals) > 5 else 0
        return results

    table_ends = [_table_csm_end_all(ti) for ti in range(n)]

    best = None
    best_err = float('inf')
    best_col_key = None

    col_keys_to_try = set()
    for te in table_ends:
        col_keys_to_try.update(te.keys())

    half = n // 2 if n >= 4 else n
    for col_key in col_keys_to_try:
        for ti in range(n):
            v = table_ends[ti].get(col_key, 0)
            err = abs(v - csm_end_target)
            if err < best_err:
                best_err = err
                best = (ti,)
                best_col_key = col_key
        for size in range(2, min(half + 1, n + 1)):
            for combo in combinations(range(n), size):
                s = sum(table_ends[ti].get(col_key, 0) for ti in combo)
                err = abs(s - csm_end_target)
                if err < best_err:
                    best_err = err
                    best = combo
                    best_col_key = col_key
            if best_err < 10:
                break
        if best_err < 10:
            break

    if not best or best_err > csm_end_target * 0.02:
        return {}

    if isinstance(best_col_key, int):
        final_col_idx = best_col_key
        sum_mode = False
    elif best_col_key == 's2_5':
        final_col_idx = 2
        sum_mode = True
    elif best_col_key == 's2_6':
        final_col_idx = 2
        sum_mode = True
    else:
        final_col_idx = 5
        sum_mode = False

    def _extract_row_csm(row, col_idx, is_sum_mode, sum_end=5):
        """Extract CSM value from a row given column info."""
        vals = [_parse_num(v) for v in row]
        if is_sum_mode:
            return sum(vals[col_idx:min(sum_end, len(vals))])
        if col_idx < len(vals):
            return vals[col_idx]
        return 0

    sum_end = 5 if best_col_key == 's2_5' else 6

    result = {76: 0, 77: 0, 78: 0, 79: 0, 80: 0}

    for idx in best:
        text = target_section[idx][1]
        lines = text.split('|')
        begin_found = False

        for i, cell in enumerate(lines):
            cs = cell.strip()
            row = [c.strip() for c in lines[i + 1:i + 10]]
            if len(row) < 3:
                continue
            csm_val = _extract_row_csm(row, final_col_idx, sum_mode, sum_end)

            is_begin = ('기초' in cs and '기말' not in cs and
                        ('순장부' in cs or '순부채' in cs or '보험계약마진' in cs))
            if is_begin and not begin_found:
                begin_found = True
                result[80] += csm_val
            elif cs == '기초' and i + 1 < len(lines) and '보험계약' in lines[i + 1].strip():
                result[80] += _extract_row_csm(
                    [c.strip() for c in lines[i + 2:i + 12]], final_col_idx, sum_mode, sum_end)
            elif ('최초' in cs and ('인식' in cs or '계약' in cs)):
                result[76] += csm_val
            elif '신계약' in cs and ('효과' in cs or '인식' in cs):
                result[76] += csm_val
            elif '보험계약마진' in cs and '상각' in cs:
                result[77] += abs(csm_val)
            elif '서비스 이전' in cs and '보험계약마진' in cs:
                result[77] += abs(csm_val)
            elif '서비스 제공' in cs and '보험계약마진' in cs:
                result[77] += abs(csm_val)
            elif '당기손익으로 인식한 보험계약마진' in cs:
                result[77] += abs(csm_val)
            elif '제공된 서비스' in cs and '당기손익' in cs:
                result[77] += abs(csm_val)
            elif '보험계약마진을 조정' in cs and ('추정' in cs or '변동' in cs):
                result[78] += csm_val
            elif '보험계약마진' in cs and '조정' in cs and '추정' in cs:
                result[78] += csm_val
            elif '순보험금융손익' in cs:
                result[79] += csm_val
            elif cs.startswith('6') and '보험금융' in cs:
                result[79] += csm_val
            elif '보험금융손익' in cs and '소' not in cs and '기타' not in cs and '당기' not in cs and '결과' not in cs:
                fin_row = row
                if fin_row and any(c.isalpha() for c in fin_row[0]):
                    fin_row = fin_row[1:]
                fin_val = _extract_row_csm(fin_row, final_col_idx, sum_mode, sum_end)
                if fin_val != 0:
                    result[79] += fin_val

    return result


def _get_annual_doc(corp_name, year="2025"):
    """사업보고서 document.xml 경로 반환 (없으면 다운로드)."""
    cache_path = os.path.join("data", corp_name, f"doc_annual_{year}.xml")
    os.makedirs(os.path.dirname(cache_path), exist_ok=True)
    if os.path.exists(cache_path):
        return cache_path
    corp_code = ALL_CORP_CODES.get(corp_name)
    if not corp_code:
        return None
    try:
        from dart_api import get_disclosures, _fetch_document_content
        next_year = str(int(year) + 1)
        data = get_disclosures(corp_code=corp_code, bgn_de=f"{next_year}0101",
                               end_de=f"{next_year}0630", page_count=50)
        items = data.get("list", [])
        candidates = [i for i in items if "사업보고서" in i.get("report_nm", "")]
        if not candidates:
            return None
        originals = [c for c in candidates if "정정" not in c.get("report_nm", "")]
        selected = originals[0] if originals else candidates[-1]
        content = _fetch_document_content(selected["rcept_no"])
        os.makedirs("data", exist_ok=True)
        with open(cache_path, "w", encoding="utf-8") as f:
            f.write(content)
        print(f"    사업보고서 doc.xml 다운로드: {cache_path}")
        return cache_path
    except Exception as e:
        print(f"    doc.xml 다운로드 실패 ({corp_name}): {e}")
        return None


def _extract_kics_from_q1(filepath, target_year="2025"):
    """1분기보고서 doc.xml에서 전년도(target_year) K-ICS 최종 확정치 추출.
    TR 기반: 헤더 TR에서 target_year 포함 TD의 열 인덱스를 찾고,
    데이터 TR에서 같은 열 인덱스 값을 가져옴.
    반환: {95: 비율, 96: 가용자본(백만원), 97: 요구자본(백만원)}
    """
    try:
        with open(filepath, encoding="utf-8") as f:
            content = f.read()
    except Exception:
        return {}

    def _pf(s):
        try:
            return float(s.replace(',', '').replace('(', '-').replace(')', '').strip())
        except (ValueError, AttributeError):
            return None

    AVAIL_KWS = ('지급여력(A)', '지급여력금액(A)', '가용자본(A)', '가용자본')
    REQ_KWS   = ('지급여력기준(B)', '지급여력기준금액(B)', '요구자본', '지급여력(B)', '지급여력기준')
    RATIO_KWS = ('지급여력비율(A/B)', '지급여력비율', 'A/B')

    table_pattern = re.compile(r'<TABLE[^>]*>(.*?)</TABLE>', re.DOTALL | re.IGNORECASE)
    for m in table_pattern.finditer(content):
        t = m.group(1)
        plain = re.sub(r'<[^>]+>', '', t)
        if not ('지급여력' in plain[:500] or '가용자본' in plain[:500]):
            continue
        if '개선권고' in plain[:200] or '경영개선' in plain[:200]:
            continue
        # 숫자 데이터 있어야 함 (쉼표 포함 금액 또는 5자리 이상 연속 숫자)
        if not re.search(r'\d[\d,]{4,}', plain):
            continue

        # 단위 감지
        before_text = re.sub(r'<[^>]+>', ' ', content[max(0, m.start()-500):m.start()])
        unit_m = re.search(r'단위\s*[:\s]\s*(억원|백만원|원)', before_text + plain[:300])
        unit_factor = 1
        if unit_m:
            u = unit_m.group(1)
            if '억원' in u:
                unit_factor = 100
            elif '원' in u and '백만' not in u:
                unit_factor = 1e-6

        # TR 기반 파싱
        rows = re.split(r'<TR[^>]*>', t, flags=re.IGNORECASE)

        def tr_cells(row):
            tds = re.findall(r'<(?:TD|TH)[^>]*>(.*?)</(?:TD|TH)>', row, re.DOTALL | re.IGNORECASE)
            return [re.sub(r'<[^>]+>', '', td).strip() for td in tds]

        # 헤더 TR에서 target_year 포함 TD의 열 인덱스(0-based) 찾기
        # COLSPAN 고려: 헤더 TD의 COLSPAN만큼 데이터 열 인덱스를 누적
        year_col = None
        for row in rows[:6]:
            tds_raw = re.findall(r'<(?:TD|TH)([^>]*)>(.*?)</(?:TD|TH)>', row, re.DOTALL | re.IGNORECASE)
            if not tds_raw:
                continue
            col_offset = 0  # 현재까지 누적된 실제 열 인덱스
            for attrs, inner in tds_raw:
                label = re.sub(r'<[^>]+>', '', inner).strip()
                cs_m = re.search(r'COLSPAN=["\'](\d+)', attrs, re.IGNORECASE)
                colspan = int(cs_m.group(1)) if cs_m else 1
                if col_offset > 0 and target_year in label:
                    year_col = col_offset  # 이 TD가 시작하는 실제 열 인덱스
                    break
                col_offset += colspan
            if year_col is not None:
                break

        # target_year를 못 찾으면 → 기수(期數) 방식: 두 번째 데이터열(index=2) 사용
        if year_col is None:
            for row in rows[:3]:
                tds_raw = re.findall(r'<(?:TD|TH)([^>]*)>(.*?)</(?:TD|TH)>', row, re.DOTALL | re.IGNORECASE)
                labels = [re.sub(r'<[^>]+>', '', inner).strip() for _, inner in tds_raw]
                # 헤더 행: 구분열 + 연도/기수 레이블 여러 개
                is_header = any(re.search(r'\d{2,}기|분기|FY\d{4}|\d{4}년', c) for c in labels[1:])
                if is_header and len(labels) >= 3:
                    year_col = 2  # 0-based: 구분=0, 당기=1, 전기(target_year)=2
                    break

        if year_col is None:
            continue

        # 데이터 TR에서 year_col 위치 값 수집
        avail = req = ratio = None
        for row in rows:
            cells = tr_cells(row)
            if not cells or year_col >= len(cells):
                continue
            label = cells[0].strip()
            val_str = cells[year_col].strip()
            if val_str in ('-', '산출중', '미정', ''):
                continue
            val = _pf(val_str)
            if val is None:
                continue

            if any(kw in label for kw in AVAIL_KWS) and '요구' not in label and '기준' not in label:
                if avail is None:
                    avail = val
            elif any(kw in label for kw in REQ_KWS):
                if req is None:
                    req = val
            elif any(kw in label for kw in RATIO_KWS):
                if ratio is None:
                    ratio = val

        if avail is None or req is None or ratio is None:
            continue

        # 단위 적용
        def _apply_unit(v, uf):
            if uf != 1:
                return v * uf
            return v * 100 if v < 100000 else v

        avail2 = _apply_unit(avail, unit_factor)
        req2   = _apply_unit(req,   unit_factor)
        ratio2 = ratio / 100 if ratio > 10 else ratio

        # 일관성 검증: avail/req ≈ ratio (5% 오차)
        if req2 > 0 and abs(avail2 / req2 - ratio2) / max(ratio2, 1e-9) > 0.05:
            continue

        return {95: ratio2, 96: avail2, 97: req2}

    return {}


def _extract_kics(filepath):
    """사업보고서 doc.xml에서 K-ICS 지급여력비율/가용자본/요구자본 추출.
    반환: {92: 비율(소수), 93: 가용자본(백만원), 94: 요구자본(백만원)}
    """
    try:
        with open(filepath, encoding="utf-8") as f:
            content = f.read()
    except Exception:
        return {}

    def _parse_float(s):
        try:
            return float(s.replace(',', '').replace('(', '-').replace(')', '').strip())
        except (ValueError, AttributeError):
            return 0.0

    def _fix_unit(lst):
        return [v * 100 if v < 100000 else v for v in lst]

    def _build_result(ratios, avails, reqs):
        if not ratios or not avails or not reqs:
            return {}
        avails = _fix_unit(avails)
        reqs = _fix_unit(reqs)
        r = {}
        if ratios: r[95] = ratios[0]; r[92] = ratios[0]
        if avails: r[96] = avails[0]; r[93] = avails[0]
        if reqs:   r[97] = reqs[0];  r[94] = reqs[0]
        return r

    def _parse_table_cells(text):
        """TABLE 텍스트에서 K-ICS 수치 추출. (기존 로직)"""
        cells = text.split('|')
        ratios, avails, reqs = [], [], []
        for i, cell in enumerate(cells):
            cs = cell.strip()
            if '지급여력비율' in cs or ('비율' in cs and 'A/B' in cs):
                row = [c.strip() for c in cells[i+1:i+10]]
                if row and row[0] in ('-', '산출중', '미정'):
                    break
                for v in row:
                    n = _parse_float(v)
                    if 50 < n < 1000:
                        ratios.append(n / 100)
            elif '지급여력금액' in cs or '지급여력(A)' in cs or ('가용자본' in cs and '요구' not in cs and '기준' not in cs):
                row = [c.strip() for c in cells[i+1:i+8]]
                for v in row:
                    n = _parse_num(v)
                    if n > 1000:
                        avails.append(n)
            elif '지급여력기준' in cs or '지급여력기준금액' in cs or '요구자본' in cs or '지급여력(B)' in cs:
                row = [c.strip() for c in cells[i+1:i+8]]
                for v in row:
                    n = _parse_num(v)
                    if n > 1000:
                        reqs.append(n)
        return ratios, avails, reqs

    table_pattern = re.compile(r'<TABLE[^>]*>(.*?)</TABLE>', re.DOTALL | re.IGNORECASE)

    # ── 패턴1: TABLE 방식 (기존 로직) ─────────────────────────────────────
    idx = content.find('지급여력비율')
    if idx < 0: idx = content.find('지급여력금액')
    if idx < 0: idx = content.find('지급여력(A)')
    if idx < 0: idx = content.find('가용자본')
    search_start = max(0, idx - 2000) if idx >= 0 else 0
    search_end = min(len(content), idx + 50000) if idx >= 0 else len(content)

    for m in table_pattern.finditer(content, search_start, search_end):
        text = re.sub(r'<[^>]+>', '|', m.group(1))
        text = re.sub(r'\|[\s|]+\|', '|', text)
        text = re.sub(r'\s+', ' ', text).strip()
        if not ('지급여력' in text[:500] or '가용자본' in text[:500]):
            continue
        if '개선권고' in text[:200] or '경영개선' in text[:200]:
            continue
        ratios, avails, reqs = _parse_table_cells(text)
        r = _build_result(ratios, avails, reqs)
        if r:
            return r

    # ── 패턴2: TD 방식 (한화생명 등 TABLE이 없는 경우) ──────────────────────
    # 지급여력비율/가용자본 키워드가 TD 안에 있는 경우
    for kw in ['지급여력(A)', '가용자본(A)', '지급여력금액(A)', '지급여력금액']:
        idx2 = content.find(kw)
        if idx2 < 0:
            continue
        # kw 위치에서 앞 200자 내 TR을 포함하는 구간 전체를 가상 TABLE로 처리
        region = content[max(0, idx2 - 200):idx2 + 3000]
        text = re.sub(r'<[^>]+>', '|', region)
        text = re.sub(r'\|[\s|]+\|', '|', text)
        text = re.sub(r'\s+', ' ', text).strip()
        ratios, avails, reqs = _parse_table_cells(text)
        r = _build_result(ratios, avails, reqs)
        if r:
            return r

    return {}


def _extract_csm_maturity(filepath, csm_total):
    """사업보고서 doc.xml에서 CSM 기간별 기대수익인식금액 추출.
    csm_total: Col88(합계) 기준으로 당기 테이블 식별.
    반환: {83: 1년이하, 84: 1~3년, 85: 3~5년, 86: 5~10년, 87: 10년초과, 88: 합계}
    """
    try:
        with open(filepath, encoding="utf-8") as f:
            content = f.read()
    except Exception:
        return {}

    table_pattern = re.compile(r'<TABLE[^>]*>(.*?)</TABLE>', re.DOTALL | re.IGNORECASE)
    # 테이블 앞 2000자 문맥에서 CSM 기간별 설명이 있는지 확인하는 키워드
    CSM_MATURITY_CONTEXT_KWS = ('보험계약마진이 미래에', '예상 당기손익 인식기간', '기대수익인식금액',
                                  '보험계약마진의 예상', '기간별 기대수익', '기간별 예상수익',
                                  '기간별 예상상각', '보험계약마진의 향후')
    candidates = []
    for m in table_pattern.finditer(content):
        text = re.sub(r'<[^>]+>', '|', m.group(1))
        text = re.sub(r'\|[\s|]+\|', '|', text)
        text = re.sub(r'\s+', ' ', text).strip()
        # 테이블 앞 문맥 확인
        before = re.sub(r'<[^>]+>', ' ', content[max(0, m.start()-2000):m.start()])
        before = re.sub(r'\s+', ' ', before)
        context_has_csm = any(kw in before for kw in CSM_MATURITY_CONTEXT_KWS)
        has_csm = '보험계약마진' in text[:800] or '기대수익' in text[:800] or context_has_csm
        # has_time: 버킷 구조 키워드
        has_time_strict = (any(k in text[:600] for k in ('1년이하', '1년 이하', '10년 초과', '5~10')) or
                           bool(re.search(r'\d+년\s*이하|\d+년\s*초과|\d+년\s*이내', text[:600])))
        # context가 있으면 단순 '1년', '2년' 패턴도 허용
        has_time_loose = context_has_csm and bool(re.search(r'\|\s*\d+년\s*\|', text[:600]))
        has_time = has_time_strict or has_time_loose
        if has_csm and has_time:
            candidates.append((m.start(), text))

    if not candidates:
        return {}

    # 연결/별도 section split
    sections = [[candidates[0]]]
    for i in range(1, len(candidates)):
        if candidates[i][0] - candidates[i-1][0] > 500000:
            sections.append([candidates[i]])
        else:
            sections[-1].append(candidates[i])
    # Choose section whose total (or single-table total) is closest to csm_total
    def _quick_total(text):
        """테이블의 총합 추정: 모든 양수 값의 최대값 (전체 합계가 가장 큰 값)."""
        all_nums = [abs(_parse_num(v)) for v in text.split('|') if _parse_num(v) > 0]
        return max(all_nums) if all_nums else 0

    # 연결/별도 section split
    sections = [[candidates[0]]]
    for i in range(1, len(candidates)):
        if candidates[i][0] - candidates[i-1][0] > 500000:
            sections.append([candidates[i]])
        else:
            sections[-1].append(candidates[i])

    # First try: find a single table whose total matches csm_total
    best_section = sections[-1]
    best_section_err = float('inf')
    for sec in sections:
        for pos, text in sec:
            t = _quick_total(text)
            err = abs(t - csm_total)
            if err < best_section_err and err < csm_total * 0.02:
                best_section_err = err
                best_section = [(pos, text)]  # single-table section
    # If no single table matched, try section sums
    if best_section_err > csm_total * 0.02:
        for sec in sections:
            sec_total = sum(_quick_total(text) for _, text in sec)
            err = abs(sec_total - csm_total)
            if err < best_section_err:
                best_section_err = err
                best_section = sec
    target_section = best_section

    def _get_total(text):
        cells = text.split('|')
        for i, cell in enumerate(cells):
            cs = cell.strip()
            is_total_row = (cs in ('원수보험계약', '합계', '합 계', '소 계', '발행한 보험계약') or
                            ('보험계약마진' in cs and '출재' not in cs and '재보험' not in cs))
            if not is_total_row:
                continue
            # Collect nums until we hit a non-numeric cell
            row_nums = []
            for v in cells[i+1:i+25]:
                vs = v.strip()
                n = _parse_num(vs)
                if n != 0:
                    row_nums.append(n)
                elif vs and any(c.isalpha() for c in vs):
                    break  # hit a label cell, stop
            if row_nums:
                return abs(row_nums[-1])  # abs for negative tables
        return 0

    best, best_err = None, float('inf')
    for ti, (pos, text) in enumerate(target_section):
        t = _get_total(text)
        err = abs(t - csm_total)
        if err < best_err:
            best_err = err
            best = ti

    if best is None or best_err > csm_total * 0.02:
        return {}

    text = target_section[best][1]
    cells = text.split('|')

    # Detect header bucket structure from table text
    header = text[:600]

    # Collect header bucket labels (year ranges)
    year_buckets = re.findall(r'(\d+)년\s*(?:초과\s*)?(\d+)년\s*이내', header)
    fine_buckets = len(year_buckets) >= 8  # 동양/DB style: many fine buckets

    # 헤더에 연도 범위 레이블이 있을 때만 패턴A 적용 (없으면 패턴B로)
    has_bucket_header = bool(re.search(r'\d+년\s*이하|\d+년\s*초과|\d+년\s*이내|\d+년\s*[~～]\s*\d+년', header))

    # Find data row (원수보험계약 or 보험계약마진 합산 row)
    if has_bucket_header:
        for i, cell in enumerate(cells):
            cs = cell.strip()
            is_data_row = (cs in ('원수보험계약', '합 계', '합계', '보험계약마진', '발행한 보험계약') or
                           ('보험계약마진' in cs and '출재' not in cs and '상각' not in cs and '재보험' not in cs))
            if not is_data_row:
                continue

            row = []
            for v in cells[i+1:i+25]:
                vs = v.strip()
                if any(c.isalpha() for c in vs) and '.' not in vs:
                    if row:
                        break
                    continue
                n = _parse_num(vs)
                row.append(n)
                if len(row) >= 20:
                    break

            real_nums = [n for n in row if n > 0]
            if len(real_nums) < 4:
                continue

            # 합계는 마지막 값
            total = real_nums[-1]
            buckets = real_nums[:-1]

            if not buckets:
                continue

            nb = len(buckets)
            if nb >= 10:
                r = {83: buckets[0],
                     84: sum(buckets[1:3]),
                     85: sum(buckets[3:5]),
                     86: sum(buckets[5:10]),
                     87: sum(buckets[10:]),
                     88: total}
            elif nb == 9:
                r = {83: buckets[0], 84: sum(buckets[1:3]), 85: sum(buckets[3:5]),
                     86: buckets[5], 87: sum(buckets[6:]), 88: total}
            elif nb == 8:
                r = {83: buckets[0], 84: sum(buckets[1:3]), 85: sum(buckets[3:5]),
                     86: buckets[5], 87: sum(buckets[6:]), 88: total}
            elif nb == 7:
                r = {83: buckets[0], 84: buckets[1]+buckets[2],
                     85: buckets[3]+buckets[4], 86: buckets[5], 87: buckets[6], 88: total}
            elif nb == 6:
                r = {83: buckets[0], 84: buckets[1], 85: buckets[2],
                     86: buckets[3], 87: buckets[4], 88: total}
            elif nb == 5:
                r = {83: buckets[0], 84: buckets[1], 85: buckets[2],
                     86: buckets[3], 87: 0, 88: total}
            elif nb == 4:
                r = {83: buckets[0], 84: buckets[1], 85: buckets[2],
                     86: buckets[3], 87: 0, 88: total}
            else:
                continue

            # Verify consistency
            s = r[83] + r[84] + r[85] + r[86] + r.get(87, 0)
            if abs(s - total) < total * 0.02:
                return r

    # ── 패턴B: 연도×포트폴리오 행렬 구조 (KB생명 등)
    # 패턴A에서 실패한 경우에만 시도
    # 행: 1년, 2년, ..., n년, 합계 / 열: 포트폴리오1...포트폴리오N, 합계
    # 마지막 열 = 발행한 보험계약 합계
    all_candidates = [(pos, text) for sec in sections for pos, text in sec] if candidates else []
    for pos, text in all_candidates:
        cells = text.split('|')
        # 연도 행 수집: '1년', '2년', '11년~15년' 등 레이블로 시작하는 행
        year_rows = {}   # 단일 연도
        band_rows = {}   # 범위 구간 (key: (start, end))
        for i, c in enumerate(cells):
            cs = c.strip()
            # 단일 연도: '1년', '10년' 등
            m_yr = re.match(r'^(\d+)년$', cs)
            # 범위 구간: '11년~15년', '30년 이후' 등
            m_band = re.match(r'^(\d+)년\s*[~～]\s*(\d+)년', cs)
            m_after = re.match(r'^(\d+)년\s*이후', cs)
            if not (m_yr or m_band or m_after):
                continue
            # 이 행의 숫자들 수집 (마지막 열이 발행한 보험계약 합계)
            row_nums = []
            for v in cells[i+1:i+30]:
                vs = v.strip()
                if re.match(r'^[\d,]+$', vs):
                    row_nums.append(int(vs.replace(',', '')))
                elif vs and any(ch.isalpha() for ch in vs) and row_nums:
                    break
            if not row_nums:
                continue
            val = row_nums[-1]
            if m_yr:
                year_rows[int(m_yr.group(1))] = val
            elif m_band:
                band_rows[(int(m_band.group(1)), int(m_band.group(2)))] = val
            elif m_after:
                band_rows[(int(m_after.group(1)), 999)] = val

        if len(year_rows) + len(band_rows) < 5:
            continue

        # 단위 감지 (테이블 앞 문맥 포함)
        unit_factor = 1
        before_text = re.sub(r'<[^>]+>', ' ', content[max(0, pos-500):pos])
        combined = text + before_text
        if '단위 : 원' in combined or '단위:원' in combined or '(단위: 원)' in combined:
            unit_factor = 1e-6
        elif '단위 : 억원' in combined or '(단위: 억원)' in combined:
            unit_factor = 100

        # Col83~87 매핑
        # year_rows로 1~10년 처리, band_rows로 범위 구간 처리
        def _get_band_sum(start, end):
            """start <= yr <= end인 year_rows 합산 + 해당 범위를 커버하는 band_rows 합산"""
            s = sum(v for yr, v in year_rows.items() if start <= yr <= end)
            s += sum(v for (a, b), v in band_rows.items() if a >= start and b <= end)
            return s * unit_factor

        # 11년 이상은 band_rows에서 합산
        over10 = sum(band_rows.values()) * unit_factor

        r = {
            83: year_rows.get(1, 0) * unit_factor,
            84: (year_rows.get(2, 0) + year_rows.get(3, 0)) * unit_factor,
            85: (year_rows.get(4, 0) + year_rows.get(5, 0)) * unit_factor,
            86: sum(year_rows.get(y, 0) for y in range(6, 11)) * unit_factor,
            87: over10,
        }
        total_est = sum(r.values())
        if csm_total > 0 and abs(total_est - csm_total) < csm_total * 0.05:
            r[88] = total_est
            return r

    return {}


def _extract_sensitivity(filepath):
    """사업보고서 doc.xml에서 감응도/가정변경 추출.
    BEL+RA 기준: Col146-150, CSM 기준: Col153-158
    """
    try:
        with open(filepath, encoding="utf-8") as f:
            content = f.read()
    except Exception:
        return {}

    table_pattern = re.compile(r'<TABLE[^>]*>(.*?)</TABLE>', re.DOTALL | re.IGNORECASE)
    candidates = []
    for m in table_pattern.finditer(content):
        text = re.sub(r'<[^>]+>', '|', m.group(1))
        text = re.sub(r'\|[\s|]+\|', '|', text)
        text = re.sub(r'\s+', ' ', text).strip()
        has_s = ('가정변경' in text[:500] or '감응도' in text[:500] or '민감도' in text[:500])
        has_f = '해지율' in text and '위험률' in text
        if has_s and has_f and len(text.split('|')) >= 20:
            candidates.append((m.start(), text))

    if not candidates:
        return {}

    # 섹션 분리 (연결/별도)
    sections = [[candidates[0]]]
    for i in range(1, len(candidates)):
        if candidates[i][0] - candidates[i-1][0] > 500000:
            sections.append([candidates[i]])
        else:
            sections[-1].append(candidates[i])

    target_tables = sections[-1]

    def _parse_row(cells, i, unit_mult):
        """Extract BEL+RA value and CSM value from row at index i."""
        row = []
        for v in cells[i+1:i+15]:
            vs = v.strip()
            if any(c.isalpha() for c in vs) and not re.match(r'^[\d,.()\-\s]+$', vs):
                if row:
                    break
                continue
            n = _parse_num(vs)
            row.append(n)
            if len(row) >= 8:
                break

        nonzero_vals = [n for n in row if n != 0]
        if not nonzero_vals:
            return None, None

        # BEL+RA = sum of all except last (which is typically total or CSM)
        # Pattern: BEL | RA | CSM | (total) - or - BEL | CSM
        if len(nonzero_vals) >= 3:
            # Last = CSM or total, second-to-last = CSM if 4 cols
            csm_val = nonzero_vals[-1] * unit_mult
            bel_ra_val = sum(nonzero_vals[:-1]) * unit_mult
        elif len(nonzero_vals) == 2:
            bel_ra_val = nonzero_vals[0] * unit_mult
            csm_val = nonzero_vals[1] * unit_mult
        else:
            bel_ra_val = nonzero_vals[0] * unit_mult
            csm_val = nonzero_vals[0] * unit_mult

        return bel_ra_val, csm_val

    def _detect_unit(text):
        """Detect if table uses 억원 units."""
        if '억원' in text[:300]:
            return 100
        return 1

    results = {}
    FACTOR_MAP = {
        '해지율': (146, 153),
        '위험률': (147, 154),
        '사업비': (148, 155),
        '기타': (149, 156),
        '물량': (150, 157),
        '예실차': (150, 157),
        '손실요소': (158, None),
    }

    for pos, text in target_tables:
        unit_mult = _detect_unit(text)
        cells = text.split('|')

        for i, cell in enumerate(cells):
            cs = cell.strip()
            for keyword, (bel_col, csm_col) in FACTOR_MAP.items():
                if keyword not in cs:
                    continue
                # Skip header-like labels (too short or pure label)
                bel_val, csm_val = _parse_row(cells, i, unit_mult)
                if bel_val is None:
                    continue

                # Assign BEL col if not set
                if bel_col not in results and bel_val != 0:
                    results[bel_col] = bel_val
                # Assign CSM col if not set
                if csm_col and csm_col not in results and csm_val != 0:
                    results[csm_col] = csm_val
                break

    return results


def _extract_loss_ratios(filepath):
    """사업보고서 doc.xml에서 보험금 예실차 비율 추출.
    반환: {101: 예상손해율, 102: 실제손해율, 103: 예실차비율, 104: 실제/예상}
    """
    try:
        with open(filepath, encoding="utf-8") as f:
            content = f.read()
    except Exception:
        return {}

    def _pf(s):
        s = str(s).replace(",", "").replace("%", "").strip()
        try:
            return float(s)
        except ValueError:
            return None

    result = {}

    # Method 1: ACODE TE tag format (한화생명, 동양생명, DB손보)
    col_priority = {}  # col -> priority (0=Separate, 1=Consolidated)
    # 더 구체적인 태그를 먼저, 덜 구체적인 것을 나중에 (fallback 순서)
    for kw, col in [("ExpectedLossRateOfInsuranceClaim", 101), ("ActualLossRateOfInsuranceMoney", 102),
                    ("ActualLossRateOf", 102), ("ExpectedLossRateOf", 101),
                    ("ExpectedLossRatioOf", 101), ("ActualLossRatioOf", 102)]:
        for m in re.finditer(
            kw + r'[^"]*?"[^"]*?ACONTEXT="([^"]+)"[^>]*?>([0-9.]+)</TE>',
            content, re.DOTALL
        ):
            ctx, val = m.group(1), _pf(m.group(2))
            if val is None:
                continue
            if "SeparateMember" not in ctx and "ConsolidatedMember" not in ctx:
                continue
            priority = 0 if "SeparateMember" in ctx else 1
            if col not in result or priority < col_priority.get(col, 99):
                # 소수점 형식(0.86)이면 그대로, % 형식(86.0)이면 /100
                result[col] = val if val <= 2 else val / 100
                col_priority[col] = priority

    # Method 2: Table/text extraction
    if 101 not in result:
        table_pattern = re.compile(r'<TABLE[^>]*>(.*?)</TABLE>', re.DOTALL | re.IGNORECASE)
        for m2 in table_pattern.finditer(content):
            text = re.sub(r'<[^>]+>', '|', m2.group(1))
            text = re.sub(r'\|[\s|]+\|', '|', text)
            text = re.sub(r'\s+', ' ', text).strip()
            if '예상손해율' not in text[:400]:
                continue
            cells = text.split('|')

            exp_idx = next((i for i, c in enumerate(cells) if '예상손해율' in c.strip()), None)
            act_idx = next((i for i, c in enumerate(cells) if '실제손해율' in c.strip()), None)

            if exp_idx is None:
                continue

            def _normalize(n):
                """Normalize to 0~2 range (% → decimal)."""
                if n is None:
                    return None
                a = abs(n)
                if a > 1.5:
                    return n / 100  # percentage form
                return n  # already decimal

            def _get_label_val(start_idx):
                """Get first ratio value immediately after a label cell."""
                for v in cells[start_idx + 1: start_idx + 5]:
                    vs = v.strip()
                    n = _pf(vs)
                    nv = _normalize(n)
                    if nv is not None and 0.3 <= abs(nv) <= 1.5:
                        return nv
                return None

            # Try inline value (신한라이프 style: label immediately followed by value)
            exp_inline = _get_label_val(exp_idx)
            act_inline = _get_label_val(act_idx) if act_idx is not None else None

            # Check if inline values are distinct (not same as each other)
            if exp_inline is not None and act_inline is not None and exp_inline != act_inline:
                if 101 not in result:
                    result[101] = exp_inline
                if 102 not in result:
                    result[102] = act_inline
            else:
                # Fall back to after-all-headers method (삼성화재, 교보 style)
                header_end = max(x for x in [exp_idx, act_idx or 0] if x is not None)
                nums = []
                for v in cells[header_end + 1: header_end + 10]:
                    vs = v.strip()
                    if vs and not any(c.isalpha() for c in vs.replace('%', '')):
                        n = _pf(vs)
                        nv = _normalize(n)
                        if nv is not None and 0.3 <= abs(nv) <= 1.5:
                            nums.append(nv)
                    elif vs and any(c.isalpha() for c in vs) and nums:
                        break
                if len(nums) >= 2:
                    if 101 not in result:
                        result[101] = nums[0]
                    if 102 not in result:
                        result[102] = nums[1]
                elif exp_inline is not None and 101 not in result:
                    result[101] = exp_inline

            if result.get(101) or result.get(102):
                break

    if 101 in result and 102 in result:
        result[103] = result[101] - result[102]  # A - B = 예상 - 실제 (음수면 실제 > 예상)
        if result[101] != 0:
            result[104] = result[102] / result[101]

    return {k: v for k, v in result.items() if isinstance(v, (int, float)) and k in (101, 102, 103, 104)}


def _extract_investment_assets(filepath, total_assets=None):
    """사업보고서 doc.xml에서 운용자산 및 자산운용률 추출.
    total_assets: Col4 총자산 (백만원) - 단위 확인용
    반환: {22: 운용자산(백만원), 23: 자산운용률(소수)}
    """
    try:
        with open(filepath, encoding="utf-8") as f:
            content = f.read()
    except Exception:
        return {}

    table_pattern = re.compile(r'<TABLE[^>]*>(.*?)</TABLE>', re.DOTALL | re.IGNORECASE)
    for m in table_pattern.finditer(content):
        text = re.sub(r'<[^>]+>', '|', m.group(1))
        text = re.sub(r'\|[\s|]+\|', '|', text)
        text = re.sub(r'\s+', ' ', text).strip()
        if '운용자산' not in text[:400] or '자산운용률' not in text[:400]:
            continue

        cells = text.split('|')
        inv_assets = None
        util_rate = None
        total_assets_doc = None

        for i, cell in enumerate(cells):
            cs = cell.strip()
            if ('총자산' in cs or '총 자 산' in cs) and total_assets_doc is None:
                for v in cells[i + 1:i + 6]:
                    vs = v.strip()
                    try:
                        n = float(vs.replace(',', ''))
                        if n > 0:
                            total_assets_doc = n
                            break
                    except ValueError:
                        pass
            if '운용자산' in cs and '총' not in cs and '자산운용률' not in cs and inv_assets is None:
                for v in cells[i + 1:i + 6]:
                    vs = v.strip()
                    try:
                        n = float(vs.replace(',', ''))
                        if n > 0:
                            inv_assets = n
                            break
                    except ValueError:
                        pass
            elif '자산운용률' in cs and util_rate is None:
                for v in cells[i + 1:i + 6]:
                    vs = v.strip()
                    try:
                        n = float(vs.replace(',', '').rstrip('%'))
                        if n > 0:
                            util_rate = n / 100 if n > 1 else n
                            break
                    except ValueError:
                        pass

        if inv_assets is None:
            continue

        # Detect unit: if total_assets is known and inv_assets is ~100x smaller, it's in 억원
        if total_assets and total_assets > 0:
            ratio = inv_assets / total_assets
            if ratio < 0.05:  # 억원 단위 → × 100
                inv_assets *= 100

        result = {}
        if total_assets_doc:
            # Unit detection for total assets
            ta = total_assets_doc
            if total_assets and total_assets > 0 and ta / total_assets < 0.05:
                ta *= 100
            result[21] = ta
        if inv_assets:
            result[22] = inv_assets
        if util_rate:
            result[23] = util_rate
        if result:
            return result

    return {}


def _extract_oci_detail(filepath, oci_total=None):
    """사업보고서 doc.xml에서 기타포괄손익누계액 세부항목 추출 (별도 기준 기말잔액).
    oci_total: Col24(기타포괄손익누계액 합계) - 당기 테이블 식별용
    반환: {25: FVOCI, 26: 보험계약금융손익, 27: 재보험계약금융손익, 28: 현금흐름위험회피, 29: 재평가잉여금, 30: 확정급여}
    """
    try:
        with open(filepath, encoding="utf-8") as f:
            content = f.read()
    except Exception:
        return {}

    def _detect_oci_unit(raw_table_html, table_start_pos, fallback_range=3000):
        """테이블 단위 감지. 반환: 1=백만원, 1e-6=원(/1e6→백만원), 100=억원(×100→백만원)."""
        unit_m = re.search(r'단위\s*[:\s]\s*(억원|백만원|원)', raw_table_html)
        if unit_m:
            u = unit_m.group(1)
            if '억원' in u:
                return 100
            if '원' in u and '백만' not in u:
                return 1e-6
            return 1
        before = content[max(0, table_start_pos - fallback_range): table_start_pos]
        before_clean = re.sub(r'<[^>]+>', ' ', before)
        unit_m2 = re.search(r'단위\s*[:\s]\s*(억원|백만원|원)', before_clean)
        if unit_m2:
            u = unit_m2.group(1)
            if '억원' in u:
                return 100
            if '원' in u and '백만' not in u:
                return 1e-6
        return 1

    table_pattern = re.compile(r'<TABLE[^>]*>(.*?)</TABLE>', re.DOTALL | re.IGNORECASE)
    candidates = []
    for m in table_pattern.finditer(content):
        raw_html = m.group(0)
        text = re.sub(r'<[^>]+>', '|', m.group(1))
        text = re.sub(r'\|[\s|]+\|', '|', text)
        text = re.sub(r'\s+', ' ', text).strip()
        has_fvoci = any(kw in text[:800] for kw in
                        ['공정가치측정금융자산평가손익', 'FVOCI', '공정가치금융자산평가',
                         '공정가치측정유가증권평가손익', '공정가치측정금융상품평가손익'])
        has_ins = (('보험계약' in text[:800] or '보험금융손익' in text[:800]) and
                   ('금융손익' in text[:800] or '기타포괄' in text[:800] or '손익' in text[:800]))
        if has_fvoci and has_ins:
            unit_factor = _detect_oci_unit(raw_html, m.start())
            candidates.append((m.start(), text, unit_factor))

    if not candidates:
        return {}

    # 섹션 분리 (연결/별도): (pos, text, unit_factor) 튜플
    sections = [[candidates[0]]]
    for i in range(1, len(candidates)):
        if candidates[i][0] - candidates[i-1][0] > 500000:
            sections.append([candidates[i]])
        else:
            sections[-1].append(candidates[i])

    COL_KW = {
        25: [
            (['공정가치측정금융자산평가손익'], ['보험']),
            (['공정가치측정유가증권평가손익'], ['보험']),
            (['공정가치측정금융상품평가손익'], ['보험']),
        ],
        26: [
            (['보험계약', '금융손익'], ['재보험']),
            (['보험금융손익'], ['재보험']),
            (['순보험계약금융손익'], []),
            (['보험계약부채순금융손익'], []),
            (['보험계약자산', '순금융손익'], []),
            (['보험계약손익'], ['재보험']),
        ],
        27: [
            (['재보험계약', '금융손익'], []),
            (['재보험금융손익'], []),
            (['재보험계약자산', '순금융손익'], []),
            (['재보험계약손익'], []),
        ],
        28: [
            (['현금흐름위험회피'], []),
            (['위험회피목적파생상품평가손익'], []),
            (['위험회피목적'], ['재보험', '보험계약']),
        ],
        29: [
            (['재평가잉여금'], []),
            (['자산재평가손익'], []),
        ],
        30: [
            (['확정급여'], []),
        ],
    }

    def _label_matches(cs, patterns):
        """셀 레이블이 COL_KW 패턴 중 하나에 매칭되면 해당 col 번호 반환, 없으면 None."""
        for col, pats in patterns.items():
            for req_kws, exc_kws in pats:
                if all(kw in cs for kw in req_kws) and not any(kw in cs for kw in exc_kws):
                    return col
        return None

    def _parse_row_nums(cells, start, count=8):
        """start 이후 최대 count개 셀에서 숫자 파싱. None=비숫자/빈칸."""
        nums = []
        for v in cells[start:start + count]:
            vs = v.strip()
            if not vs:
                nums.append(None)
            elif any(c.isalpha() for c in vs):
                break  # 다음 레이블 시작 → 중단
            else:
                nums.append(_parse_num(vs))
        return nums

    def _extract_from_table(text, col_idx=None):
        """
        테이블에서 OCI 세부항목 추출.
        col_idx: 각 행에서 사용할 숫자 컬럼 인덱스 (None이면 첫 번째 숫자 사용).
        """
        cells = text.split('|')
        r = {}
        for i, cell in enumerate(cells):
            cs = cell.strip()
            col = _label_matches(cs, COL_KW)
            if col is None or col in r:
                continue
            nums = _parse_row_nums(cells, i + 1)
            valid = [n for n in nums if n is not None]
            if not valid:
                continue
            if col_idx is not None and col_idx < len(valid):
                r[col] = valid[col_idx]
            else:
                r[col] = valid[0]
        return r

    def _find_best_col_idx(text, oci_total_mm, unit_factor):
        """
        자본변동표 스타일 테이블에서 기말잔액 컬럼 인덱스 탐색.
        oci_total_mm: 백만원 단위 OCI 합계 (XBRL Col17 값)
        unit_factor: 테이블 단위 (1=백만원, 1e-6=원/1e6, 100=억원)
        합계행의 숫자들 중 oci_total_mm에 가장 가까운 값의 인덱스 반환.
        """
        cells = text.split('|')
        # oci_total을 테이블 단위로 변환
        # unit_factor=1e-6 → 테이블값이 원, oci_total은 백만원 → 테이블기준=oci_total*1e6
        if unit_factor == 1e-6:
            oci_target = oci_total_mm * 1e6
        elif unit_factor == 100:
            oci_target = oci_total_mm / 100
        else:
            oci_target = oci_total_mm
        # 합계/기말 행 찾기
        for i, cell in enumerate(cells):
            cs = cell.strip()
            if cs in ('합계', '소 계', '소계', '합 계', '기말') or cs.startswith('합'):
                nums = _parse_row_nums(cells, i + 1, count=10)
                valid = [(idx, n) for idx, n in enumerate(nums) if n is not None]
                if not valid:
                    continue
                # oci_target에 가장 가까운 컬럼 인덱스
                best_idx, best_err = 0, float('inf')
                for idx, n in valid:
                    err = abs(n - oci_target)
                    if err < best_err:
                        best_err = err
                        best_idx = idx
                # 오차가 oci_target의 2% 이내이면 신뢰
                if oci_target != 0 and best_err < abs(oci_target) * 0.02:
                    return best_idx
        return None

    def _get_section_best(sec, oci_total_hint):
        """섹션의 첫 번째 테이블에서 OCI 총합을 추출 (컬럼 자동 탐색 포함).
        반환: (sum_mm, col_idx, unit_factor)  sum_mm은 백만원 단위
        """
        if not sec:
            return None, None, 1
        pos, text, unit_factor = sec[0]
        # 컬럼 인덱스 탐색 (oci_total_hint가 있으면)
        col_idx = None
        if oci_total_hint is not None:
            col_idx = _find_best_col_idx(text, oci_total_hint, unit_factor)
        r = _extract_from_table(text, col_idx=col_idx)
        if len(r) >= 2:
            # 테이블 단위를 백만원으로 환산
            if unit_factor == 1e-6:
                raw_sum = sum(r.values()) * 1e-6
            elif unit_factor == 100:
                raw_sum = sum(r.values()) * 100
            else:
                raw_sum = sum(r.values())
            return raw_sum, col_idx, unit_factor
        return None, col_idx, unit_factor

    # 각 섹션 중 OCI합계에 가장 가까운 섹션 선택
    best_table_text = None
    best_col_idx = None
    best_unit_factor = 1
    best_err = float('inf')

    for sec in sections:
        pos, text, unit_factor = sec[0]
        if oci_total is not None:
            t, cidx, uf = _get_section_best(sec, oci_total)
            if t is not None:
                err = abs(t - oci_total)
                if err < best_err:
                    best_err = err
                    best_table_text = text
                    best_col_idx = cidx
                    best_unit_factor = uf
        else:
            pos, best_table_text, best_unit_factor = sections[-1][0]
            break

    if best_table_text is None:
        pos, best_table_text, best_unit_factor = sections[-1][0]

    result = _extract_from_table(best_table_text, col_idx=best_col_idx)
    # 단위 환산: 테이블이 원(1e-6) 또는 억원(100)이면 백만원으로 변환
    if best_unit_factor == 1e-6:
        result = {col: val * 1e-6 for col, val in result.items()}
    elif best_unit_factor == 100:
        result = {col: val * 100 for col, val in result.items()}
    return {col: val for col, val in result.items() if val != 0}


def parse_annual_doc_data(corp_name, csm_total=None, oci_total=None, year="2025", **kwargs):
    """사업보고서 doc.xml에서 K-ICS, CSM기간별, 감응도, OCI세부 통합 추출.
    반환: {col: value, ...}
    """
    doc_path = _get_annual_doc(corp_name, year)
    if not doc_path:
        return {}

    result = {}

    # Col92~94: 사업보고서 (잠정치), Col95~97: 1분기보고서 (최종치)
    kics_ann = _extract_kics(doc_path)
    next_year = str(int(year) + 1)
    q1_path = _get_q1_doc_path(corp_name, next_year)
    kics_q1 = _extract_kics_from_q1(q1_path, target_year=year) if q1_path else {}
    kics = dict(kics_ann)
    if kics_q1:
        # 1Q 최종치로 Col95~97 채움 (항상)
        kics[95] = kics_q1[95]
        kics[96] = kics_q1[96]
        kics[97] = kics_q1[97]
        if not kics_ann:
            # 사업보고서 값 없으면 잠정치도 1Q로
            kics[92] = kics_q1[95]
            kics[93] = kics_q1[96]
            kics[94] = kics_q1[97]
    result.update(kics)

    loss = _extract_loss_ratios(doc_path)
    result.update(loss)

    inv = _extract_investment_assets(doc_path, total_assets=kwargs.get('total_assets'))
    result.update(inv)

    oci = _extract_oci_detail(doc_path, oci_total=oci_total)
    result.update(oci)

    if csm_total and csm_total > 0:
        maturity = _extract_csm_maturity(doc_path, csm_total)
        result.update(maturity)

    # 복잡한 구조로 파싱 불가 또는 XBRL로 이미 추출되는 회사 → 스킵
    SKIP_SENSITIVITY = {"동양생명", "삼성화재", "삼성생명"}
    if corp_name not in SKIP_SENSITIVITY:
        sensitivity = _extract_sensitivity(doc_path)
        result.update(sensitivity)

    return result


def parse_new_csm(xbrl_path, target_year="2025", ref_date=None, start_date=None):
    """
    신계약 CSM 추출: 별도 + duration + InsuranceContractsIssuedMember + ContractualServiceMarginMember
    ChangesInFutureServicesDueToNewContracts... 태그
    """
    if ref_date is None:
        ref_date = f"{target_year}-12-31"
    if start_date is None:
        start_date = f"{target_year}-01-01"
    tree = ET.parse(xbrl_path)
    root = tree.getroot()

    for ctx in root.findall(f"{{{NS_XBRLI}}}context"):
        ctx_id = ctx.get("id", "")
        period = ctx.find(f"{{{NS_XBRLI}}}period")
        start = period.find(f"{{{NS_XBRLI}}}startDate")
        end = period.find(f"{{{NS_XBRLI}}}endDate")
        if start is None or end is None:
            continue
        if start.text != start_date or end.text != ref_date:
            continue

        members = ctx.findall(f".//{{{NS_XBRLDI}}}explicitMember")
        dims = {}
        for m in members:
            dim_name = m.get("dimension", "").split(":")[-1]
            val = (m.text or "").split(":")[-1]
            dims[dim_name] = val

        if "SeparateMember" != dims.get("ConsolidatedAndSeparateFinancialStatementsAxis", ""):
            continue
        if "ContractualServiceMarginMember" not in dims.get("InsuranceContractsByComponentsAxis", ""):
            continue
        if "InsuranceContractsIssuedMember" not in dims.get("DisaggregationOfInsuranceContractsAxis", ""):
            continue
        if len(dims) != 3:
            continue

        # 이 context의 신계약CSM 태그 찾기
        for elem in root:
            if elem.get("contextRef") != ctx_id or not elem.text:
                continue
            tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
            if "ChangesInFutureServicesDueToNewContracts" in tag:
                try:
                    return float(elem.text.strip())
                except ValueError:
                    pass
    return None


def parse_equity_changes(xbrl_path, target_year="2025", ref_date=None, start_date=None):
    """자기주식 소각(Col67), 기타자본변동(Col69) 추출.
    RetainedEarnings component context (2-dim: Separate + RetainedEarnings)에서 추출.
    반환: {67: treasury_cancellation, 69: other_equity_change}
    """
    if ref_date is None:
        ref_date = f"{target_year}-12-31"
    if start_date is None:
        start_date = f"{target_year}-01-01"
    tree = ET.parse(xbrl_path)
    root = tree.getroot()

    COL67_TAGS = {"CancellationOfTreasuryShares"}
    COL69_TAGS = {
        "OtherComprehensiveIncomeNetOfTaxFinancialAssetsMeasuredAtFairValueThroughOtherComprehensiveIncome",
        "OtherComprehensiveIncomeNetOfTaxFinancialAssetsMeasuredAtFairValueThroughOtherComprehensiveIncomeTotal",
        "TransferOfAmountRecognisedInOtherComprehensiveIncomeAndAccumulatedInEquityRelatingToDisposalOfInvestmentsInEquityInstruments",
        "ReclassificationOfGainOnDisposalOfEquityInstrumentsAtFairValueThroughOtherComprehensiveIncomeNetOfTax",
    }

    # Build RetainedEarnings component context set (2-dim, duration)
    re_ctx = set()
    for ctx in root.findall(f"{{{NS_XBRLI}}}context"):
        period = ctx.find(f"{{{NS_XBRLI}}}period")
        start = period.find(f"{{{NS_XBRLI}}}startDate") if period is not None else None
        end = period.find(f"{{{NS_XBRLI}}}endDate") if period is not None else None
        if start is None or start.text != start_date:
            continue
        if end is None or end.text != ref_date:
            continue
        members = ctx.findall(f".//{{{NS_XBRLDI}}}explicitMember")
        dims = {m.get("dimension", "").split(":")[-1]: (m.text or "").split(":")[-1] for m in members}
        if dims.get("ConsolidatedAndSeparateFinancialStatementsAxis") != "SeparateMember":
            continue
        comp = dims.get("ComponentsOfEquityAxis", "")
        if "RetainedEarning" in comp or "AccumulatedDeficit" in comp:
            re_ctx.add(ctx.get("id", ""))

    result = {}
    for elem in root:
        ctx_ref = elem.get("contextRef", "")
        if ctx_ref not in re_ctx or not elem.text:
            continue
        tag_local = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
        try:
            v = float(elem.text.strip())
        except ValueError:
            continue
        if v == 0:
            continue
        if tag_local in COL67_TAGS and 67 not in result:
            result[67] = v
        elif tag_local in COL69_TAGS and 69 not in result:
            result[69] = v

    return result


def parse_surrender_reserve(xbrl_path, target_year="2025", ref_date=None, start_date=None):
    """해약환급금준비금 추출.
    패턴A(생보): 2-dim (Separate + SurrenderValueReserveMember), tag=LoanLossReserveBalance 등
    패턴B(손보/삼성생명): 1-dim (Separate), tag=SurrenderValueReserve, instant 당기말
    """
    if ref_date is None:
        ref_date = f"{target_year}-12-31"
    if start_date is None:
        start_date = f"{target_year}-01-01"
    tree = ET.parse(xbrl_path)
    root = tree.getroot()

    best_a = None   # 패턴A (ReservesWithinEquityAxis)
    best_b = None   # 패턴B SurrenderValueReserve
    best_b_added = None  # SurrenderValueReserveToBeAdded (당기 추가분)

    for ctx in root.findall(f"{{{NS_XBRLI}}}context"):
        ctx_id = ctx.get("id", "")
        period = ctx.find(f"{{{NS_XBRLI}}}period")
        members = ctx.findall(f".//{{{NS_XBRLDI}}}explicitMember")
        dims = {}
        for m in members:
            dim_name = m.get("dimension", "").split(":")[-1]
            val = (m.text or "").split(":")[-1]
            dims[dim_name] = val

        if "SeparateMember" not in dims.get("ConsolidatedAndSeparateFinancialStatementsAxis", ""):
            continue

        instant = period.find(f"{{{NS_XBRLI}}}instant")
        start = period.find(f"{{{NS_XBRLI}}}startDate")
        end = period.find(f"{{{NS_XBRLI}}}endDate")
        is_cur = (
            (instant is not None and instant.text == ref_date) or
            (start is not None and start.text == start_date and
             end is not None and end.text == ref_date)
        )
        if not is_cur:
            continue

        has_reserve_axis = "SurrenderValueReserveMember" in dims.get("ReservesWithinEquityAxis", "")
        is_1dim_instant = (len(dims) == 1 and instant is not None)
        is_reported_amount = (
            dims.get("CarryingAmountAccumulatedDepreciationAmortisationAndImpairmentAndGrossCarryingAmountAxis", "") == "ReportedAmountMember"
            and len(dims) == 2 and instant is not None
        )

        if not has_reserve_axis and not is_1dim_instant and not is_reported_amount:
            continue

        for elem in root:
            if elem.get("contextRef") != ctx_id or not elem.text:
                continue
            tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
            try:
                v = float(elem.text.strip())
                if has_reserve_axis and "LoanLossReserveBalance" in tag:
                    if v > 0: best_a = v
                elif (is_1dim_instant or is_reported_amount) and tag == "SurrenderValueReserve":
                    if v > 0: best_b = v
                elif (is_1dim_instant or is_reported_amount) and tag == "SurrenderValueReserveToBeAdded":
                    best_b_added = v
                elif has_reserve_axis and ("SurrenderValueReserve" in tag or "LoanLossReserve" in tag):
                    if best_a is None:
                        best_a = v
            except ValueError:
                pass

    # Col99 = SurrenderValueReserve + SurrenderValueReserveToBeAdded (손보사는 두 값의 합이 정답)
    if best_a:
        return best_a
    if best_b and best_b_added:
        return best_b + best_b_added
    return best_b or best_b_added


def extract_company_data(corp_name, year="2025", report_type="사업보고서", month="12"):
    """한 회사의 보고서에서 DATA 시트에 필요한 모든 값 추출.
    month: 보고서 결산월 ("03"=1분기, "06"=반기, "09"=3분기, "12"=사업보고서)
    """
    xbrl_path = find_xbrl_file(corp_name, report_type, year)
    if not xbrl_path:
        print(f"  ERROR: XBRL not found for {corp_name} ({report_type} {year})")
        return None

    print(f"  파싱: {xbrl_path}")
    result = {}

    # 분기 보고서는 결산일이 다름 (3, 6, 9월말)
    ref_date = f"{year}-{month}-{'31' if month == '12' else '30' if month in ('06','09') else '31'}"
    # 실제 월말일 정확히
    import calendar
    last_day = calendar.monthrange(int(year), int(month))[1]
    ref_date = f"{year}-{month}-{last_day:02d}"
    start_date = f"{year}-01-01"  # PL 시작일 (항상 연초)

    # 단순 context (BS/PL) — 분기는 해당 분기말/분기 기간 기준
    simple = parse_simple_contexts(xbrl_path, year, ref_date=ref_date, month=month)
    bs = simple["bs"]
    pl = simple["pl"]
    is_annual = (month == "12")  # 사업보고서 여부

    def _get_bs(account_id):
        v = bs.get(account_id)
        if v is not None:
            return v
        alt = account_id.replace("dart_", "ifrs-full_") if account_id.startswith("dart_") else account_id.replace("ifrs-full_", "dart_")
        return bs.get(alt)

    # BS 매핑
    for col, account_id in BS_MAPPING.items():
        val = _get_bs(account_id)
        if val is not None:
            result[col] = val / 1e6  # 원 → 백만원

    # Col13 fallback: 참여특성 유/무 합산 (KB생명 패턴)
    if 13 not in result:
        v1 = _get_bs("dart_InsuranceContractLiabilitiesWithParticipationFeaturesOfLiabilitiesAbstract") or 0
        v2 = _get_bs("dart_InsuranceContractLiabilitiesWithoutParticipationFeaturesOfLiabilitiesAbstract") or 0
        if v1 + v2 > 0:
            result[13] = (v1 + v2) / 1e6

    # Col6 추가 fallback (신한라이프 패턴: SecuritiesAtFairValueThroughOCI)
    if 6 not in result:
        v = bs.get("dart_SecuritiesAtFairValueThroughOtherComprehensiveIncome")
        if v is not None:
            result[6] = v / 1e6

    # Col8 추가 fallback (InvestmentAccountedForUsingEquityMethod - 신한라이프 패턴)
    if 8 not in result:
        v = bs.get("ifrs-full_InvestmentAccountedForUsingEquityMethod")
        if v is not None:
            result[8] = v / 1e6

    # PL 매핑
    for col, account_id in PL_MAPPING.items():
        val = pl.get(account_id)
        if val is not None:
            result[col] = val / 1e6
        else:
            # ifrs-full과 dart prefix 둘 다 시도
            alt_id = account_id.replace("dart_", "ifrs-full_") if account_id.startswith("dart_") else account_id.replace("ifrs-full_", "dart_")
            val = pl.get(alt_id)
            if val is not None:
                result[col] = val / 1e6

    def _get_pl(account_id):
        """PL에서 dart_/ifrs-full_ 둘 다 시도"""
        v = pl.get(account_id)
        if v is not None:
            return v
        alt = account_id.replace("dart_", "ifrs-full_") if account_id.startswith("dart_") else account_id.replace("ifrs-full_", "dart_")
        return pl.get(alt, 0)

    # 계산 항목
    for col, (id1, id2, op) in CALC_MAPPING.items():
        if op == "direct":
            val = _get_pl(id1)
            if val:
                result[col] = val / 1e6
        elif op == "subtract":
            v1 = _get_pl(id1)
            v2 = _get_pl(id2)
            if v1 or v2:
                result[col] = (v1 - v2) / 1e6

    # Col56: 2-dim (Separate + ReportedAmountMember) context 우선
    fvoci_v = parse_fvoci_oci(xbrl_path, year, ref_date=ref_date, start_date=start_date)
    if fvoci_v:
        result[56] = fvoci_v / 1e6

    # OCI 매핑 (태그 후보 리스트) - Col56은 이미 처리됨
    for col, candidates in OCI_MAPPING.items():
        if col in result:
            continue
        for account_id in candidates:
            val = _get_pl(account_id)
            if val:
                result[col] = val / 1e6
                break

    # OCI 합산 매핑
    for col, (id1, id2) in OCI_SUM_MAPPING.items():
        v1 = _get_pl(id1)
        v2 = _get_pl(id2)
        if v1 or v2:
            result[col] = (v1 + v2) / 1e6

    # 전기말 BS 매핑
    prior_bs = simple["prior_bs"]
    prior_hybrid = 0.0
    for col, account_id in PRIOR_BS_MAPPING.items():
        if isinstance(col, str):
            # Special keys (e.g. "_prior_hybrid")
            v = prior_bs.get(account_id)
            if v is None:
                alt = account_id.replace("dart_", "ifrs-full_") if account_id.startswith("dart_") else account_id.replace("ifrs-full_", "dart_")
                v = prior_bs.get(alt)
            if v is not None and col == "_prior_hybrid":
                prior_hybrid = v
            continue
        val = prior_bs.get(account_id)
        if val is None:
            alt = account_id.replace("dart_", "ifrs-full_") if account_id.startswith("dart_") else account_id.replace("ifrs-full_", "dart_")
            val = prior_bs.get(alt)
        if val is not None:
            result[col] = val / 1e6

    # Col7 = AC 자산 (상각후원가 금융자산)
    ac_fin = _get_bs("ifrs-full_FinancialAssetsAtAmortisedCost") or 0
    loans_ac = _get_bs("ifrs-full_LoansAtAmortisedCost") or 0
    sec_ac = _get_bs("ifrs-full_SecuritiesAtAmortisedCost") or 0
    other_ac = (_get_bs("ifrs-full_OtherFinancialAssetsAtAmortisedCost") or
                _get_bs("dart_OtherFinancialAssetsAtAmortisedCost") or 0)

    if ac_fin > 0:
        if loans_ac > ac_fin:
            # loans가 ac_fin보다 크면 loans+sec+other로 합산
            result[7] = (loans_ac + sec_ac + other_ac) / 1e6
        elif loans_ac == 0 and sec_ac == 0 and other_ac > 0:
            # 메리츠화재 패턴: FinancialAssets + OtherFinancialAssets 별도
            result[7] = (ac_fin + other_ac) / 1e6
        else:
            result[7] = ac_fin / 1e6
    elif sec_ac > 0:
        if sec_ac >= loans_ac:
            result[7] = sec_ac / 1e6  # 신한라이프: 대출채권(보험계약대출) 제외
        else:
            result[7] = (loans_ac + sec_ac + other_ac) / 1e6
    elif loans_ac > 0:
        result[7] = (loans_ac + other_ac) / 1e6

    # Col14 = 자본(신종자본증권 제외) = Col15 - Col16
    if 15 in result:
        result[14] = result[15] - result.get(16, 0)

    # Col19 = 기초자본 (전기말 Equity) → overwrite PL_MAPPING value
    if 19 in result and result[19] == 0:
        # PL에서 못 찾은 경우만 prior_bs 사용
        pass
    # Col20 = 기초자본(신종자본증권 제외) = Col19 - 전기말 HybridBonds
    if 19 in result:
        result[20] = result[19] - prior_hybrid / 1e6

    # Col21: 자산운용률 총자산 (일부 회사는 총자산=운용자산)
    COL21_EQUALS_COL4 = {"한화생명", "삼성화재", "현대해상", "메리츠화재", "KB손해보험", "DB손해보험"}
    if 21 not in result and 4 in result and corp_name in COL21_EQUALS_COL4:
        result[21] = result[4]

    # 비율 계산 (Col4 총자산 기준)
    denom = result.get(4, 0)
    if denom > 0:
        if 5 in result:
            result[9] = result[5] / denom  # FVPL 비율
        if 6 in result:
            result[10] = result[6] / denom  # FVOCI 비율
        if 7 in result:
            result[11] = result[7] / denom  # AC 비율

    # 보험계약부채 구성요소 (CSM/BEL/RA)
    components = parse_insurance_components(xbrl_path, year, ref_date=ref_date)
    if components["CSM"] == 0 and components["BEL"] == 0:
        if is_annual and corp_name in DOC_FALLBACK_CORP_CODES:
            # XBRL에 component axis 없는 회사: 다음연도 1Q document.xml fallback (사업보고서)
            col13_val = result.get(13, 0)
            if col13_val > 0:
                components = parse_insurance_components_from_doc(corp_name, col13_val, year)
        elif not is_annual:
            # 분기/반기 보고서: 해당 연도 1Q doc.xml에서 당기말 잔액 테이블 파싱
            doc_path = _get_q1_doc_path(corp_name, year)
            if doc_path:
                bel_q, ra_q, csm_q = _extract_components_from_doc_q1(doc_path)
                if bel_q is not None:
                    components = {"BEL": bel_q * 1e6, "RA": (ra_q or 0) * 1e6, "CSM": csm_q * 1e6}
                    print(f"    doc_q1 BEL/RA/CSM: {bel_q:.0f}/{ra_q:.0f}/{csm_q:.0f} 백만원")
    if components["CSM"] > 0:
        result[75] = components["CSM"] / 1e6
    if components["BEL"] > 0:
        result[73] = components["BEL"] / 1e6
    if components["RA"] > 0:
        result[74] = components["RA"] / 1e6

    # 신계약 CSM
    new_csm = parse_new_csm(xbrl_path, year, ref_date=ref_date, start_date=start_date)
    if new_csm is not None:
        result[76] = new_csm / 1e6

    # 해약환급금준비금
    surr = parse_surrender_reserve(xbrl_path, year, ref_date=ref_date, start_date=start_date)
    if surr is not None:
        result[99] = surr / 1e6
    # fallback: doc.xml에서 '해약환급금준비금 잔액' 테이블 파싱
    if 99 not in result and is_annual:
        doc_path = _get_annual_doc(corp_name, year)
        if doc_path:
            try:
                with open(doc_path, encoding='utf-8') as _f:
                    _content = _f.read()
                for _kw in ['해약환급금준비금 잔액', '해약환급금준비금잔액']:
                    _idx = _content.find(_kw)
                    if _idx >= 0:
                        _region = _content[_idx:_idx+500]
                        _clean = re.sub(r'<[^>]+>', '|', _region)
                        _cells = [c.strip() for c in re.split(r'\|+', _clean) if c.strip()]
                        for _v in _cells[1:5]:
                            try:
                                _n = float(_v.replace(',', ''))
                                if _n > 100000:
                                    result[99] = _n
                                    break
                            except: pass
                        if 99 in result:
                            break
            except: pass

    # 추가 PL 항목 (중복 컬럼)
    for col, account_id in EXTRA_PL_MAPPING.items():
        if col not in result:
            val = _get_pl(account_id)
            if val:
                result[col] = val / 1e6

    # Col40: 기타투자손익 = FeeIncome + RentalIncome + OtherIncomeInvestment - OtherExpenseInvestment
    def _get_pl_multi(candidates):
        for aid in candidates:
            v = _get_pl(aid)
            if v:
                return v
        return 0

    def _get_pl_by_keyword(keywords):
        """PL에서 tag 이름에 keyword가 포함된 값들 합산"""
        total = 0.0
        for k, v in pl.items():
            for kw in keywords:
                if kw in k:
                    # 단순 OperatingExpenseInvestment(전체) 태그 제외
                    if k in ("dart_OperatingExpenseInvestment", "ifrs-full_OperatingExpenseInvestment"):
                        continue
                    # 이익/수익/관련자 거래 태그 제외
                    if any(x in k for x in ("Gain", "Income", "Revenue", "RelatedParty",
                                             "DisclosureOfTransactions")):
                        continue
                    # udf_ 계열 OfOperatingExpenseInvestment는 포함 (DB손보 패턴)
                    total += v
                    break
            # udf_ 계열 OfOperatingExpenseInvestment는 OtherOperatingExpenseInvestment와 함께 존재할 때만 포함
            # (동일 회사 내에서 중복 방지)
        return total

    fee = _get_pl_multi(COL40_FEE)
    rent = _get_pl_multi(COL40_RENT)
    other_inc = _get_pl_multi(COL40_OTHER_INC)
    other_exp = _get_pl_by_keyword(COL40_OTHER_EXP_KEYWORDS)
    col40_val = fee + rent + other_inc - other_exp
    if col40_val:
        result[40] = col40_val / 1e6

    # Col66: 결산배당 (음수, 이익잉여금에서 유출)
    for tag in DIVIDEND_TAGS:
        v = _get_pl(tag)
        if v:
            result[66] = -abs(v) / 1e6  # 항상 음수
            break

    # Col70: 신종자본증권 배당 (음수)
    for tag in HYBRID_DIVIDEND_TAGS:
        v = _get_pl(tag)
        if v:
            result[70] = -abs(v) / 1e6
            break

    # Col67/69: 자기주식 소각, 기타자본변동 (RetainedEarnings component 2-dim context)
    eq_changes = parse_equity_changes(xbrl_path, year, ref_date=ref_date, start_date=start_date)
    if 67 not in result and eq_changes.get(67):
        result[67] = eq_changes[67] / 1e6
    if 69 not in result and eq_changes.get(69):
        result[69] = eq_changes[69] / 1e6

    # Col38: 금융손익 = Col35 - Col36 - Col37 + Col39 - Col40 (Col37 없으면 0)
    if all(c in result for c in [35, 36, 39, 40]):
        result[38] = result[35] - result[36] - result.get(37, 0) + result[39] - result[40]

    # Col42/Col50 fallback: 영업외손익 = 세전이익 - 영업이익
    if 42 not in result and 43 in result and 41 in result:
        result[42] = result[43] - result[41]
        result[50] = result[42]

    # Col51: 영업손익 합계 = Col48 + Col49 + Col50
    if any(c in result for c in [48, 49, 50]):
        result[51] = result.get(48, 0) + result.get(49, 0) + result.get(50, 0)

    # OCI 누계 항목 (당기말/전기말 BS 3-dim context)
    oci_accum = parse_oci_accumulated(xbrl_path, year, ref_date=ref_date)
    oci_cur = oci_accum["cur"]
    oci_prior = oci_accum["prior"]
    for col, keywords in OCI_ACCUM_MAPPING.items():
        # Col29 재평가잉여금은 전기말
        src = oci_prior if col == 29 else oci_cur
        total = 0.0
        for member, v in src.items():
            if any(kw in member for kw in keywords):
                total += v
        if total != 0.0:
            result[col] = total / 1e6

    # 패턴C direct: CarryingAmountAxis→ReportedAmountMember (DB손보/현대해상)
    # Col25~30 중 패턴A/B로 못 채운 항목을 패턴C 결과로 보완
    oci_cur_direct = oci_accum.get("cur_direct", {})
    oci_prior_direct = oci_accum.get("prior_direct", {})
    for col in (25, 26, 27, 28, 30):
        if col not in result and col in oci_cur_direct:
            result[col] = oci_cur_direct[col] / 1e6
    # Col29 재평가잉여금: 전기말 direct 사용
    if 29 not in result and 29 in oci_prior_direct:
        result[29] = oci_prior_direct[29] / 1e6

    # Col24 = 기타포괄손익누계액 합계 (= Col17, AccumulatedOCI)
    if 17 in result:
        result[24] = result[17]

    # Col63: 기말 기타포괄손익누계액 (당기말 = Col17)
    if 17 in result:
        result[63] = result[17]

    # Col71: 이익잉여금 기말 (당기말 BS = Col18)
    if 18 in result:
        result[71] = result[18]

    # CSM 변동 항목 (XBRL 기반 — 분기/반기/사업보고서 모두 시도)
    csm_mov = parse_csm_movement(xbrl_path, year, ref_date=ref_date, start_date=start_date)
    if csm_mov["csm_amortization"]:
        result[77] = abs(csm_mov["csm_amortization"]) / 1e6
    if csm_mov["csm_adjustment"]:
        result[78] = csm_mov["csm_adjustment"] / 1e6
    if csm_mov["csm_finance"]:
        result[79] = csm_mov["csm_finance"] / 1e6
    if 76 not in result and csm_mov.get("new_csm"):
        result[76] = abs(csm_mov["new_csm"]) / 1e6

    prior_csm = parse_prior_csm(xbrl_path, year)
    if prior_csm:
        result[80] = prior_csm / 1e6

    # doc.xml CSM 변동 fallback — 사업보고서만 (분기 doc.xml은 별도 작업)
    if is_annual:
        if 77 not in result and 75 in result and corp_name in DOC_CSM_MOVEMENT_COMPANIES:
            doc_csm = parse_csm_movement_from_doc(corp_name, result[75], year)
            for col, val in doc_csm.items():
                if col not in result and val != 0:
                    result[col] = val

    # Col88: CSM 합계 (= Col75)
    if 75 in result:
        result[88] = result[75]

    if is_annual:
        # FISIS로 Col16/22/74/75/99 채움
        # Col99(해약환급금)/Col22(운용자산): FISIS가 XBRL보다 신뢰 → 무조건 덮어씀
        # Col16/74/75: XBRL 없을 때만
        _fisis_always = {22, 99}
        _fisis_fallback = {16, 74, 75}
        _fisis_needed = _fisis_always | {c for c in _fisis_fallback if c not in result or result[c] == 0}
        base_month = f"{year}{month}"
        try:
            from dart_api import fetch_fisis_balance
            _fbal = fetch_fisis_balance(corp_name, base_month)
            for col, val in _fbal.items():
                if col in _fisis_needed and val and val != 0:
                    result[col] = val
        except Exception:
            pass

        # 사업보고서 doc.xml: K-ICS, CSM기간별, 감응도, OCI 세부
        csm_total_for_doc = result.get(88, result.get(75, 0))
        oci_total_for_doc = result.get(17, result.get(24, None))
        total_assets_for_doc = result.get(4, 0)
        doc_data = parse_annual_doc_data(corp_name, csm_total=csm_total_for_doc,
                                         oci_total=oci_total_for_doc, year=year,
                                         total_assets=total_assets_for_doc)
        for col, val in doc_data.items():
            if col not in result or result[col] == 0:
                result[col] = val
            elif col in (25, 26, 27, 28, 29, 30) and result[col] != 0:
                pass
    else:
        # 분기/반기 보고서: FISIS API로 K-ICS 및 재무 항목 조회
        base_month = f"{year}{month}"
        try:
            from dart_api import fetch_fisis_kics, fetch_fisis_balance
            fisis_kics = fetch_fisis_kics(corp_name, base_month)
            for col, val in fisis_kics.items():
                if col not in result or result[col] == 0:
                    result[col] = val
            fisis_bal = fetch_fisis_balance(corp_name, base_month)
            # Col74/75(RA/CSM)/22(운용자산)/99(해약환급금): FISIS가 더 신뢰 → 무조건 덮어씀
            # Col16: XBRL 없을 때만
            _fisis_always_q = {22, 74, 75, 99}
            for col, val in fisis_bal.items():
                if col in _fisis_always_q:
                    if val and val != 0:
                        result[col] = val
                elif col not in result or result[col] == 0:
                    result[col] = val
        except Exception as _e:
            print(f"    FISIS 조회 실패 ({corp_name} {base_month}): {_e}")

    # Col99 부호 보정: 해약환급금준비금은 항상 양수
    if 99 in result and result[99] < 0:
        result[99] = abs(result[99])

    # 계리적 가정 변경 민감도 분석 (Col146~158) — XBRL 기반, 분기/반기/사업보고서 모두 시도
    actuarial = parse_actuarial_sensitivity(xbrl_path, year, ref_date=ref_date, start_date=start_date)
    for (col, comp), v in actuarial.items():
        result[col] = v / 1e6

    # Col158 fallback: 손실요소 (여러 패턴)
    if 158 not in result:
        tree_tmp = ET.parse(xbrl_path)
        root_tmp = tree_tmp.getroot()
        csm_members_set = {
            "ContractualServiceMarginMember",
            "ContractualServiceMarginRelatedToContractsThatExistedAtTransitionDateToWhichModifiedRetrospectiveApproachHasBeenAppliedMember",
            "ContractualServiceMarginNotRelatedToContractsThatExistedAtTransitionDateToWhichModifiedRetrospectiveApproachOrFairValueApproachHasBeenAppliedMember",
            "ContractualServiceMarginRelatedToContractsThatExistedAtTransitionDateToWhichFairValueApproachHasBeenAppliedMember",
        }
        # 패턴A: DoNotAdjustContractualServiceMargin in CSM context
        # 패턴B: AdjustContractualServiceMargin in CSM+LossComponentMember context
        csm_ctxs_a = set()  # CSM context (패턴A)
        csm_loss_ctxs_b = set()  # CSM + LossComponentMember context (패턴B)
        for ctx in root_tmp.findall(f"{{{NS_XBRLI}}}context"):
            period = ctx.find(f"{{{NS_XBRLI}}}period")
            start = period.find(f"{{{NS_XBRLI}}}startDate")
            end = period.find(f"{{{NS_XBRLI}}}endDate")
            if start is None or start.text != start_date:
                continue
            if end is None or end.text != ref_date:
                continue
            members = ctx.findall(f".//{{{NS_XBRLDI}}}explicitMember")
            dims = {}
            for m in members:
                dim_name = m.get("dimension", "").split(":")[-1]
                val = (m.text or "").split(":")[-1]
                dims[dim_name] = val
            if "SeparateMember" != dims.get("ConsolidatedAndSeparateFinancialStatementsAxis", ""):
                continue
            comp = dims.get("InsuranceContractsByComponentsAxis", "")
            disagg = dims.get("DisaggregationOfInsuranceContractsAxis", "")
            has_csm = comp in csm_members_set or disagg in csm_members_set
            has_loss = any("LossComponent" in v for v in dims.values())
            ctx_id = ctx.get("id", "")
            if has_csm and has_loss:
                csm_loss_ctxs_b.add(ctx_id)
            elif has_csm:
                csm_ctxs_a.add(ctx_id)

        # 패턴A: DoNotAdjust in CSM context
        for elem in root_tmp:
            ctx_ref = elem.get("contextRef", "")
            if ctx_ref not in csm_ctxs_a or not elem.text:
                continue
            tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
            if "DoNotAdjustContractualServiceMargin" not in tag:
                continue
            try:
                v = float(elem.text.strip())
            except ValueError:
                continue
            if v != 0:
                result[158] = v / 1e6
                break

        # 패턴B: AdjustContractualServiceMargin in CSM+LossComponent context
        if 158 not in result:
            for elem in root_tmp:
                ctx_ref = elem.get("contextRef", "")
                if ctx_ref not in csm_loss_ctxs_b or not elem.text:
                    continue
                tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
                if "AdjustContractualServiceMargin" not in tag:
                    continue
                try:
                    v = float(elem.text.strip())
                except ValueError:
                    continue
                if v != 0:
                    result[158] = v / 1e6
                    break

    # 계산값 (다른 컬럼 값들의 조합)
    # 체크열(Col31/45/46/47/64/72/82/89/90/159)은 엑셀에서 수식으로 계산 — 스크립트에서 채우지 않음

    return result


def _period_to_report(period_code):
    """기간코드(예: 2512) → (year, month, report_type) 반환"""
    pc = str(period_code)
    yy = pc[:2]
    mm = pc[2:]
    year = f"20{yy}"
    month = mm
    reprt_map = {
        "03": "1분기보고서",
        "06": "반기보고서",
        "09": "3분기보고서",
        "12": "사업보고서",
    }
    report_type = reprt_map.get(mm, "사업보고서")
    return year, month, report_type


def main(year=None, period_code=None):
    """
    year: 사업연도 (예: "2025"). 미지정 시 직전 연도 자동 사용.
    period_code: Excel Col2 기간코드 (예: 2512). 미지정 시 year에서 자동 계산 (YY12).
                 'all'을 지정하면 Excel 빈칸 파일의 모든 기간 처리.
    """
    import datetime
    excel_blank = "DATA_작업_빈칸.xlsx"

    # period_code='all'이면 Excel에서 기간 목록 읽어서 전체 처리
    if str(period_code).lower() == "all" if period_code else False:
        try:
            import openpyxl as _xl
            wb = _xl.load_workbook(excel_blank, data_only=True)
            ws = wb["DATA"]
            periods = sorted(set(
                str(ws.cell(row=r, column=2).value)
                for r in range(6, ws.max_row + 1)
                if ws.cell(row=r, column=2).value is not None
            ))
        except Exception:
            periods = []
        for p in periods:
            print(f"\n{'='*60}")
            print(f"기간코드 {p} 처리 중...")
            main(period_code=int(p))
        return

    if year is None and period_code is None:
        year = str(datetime.date.today().year - 1)
        period_code = int(year[2:] + "12")
    elif period_code is not None and year is None:
        year, month, _ = _period_to_report(period_code)
    elif year is not None and period_code is None:
        period_code = int(year[2:] + "12")

    year, month, report_type = _period_to_report(period_code)

    print("=" * 60)
    print(f"{report_type} XBRL → DATA 시트 매핑 ({year}년 {month}월, 기간코드 {period_code})")
    print("=" * 60)

    all_results = {}
    for comp in COMPANIES:
        name = comp["name"]
        short = comp["short"]
        print(f"\n[{name}] ({short})")
        data = extract_company_data(name, year, report_type=report_type, month=month)
        if data:
            key = f"{period_code}{short}"
            data[2] = period_code
            all_results[key] = data
            print(f"  추출 항목: {len(data)}개")
            if 4 in data:
                print(f"  자산: {data[4]:,.2f}M")
            if 32 in data:
                print(f"  보험손익: {data[32]:,.2f}M")
            if 44 in data:
                print(f"  당기순이익: {data[44]:,.2f}M")

    # CSV 출력
    output_path = f"data/data_sheet_{period_code}.csv"
    os.makedirs("data", exist_ok=True)

    all_cols = sorted(set(col for data in all_results.values() for col in data.keys()))
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["key"] + [f"Col{c}" for c in all_cols])
        for key, data in all_results.items():
            row = [key] + [round(data[c], 6) if isinstance(data.get(c), float) else data.get(c, "") for c in all_cols]
            writer.writerow(row)

    print(f"\n저장: {output_path}")
    print(f"회사 수: {len(all_results)}, 열 수: {len(all_cols)}")

    # Excel 빈칸 파일에 결과 채우기
    excel_blank = "DATA_작업_빈칸.xlsx"
    if os.path.exists(excel_blank):
        try:
            import openpyxl as _xl
            wb = _xl.load_workbook(excel_blank, keep_vba=False)  # 수식 유지를 위해 data_only=False
            ws = wb["DATA"]

            # row1: 컬럼번호 → Excel열 역매핑
            col_to_excel = {}
            for ec in range(1, ws.max_column + 1):
                v = ws.cell(row=1, column=ec).value
                if v is not None:
                    try:
                        col_to_excel[int(v)] = ec
                    except (ValueError, TypeError):
                        pass

            # 기간코드(col2) + 회사명(col3) → 데이터행 매핑
            short_to_row = {}
            for r in range(6, ws.max_row + 1):
                v2 = ws.cell(row=r, column=2).value
                v3 = ws.cell(row=r, column=3).value
                if str(v2) == str(period_code) and v3:
                    short_to_row[str(v3)] = r

            from openpyxl.utils import get_column_letter

            def _addr(col_num, row):
                """컬럼번호 → 셀주소 (예: 35, 12 → 'AI12')"""
                ec = col_to_excel.get(col_num)
                return f"{get_column_letter(ec)}{row}" if ec else None

            # 계산 항목별 수식 생성 함수
            def _formula(col_num, row):
                a = _addr
                r = row
                if col_num == 14:   # 자본(신종제외) = Col15 - Col16
                    c15, c16 = a(15,r), a(16,r)
                    if c15 and c16: return f"={c15}-{c16}"
                elif col_num == 9:   # FVPL 비중 = Col5/Col4
                    c5, c4 = a(5,r), a(4,r)
                    if c5 and c4: return f"={c5}/{c4}"
                elif col_num == 10:  # FVOCI 비중 = Col6/Col4
                    c6, c4 = a(6,r), a(4,r)
                    if c6 and c4: return f"={c6}/{c4}"
                elif col_num == 11:  # AC 비중 = Col7/Col4
                    c7, c4 = a(7,r), a(4,r)
                    if c7 and c4: return f"={c7}/{c4}"
                elif col_num == 20:  # 기시자본(신종제외) = Col19 - (Col19*Col16/Col15 근사)
                    pass  # 전기말 신종자본증권이 별도값이라 수식 불가
                elif col_num == 24:  # OCI합계 = Col17
                    c17 = a(17,r)
                    if c17: return f"={c17}"
                elif col_num == 31:  # OCI CHECK = Col24-(Col25+26+27+28+29+30)
                    cols = [a(c,r) for c in [24,25,26,27,28,29,30]]
                    if all(cols): return f"={cols[0]}-({cols[1]}+{cols[2]}+{cols[3]}+{cols[4]}+{cols[5]}+{cols[6]})"
                elif col_num == 38:  # 금융손익 = Col35-Col36-Col37+Col39-Col40
                    cs = [a(c,r) for c in [35,36,37,39,40]]
                    if all(cs): return f"={cs[0]}-{cs[1]}-{cs[2]}+{cs[3]}-{cs[4]}"
                elif col_num == 45:  # CHECK1 = Col41+Col42-Col43
                    cs = [a(c,r) for c in [41,42,43]]
                    if all(cs): return f"={cs[0]}+{cs[1]}-{cs[2]}"
                elif col_num == 46:  # CHECK2 = Col36+Col37+Col38-Col39+Col40-Col35
                    cs = [a(c,r) for c in [36,37,38,39,40,35]]
                    if all(cs): return f"={cs[0]}+{cs[1]}+{cs[2]}-{cs[3]}+{cs[4]}-{cs[5]}"
                elif col_num == 47:  # CHECK3 = Col32+Col35-Col41
                    cs = [a(c,r) for c in [32,35,41]]
                    if all(cs): return f"={cs[0]}+{cs[1]}-{cs[2]}"
                elif col_num == 48:  # = Col32
                    c32 = a(32,r)
                    if c32: return f"={c32}"
                elif col_num == 49:  # = Col35
                    c35 = a(35,r)
                    if c35: return f"={c35}"
                elif col_num == 50:  # = Col42
                    c42 = a(42,r)
                    if c42: return f"={c42}"
                elif col_num == 51:  # = Col48+Col49+Col50
                    cs = [a(c,r) for c in [48,49,50]]
                    if all(cs): return f"={cs[0]}+{cs[1]}+{cs[2]}"
                elif col_num == 63:  # = Col17
                    c17 = a(17,r)
                    if c17: return f"={c17}"
                elif col_num == 64:  # OCI 자본변동 CHECK
                    cs = [a(c,r) for c in [63,55,56,57,58,59,60,61,62]]
                    if cs[0] and cs[1]: return f"={cs[0]}-{cs[1]}-SUM({get_column_letter(col_to_excel.get(56,1))}{r}:{get_column_letter(col_to_excel.get(62,1))}{r})"
                elif col_num == 68:  # = Col44
                    c44 = a(44,r)
                    if c44: return f"={c44}"
                elif col_num == 71:  # = Col18
                    c18 = a(18,r)
                    if c18: return f"={c18}"
                elif col_num == 82:  # CSM CHECK = Col80+76-77+78+79-75
                    cs = [a(c,r) for c in [80,76,77,78,79,75]]
                    if all(cs): return f"={cs[0]}+{cs[1]}-{cs[2]}+{cs[3]}+{cs[4]}-{cs[5]}"
                elif col_num == 88:  # = Col75
                    c75 = a(75,r)
                    if c75: return f"={c75}"
                elif col_num == 89:  # CSM기간별 CHECK = Col88-SUM(83:87)
                    c88 = a(88,r)
                    c83 = col_to_excel.get(83); c87 = col_to_excel.get(87)
                    if c88 and c83 and c87: return f"={c88}-SUM({get_column_letter(c83)}{r}:{get_column_letter(c87)}{r})"
                elif col_num == 103:  # = Col101-Col102
                    cs = [a(c,r) for c in [101,102]]
                    if all(cs): return f"={cs[0]}-{cs[1]}"
                elif col_num == 104:  # = Col102/Col101
                    cs = [a(c,r) for c in [102,101]]
                    if all(cs): return f"={cs[0]}/{cs[1]}"
                return None

            # 수식으로 넣을 컬럼 번호 집합
            FORMULA_COLS = {9,10,11,14,24,31,38,45,46,47,48,49,50,51,63,64,68,71,82,88,89,103,104}

            filled = 0
            for key, data in all_results.items():
                short = key[len(str(period_code)):]
                row_idx = short_to_row.get(short)
                if row_idx is None:
                    continue
                for col_num, val in data.items():
                    if col_num == 2:
                        continue
                    ec = col_to_excel.get(col_num)
                    if ec is None:
                        continue
                    cell = ws.cell(row=row_idx, column=ec)
                    if cell.data_type == "f":
                        continue

                    # 계산 항목은 수식으로 넣기
                    if col_num in FORMULA_COLS:
                        formula = _formula(col_num, row_idx)
                        if formula:
                            cell.value = formula
                            filled += 1
                            continue

                    # 부동소수점 반올림 (소수점 6자리)
                    if isinstance(val, float):
                        val = round(val, 6)
                    cell.value = val
                    filled += 1

            wb.save(excel_blank)
            print(f"Excel 채우기 완료: {excel_blank} ({filled}셀)")
        except Exception as e:
            print(f"Excel 쓰기 실패: {e}")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="동업사 공시비교 DATA 시트 자동 추출")
    parser.add_argument("--year", default=None,
                        help="사업연도 (예: 2025). 미지정 시 직전 연도 자동 사용.")
    parser.add_argument("--period", default=None,
                        help="기간코드 (예: 2512) 또는 'all'(Excel 내 전체 기간). 미지정 시 year에서 자동 계산.")
    args = parser.parse_args()
    period = args.period if args.period == "all" else (int(args.period) if args.period else None)
    main(year=args.year, period_code=period)
